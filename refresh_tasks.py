from __future__ import annotations

import os
import pickle
import socket
import uuid
import argparse
import gc
import json
import logging
import math
import sqlite3
import sys
import traceback
import time as time_module
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any
from datetime import datetime, timedelta, date, time as dt_time

import openpyxl
import xlwings as xw

LOG_PATH = r"L:\LRC\common_data\ФЛЮИДЫ\ГТИ\sqlite-excel\task_refresh.log"

PROJECT_DIR = Path(__file__).resolve().parent
ERROR_LOG_PATH = PROJECT_DIR / "refresh_tasks_errors.log"

# Общий сетевой кэш внешнего журнала.
# Поскольку refresh_tasks.py находится на сетевом диске,
# этот каталог будет общим для всех пользователей.
TASK_CACHE_VERSION = 1

TASK_CACHE_PATH = (
    PROJECT_DIR
    / "cache"
    / "task_journal_cache.pkl"
)

TASK_CACHE_LOCK_PATH = (
    PROJECT_DIR
    / "cache"
    / "task_journal_cache.lock"
)

# Lock старше этого времени считается брошенным.
TASK_CACHE_LOCK_STALE_SECONDS = 180.0

# Сколько ждать, пока другой пользователь обновит кэш.
TASK_CACHE_WAIT_TIMEOUT_SECONDS = 180.0

TASK_CACHE_POLL_INTERVAL_SECONDS = 0.25

PROGRESS_FILE_PATH: Path | None = None

def set_progress_file(path: str | None) -> None:
    global PROGRESS_FILE_PATH

    if not path:
        PROGRESS_FILE_PATH = None
        return

    PROGRESS_FILE_PATH = Path(path)
    PROGRESS_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
    PROGRESS_FILE_PATH.write_text("", encoding="utf-8")


def progress(message: str) -> None:
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n"

    if PROGRESS_FILE_PATH is not None:
        for _ in range(5):
            try:
                with PROGRESS_FILE_PATH.open("a", encoding="utf-8") as f:
                    f.write(line)
                return
            except PermissionError:
                time_module.sleep(0.05)
            except OSError:
                time_module.sleep(0.05)

        # Прогресс не критичен. Не даём ему ломать основную синхронизацию.
        return

    try:
        print(line, file=sys.stderr, flush=True)
    except Exception:
        pass

def write_error_log(
    *,
    message: str,
    error_type: str,
    traceback_text: str,
) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with ERROR_LOG_PATH.open("a", encoding="utf-8") as f:
        f.write("\n" + "=" * 100 + "\n")
        f.write(f"datetime: {now}\n")
        f.write(f"error_type: {error_type}\n")
        f.write(f"message: {message}\n")
        f.write("-" * 100 + "\n")
        f.write(traceback_text)
        f.write("\n")

# =============================================================================
# Настройки колонок
# =============================================================================

PARENT_TASK_ID_COL = "Номер задания"
PARENT_SAMPLE_CODE_COL = "Код проекта"
PARENT_TASK_TYPE_COL = "Задание"
PARENT_STATUS_COL = "Статус"
PARENT_ROW_ID_COL = "ID"
PARENT_DATETIME_COL = "Дата и время"
CHILD_TASK_ID_COL = "Номер ГТИ"


# =============================================================================
# Отчёт
# =============================================================================

TASK_LOGGER = logging.getLogger("refresh_tasks")
TASK_LOGGER.setLevel(logging.INFO)
TASK_LOGGER.propagate = False

if not TASK_LOGGER.handlers:
    task_log_handler = logging.FileHandler(
        LOG_PATH,
        mode="a",
        encoding="utf-8",
    )

    task_log_handler.setFormatter(
        logging.Formatter(
            "%(asctime)s - %(levelname)s - %(message)s"
        )
    )

    TASK_LOGGER.addHandler(task_log_handler)

def log_elapsed(
    stage_name: str,
    started_at: float,
    *,
    details: str = "",
) -> float:
    elapsed = time_module.perf_counter() - started_at

    if details:
        TASK_LOGGER.info(
            "%s — %.3f с; %s",
            stage_name,
            elapsed,
            details,
        )
    else:
        TASK_LOGGER.info(
            "%s — %.3f с",
            stage_name,
            elapsed,
        )

    return elapsed

@dataclass
class RefreshTasksReport:
    ok: bool
    message: str
    project: str
    parent_rows_read: int = 0
    parent_rows_written: int = 0
    child_rows_read: int = 0
    child_rows_written: int = 0
    sqlite_statuses_applied: int = 0
    external_statuses_kept: int = 0
    error_type: str | None = None
    traceback: str | None = None


# =============================================================================
# openpyxl: чтение внешней книги с заданиями
# =============================================================================

def read_excel_table_from_workbook(
        wb,
        *,
        sheet_name: str,
        table_name: str,
) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Читает умную таблицу из уже открытой через openpyxl книги.

    Открытием и закрытием книги эта функция не занимается.
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Во внешней книге нет листа '{sheet_name}'"
        )

    ws = wb[sheet_name]

    if table_name not in ws.tables:
        raise ValueError(
            f"На листе '{sheet_name}' во внешней книге "
            f"нет таблицы '{table_name}'"
        )

    table = ws.tables[table_name]
    table_range = ws[table.ref]

    rows_values = [
        [cell.value for cell in row]
        for row in table_range
    ]

    if not rows_values:
        return [], []

    headers = [
        str(value).strip() if value is not None else ""
        for value in rows_values[0]
    ]

    rows: list[dict[str, Any]] = []

    for raw_row in rows_values[1:]:
        row = {
            headers[i]: normalize_excel_value(raw_row[i])
            for i in range(len(headers))
        }

        if is_empty_row(row):
            continue

        rows.append(row)

    return headers, rows


def normalize_excel_value(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return int(value)

    return value


def is_empty_row(row: dict[str, Any]) -> bool:
    return all(str(v).strip() == "" for v in row.values())

def build_task_cache_metadata(
        *,
        source_workbook: str,
        source_parent_sheet: str,
        source_parent_table: str,
        source_child_sheet: str,
        source_child_table: str,
) -> dict[str, Any]:
    """
    Формирует описание состояния исходного Excel-файла.

    Если размер или время изменения файла поменялись,
    существующий кэш считается устаревшим.
    """
    source_path = Path(source_workbook)
    source_stat = source_path.stat()

    try:
        normalized_path = str(
            source_path.resolve(strict=True)
        ).casefold()
    except OSError:
        normalized_path = str(
            source_path.absolute()
        ).casefold()

    return {
        "cache_version": TASK_CACHE_VERSION,
        "source_path": normalized_path,
        "source_size": source_stat.st_size,
        "source_mtime_ns": source_stat.st_mtime_ns,
        "parent_sheet": source_parent_sheet,
        "parent_table": source_parent_table,
        "child_sheet": source_child_sheet,
        "child_table": source_child_table,
    }


def read_valid_task_cache(
        expected_metadata: dict[str, Any],
) -> tuple[
    list[str],
    list[dict[str, Any]],
    list[str],
    list[dict[str, Any]],
] | None:
    """
    Читает кэш только тогда, когда он соответствует
    текущему состоянию исходного Excel-файла.
    """
    try:
        with TASK_CACHE_PATH.open("rb") as cache_file:
            payload = pickle.load(cache_file)

    except FileNotFoundError:
        return None

    except Exception:
        TASK_LOGGER.warning(
            "Не удалось прочитать сетевой кэш; "
            "он будет создан заново",
            exc_info=True,
        )
        return None

    if not isinstance(payload, dict):
        return None

    if payload.get("metadata") != expected_metadata:
        return None

    required_keys = {
        "parent_headers",
        "parent_rows",
        "child_headers",
        "child_rows",
    }

    if not required_keys.issubset(payload):
        return None

    return (
        payload["parent_headers"],
        payload["parent_rows"],
        payload["child_headers"],
        payload["child_rows"],
    )


def is_task_cache_lock_stale() -> bool:
    """
    Проверяет, не остался ли lock после аварийного
    завершения другого процесса.
    """
    try:
        lock_stat = TASK_CACHE_LOCK_PATH.stat()
    except FileNotFoundError:
        return False

    lock_age = time_module.time() - lock_stat.st_mtime

    return lock_age > TASK_CACHE_LOCK_STALE_SECONDS


def try_acquire_task_cache_lock() -> bool:
    """
    Пытается атомарно создать lock-файл.

    Режим 'x' создаёт файл только при условии,
    что его ещё не существует.
    """
    TASK_CACHE_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    for _ in range(2):
        try:
            with TASK_CACHE_LOCK_PATH.open(
                "x",
                encoding="utf-8",
            ) as lock_file:
                lock_file.write(
                    f"computer={socket.gethostname()}\n"
                    f"pid={os.getpid()}\n"
                    f"created_at={datetime.now().isoformat()}\n"
                )

            return True

        except FileExistsError:
            if not is_task_cache_lock_stale():
                return False

            TASK_LOGGER.warning(
                "Обнаружен устаревший lock кэша; "
                "пытаюсь удалить: %s",
                TASK_CACHE_LOCK_PATH,
            )

            try:
                TASK_CACHE_LOCK_PATH.unlink()
            except FileNotFoundError:
                pass
            except OSError:
                return False

    return False


def release_task_cache_lock() -> None:
    try:
        TASK_CACHE_LOCK_PATH.unlink()
    except FileNotFoundError:
        pass
    except OSError:
        TASK_LOGGER.warning(
            "Не удалось удалить lock-файл кэша: %s",
            TASK_CACHE_LOCK_PATH,
            exc_info=True,
        )


def write_task_cache_atomic(
        *,
        metadata: dict[str, Any],
        parent_headers: list[str],
        parent_rows: list[dict[str, Any]],
        child_headers: list[str],
        child_rows: list[dict[str, Any]],
) -> None:
    """
    Записывает кэш во временный файл и затем атомарно
    заменяет основной кэш.

    Читатель не увидит частично записанный pickle.
    """
    TASK_CACHE_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    payload = {
        "metadata": metadata,
        "parent_headers": parent_headers,
        "parent_rows": parent_rows,
        "child_headers": child_headers,
        "child_rows": child_rows,
    }

    temp_path = TASK_CACHE_PATH.with_name(
        f"{TASK_CACHE_PATH.name}."
        f"{socket.gethostname()}."
        f"{os.getpid()}."
        f"{uuid.uuid4().hex}.tmp"
    )

    try:
        with temp_path.open("wb") as cache_file:
            pickle.dump(
                payload,
                cache_file,
                protocol=pickle.HIGHEST_PROTOCOL,
            )

            cache_file.flush()

            try:
                os.fsync(cache_file.fileno())
            except OSError:
                # На некоторых сетевых файловых системах
                # fsync может быть недоступен.
                pass

        # Если другой пользователь в этот момент читает
        # старый кэш, Windows может ненадолго блокировать замену.
        last_error: Exception | None = None

        for _ in range(30):
            try:
                os.replace(
                    temp_path,
                    TASK_CACHE_PATH,
                )
                return

            except PermissionError as exc:
                last_error = exc
                time_module.sleep(0.1)

        if last_error is not None:
            raise last_error

        raise RuntimeError(
            "Не удалось заменить сетевой файл кэша"
        )

    finally:
        try:
            temp_path.unlink()
        except FileNotFoundError:
            pass
        except OSError:
            pass


def read_external_task_tables_from_excel(
        *,
        source_workbook: str,
        source_parent_sheet: str,
        source_parent_table: str,
        source_child_sheet: str,
        source_child_table: str,
) -> tuple[
    list[str],
    list[dict[str, Any]],
    list[str],
    list[dict[str, Any]],
]:
    """
    Открывает внешний журнал один раз и читает обе таблицы.
    """
    progress(
        f"Открываю внешний журнал заданий: "
        f"{Path(source_workbook).name}"
    )

    stage_started_at = time_module.perf_counter()

    source_wb = openpyxl.load_workbook(
        source_workbook,
        read_only=False,
        data_only=True,
    )

    log_elapsed(
        "Открытие внешнего журнала заданий",
        stage_started_at,
        details=f"файл={Path(source_workbook).name}",
    )

    try:
        progress(
            f"Читаю внешнюю родительскую таблицу: "
            f"{source_parent_sheet}/{source_parent_table}"
        )

        stage_started_at = time_module.perf_counter()

        parent_headers, parent_rows = (
            read_excel_table_from_workbook(
                source_wb,
                sheet_name=source_parent_sheet,
                table_name=source_parent_table,
            )
        )

        log_elapsed(
            "Чтение внешней родительской таблицы",
            stage_started_at,
            details=(
                f"лист={source_parent_sheet}; "
                f"таблица={source_parent_table}; "
                f"строк={len(parent_rows)}"
            ),
        )

        progress(
            f"Читаю внешнюю дочернюю таблицу: "
            f"{source_child_sheet}/{source_child_table}"
        )

        stage_started_at = time_module.perf_counter()

        child_headers, child_rows = (
            read_excel_table_from_workbook(
                source_wb,
                sheet_name=source_child_sheet,
                table_name=source_child_table,
            )
        )

        log_elapsed(
            "Чтение внешней дочерней таблицы",
            stage_started_at,
            details=(
                f"лист={source_child_sheet}; "
                f"таблица={source_child_table}; "
                f"строк={len(child_rows)}"
            ),
        )

        return (
            parent_headers,
            parent_rows,
            child_headers,
            child_rows,
        )

    finally:
        stage_started_at = time_module.perf_counter()

        source_wb.close()

        log_elapsed(
            "Закрытие внешнего журнала заданий",
            stage_started_at,
        )


def load_external_task_tables_cached(
        *,
        source_workbook: str,
        source_parent_sheet: str,
        source_parent_table: str,
        source_child_sheet: str,
        source_child_table: str,
) -> tuple[
    list[str],
    list[dict[str, Any]],
    list[str],
    list[dict[str, Any]],
    str,
]:
    """
    Возвращает обе таблицы либо из общего сетевого кэша,
    либо из исходного Excel.

    Последний элемент результата:
    - 'network-cache'
    - 'excel-cache-refresh'
    - 'excel-cache-timeout'
    """
    expected_metadata = build_task_cache_metadata(
        source_workbook=source_workbook,
        source_parent_sheet=source_parent_sheet,
        source_parent_table=source_parent_table,
        source_child_sheet=source_child_sheet,
        source_child_table=source_child_table,
    )

    cache_started_at = time_module.perf_counter()

    cached_data = read_valid_task_cache(
        expected_metadata
    )

    if cached_data is not None:
        log_elapsed(
            "Чтение внешних таблиц из сетевого кэша",
            cache_started_at,
            details=(
                f"кэш={TASK_CACHE_PATH}; "
                f"parent={len(cached_data[1])}; "
                f"child={len(cached_data[3])}"
            ),
        )

        return (*cached_data, "network-cache")

    TASK_LOGGER.info(
        "Сетевой кэш отсутствует или устарел: %s",
        TASK_CACHE_PATH,
    )

    wait_started_at = time_module.perf_counter()
    wait_deadline = (
        wait_started_at
        + TASK_CACHE_WAIT_TIMEOUT_SECONDS
    )

    while True:
        if try_acquire_task_cache_lock():
            try:
                # Другой процесс мог успеть обновить кэш
                # между первой проверкой и получением lock.
                expected_metadata = build_task_cache_metadata(
                    source_workbook=source_workbook,
                    source_parent_sheet=source_parent_sheet,
                    source_parent_table=source_parent_table,
                    source_child_sheet=source_child_sheet,
                    source_child_table=source_child_table,
                )

                cached_data = read_valid_task_cache(
                    expected_metadata
                )

                if cached_data is not None:
                    return (*cached_data, "network-cache")

                # Не создаём кэш по данным, если исходный файл
                # изменился непосредственно во время чтения.
                for attempt in range(1, 3):
                    metadata_before = build_task_cache_metadata(
                        source_workbook=source_workbook,
                        source_parent_sheet=source_parent_sheet,
                        source_parent_table=source_parent_table,
                        source_child_sheet=source_child_sheet,
                        source_child_table=source_child_table,
                    )

                    (
                        parent_headers,
                        parent_rows,
                        child_headers,
                        child_rows,
                    ) = read_external_task_tables_from_excel(
                        source_workbook=source_workbook,
                        source_parent_sheet=source_parent_sheet,
                        source_parent_table=source_parent_table,
                        source_child_sheet=source_child_sheet,
                        source_child_table=source_child_table,
                    )

                    metadata_after = build_task_cache_metadata(
                        source_workbook=source_workbook,
                        source_parent_sheet=source_parent_sheet,
                        source_parent_table=source_parent_table,
                        source_child_sheet=source_child_sheet,
                        source_child_table=source_child_table,
                    )

                    if metadata_before == metadata_after:
                        cache_write_started_at = (
                            time_module.perf_counter()
                        )

                        write_task_cache_atomic(
                            metadata=metadata_after,
                            parent_headers=parent_headers,
                            parent_rows=parent_rows,
                            child_headers=child_headers,
                            child_rows=child_rows,
                        )

                        log_elapsed(
                            "Обновление сетевого кэша",
                            cache_write_started_at,
                            details=(
                                f"кэш={TASK_CACHE_PATH}; "
                                f"parent={len(parent_rows)}; "
                                f"child={len(child_rows)}"
                            ),
                        )

                        return (
                            parent_headers,
                            parent_rows,
                            child_headers,
                            child_rows,
                            "excel-cache-refresh",
                        )

                    TASK_LOGGER.warning(
                        "Внешний журнал изменился во время чтения; "
                        "повторная попытка %s/2",
                        attempt,
                    )

                raise RuntimeError(
                    "Внешний журнал заданий несколько раз "
                    "изменился во время чтения. "
                    "Не удалось создать согласованный кэш."
                )

            finally:
                release_task_cache_lock()

        # Кэш в данный момент обновляет другой пользователь.
        current_metadata = build_task_cache_metadata(
            source_workbook=source_workbook,
            source_parent_sheet=source_parent_sheet,
            source_parent_table=source_parent_table,
            source_child_sheet=source_child_sheet,
            source_child_table=source_child_table,
        )

        cached_data = read_valid_task_cache(
            current_metadata
        )

        if cached_data is not None:
            log_elapsed(
                "Ожидание обновления сетевого кэша",
                wait_started_at,
                details="кэш обновлён другим процессом",
            )

            return (*cached_data, "network-cache")

        if (
            not TASK_CACHE_LOCK_PATH.exists()
            or is_task_cache_lock_stale()
        ):
            # На следующей итерации попробуем захватить lock.
            continue

        if (
            time_module.perf_counter()
            >= wait_deadline
        ):
            TASK_LOGGER.warning(
                "Не удалось дождаться обновления кэша "
                "за %.1f с; читаю Excel напрямую",
                TASK_CACHE_WAIT_TIMEOUT_SECONDS,
            )

            (
                parent_headers,
                parent_rows,
                child_headers,
                child_rows,
            ) = read_external_task_tables_from_excel(
                source_workbook=source_workbook,
                source_parent_sheet=source_parent_sheet,
                source_parent_table=source_parent_table,
                source_child_sheet=source_child_sheet,
                source_child_table=source_child_table,
            )

            return (
                parent_headers,
                parent_rows,
                child_headers,
                child_rows,
                "excel-cache-timeout",
            )

        time_module.sleep(
            TASK_CACHE_POLL_INTERVAL_SECONDS
        )

# =============================================================================
# xlwings: запись в открытую книгу с формами
# =============================================================================

def find_open_book(workbook: str) -> xw.Book:
    target = str(Path(workbook)).lower()
    target_name = Path(workbook).name.lower()

    for app in xw.apps:
        for book in app.books:
            full_name = (
                str(book.fullname).lower()
                if book.fullname
                else ""
            )
            book_name = book.name.lower()

            TASK_LOGGER.info(
                "Проверяю открытую книгу: "
                "full_name=%s; "
                "совпадение пути=%s; "
                "совпадение имени=%s",
                full_name,
                full_name == target,
                book_name == target_name,
            )

            if full_name == target or book_name == target_name:
                return book

    raise FileNotFoundError(
        f"Открытая книга Excel не найдена: {workbook}. "
        f"Передавай путь к уже открытой книге с формами."
    )


def get_list_object(book: xw.Book, sheet_name: str, table_name: str):
    try:
        sheet = book.sheets[sheet_name]
    except Exception as exc:
        raise ValueError(f"В книге '{book.name}' нет листа '{sheet_name}'") from exc

    list_objects = sheet.api.ListObjects

    if list_objects.Count == 0:
        raise ValueError(f"На листе '{sheet_name}' нет умных таблиц")

    try:
        return list_objects.Item(table_name)
    except Exception as exc:
        raise ValueError(
            f"На листе '{sheet_name}' нет умной таблицы '{table_name}'"
        ) from exc


def read_headers_from_list_object(table) -> list[str]:
    values = table.HeaderRowRange.Value

    if values is None:
        return []

    if isinstance(values, tuple) and values and isinstance(values[0], tuple):
        raw_headers = values[0]
    else:
        raw_headers = values

    return [str(h).strip() if h is not None else "" for h in raw_headers]


def resize_list_object(table, data_rows: int, col_count: int) -> None:
    """
    Меняет размер умной таблицы.

    data_rows — количество строк данных без заголовка.
    Оставляем минимум одну строку данных.
    """
    header = table.HeaderRowRange
    start_cell = header.Cells(1, 1)

    total_rows = max(data_rows, 1) + 1

    end_cell = start_cell.Worksheet.Cells(
        start_cell.Row + total_rows - 1,
        start_cell.Column + col_count - 1,
    )

    new_range = start_cell.Worksheet.Range(start_cell, end_cell)
    table.Resize(new_range)

def sanitize_value_for_excel(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.replace(tzinfo=None)

    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    if isinstance(value, dt_time):
        return value.strftime("%H:%M:%S")

    if isinstance(value, timedelta):
        return value.total_seconds() / 86400

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return value

    if isinstance(value, (list, tuple, dict, set)):
        return json.dumps(value, ensure_ascii=False)

    if isinstance(value, str):
        # Excel не любит управляющие символы, кроме таба/переноса строки.
        value = "".join(
            ch for ch in value
            if ch in ("\t", "\n", "\r") or ord(ch) >= 32
        )

        # Максимальная длина текста в ячейке Excel — 32767 символов.
        if len(value) > 32767:
            value = value[:32767]

        return value

    return value

def replace_table_rows(
    table,
    rows: list[dict[str, Any]],
) -> int:
    """
    Полностью заменяет данные умной таблицы в открытой книге Excel.

    Порядок колонок берётся из текущих заголовков целевой таблицы.
    """
    headers = read_headers_from_list_object(table)
    col_count = len(headers)
    row_count = len(rows)

    table_name = str(table.Name)
    worksheet = table.Parent

    # При переходе из полностью пустой таблицы Excel не всегда создаёт
    # DataBodyRange одним вызовом Resize.
    if row_count > 0 and table.DataBodyRange is None:
        table.ListRows.Add()

        # После структурного изменения получаем COM-объект заново.
        table = worksheet.ListObjects.Item(table_name)

    resize_list_object(
        table,
        data_rows=row_count,
        col_count=col_count,
    )

    # После Resize старый COM-объект может возвращать прежнее состояние.
    table = worksheet.ListObjects.Item(table_name)
    body = table.DataBodyRange

    if row_count == 0:
        if body is not None:
            body.ClearContents()

        return 0

    if body is None:
        raise RuntimeError(
            f"После изменения размера таблицы "
            f"'{table_name}' DataBodyRange отсутствует. "
            f"Ожидалось строк данных: {row_count}"
        )

    actual_row_count = int(body.Rows.Count)

    if actual_row_count != row_count:
        raise RuntimeError(
            f"Таблица '{table_name}' получила неверный размер: "
            f"ожидалось строк {row_count}, "
            f"фактически {actual_row_count}"
        )

    matrix: list[list[Any]] = []

    for row in rows:
        matrix.append([
            sanitize_value_for_excel(row.get(header, ""))
            for header in headers
        ])

    try:
        body.Value = matrix

    except Exception:
        # Диагностика: ищем конкретную ячейку или значение,
        # на котором падает блочная запись Excel COM.
        for row_idx, row_values in enumerate(matrix, start=1):
            for col_idx, value in enumerate(row_values, start=1):
                try:
                    body.Cells(row_idx, col_idx).Value = value

                except Exception as cell_exc:
                    header = (
                        headers[col_idx - 1]
                        if col_idx - 1 < len(headers)
                        else f"col_{col_idx}"
                    )

                    raise ValueError(
                        "Не удалось записать значение в Excel.\n"
                        f"Таблица: {table_name}\n"
                        f"Строка данных: {row_idx}\n"
                        f"Колонка: {header}\n"
                        f"Значение: {value!r}\n"
                        f"Тип значения: {type(value).__name__}"
                    ) from cell_exc

        # Если поячеечная запись прошла, считаем операцию успешной.
        return row_count

    return row_count


# =============================================================================
# Фильтрация и нормализация
# =============================================================================

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_task_id(value: Any) -> int | None:
    if value is None:
        return None

    text = str(value).strip()

    if text == "":
        return None

    if text.endswith(".0") or text.endswith(",0"):
        text = text[:-2]

    return int(float(text.replace(",", ".")))


def normalize_excel_datetime(value: Any) -> datetime | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, (int, float)):
        if value <= 0:
            return None
        return datetime(1899, 12, 30) + timedelta(days=float(value))

    text = str(value).strip()

    if text == "":
        return None

    for fmt in (
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y %H:%M",
            "%d.%m.%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass

    return None


def external_parent_latest_sort_key(
        row: dict[str, Any],
        original_index: int,
) -> tuple[int, float, int]:
    """
    Чем больше ключ — тем строка считается новее.

    Приоритет:
    1. ID строки внешней таблицы
    2. Дата и время
    3. Порядок строки в файле
    """
    row_id = normalize_task_id(row.get(PARENT_ROW_ID_COL))
    dt = normalize_excel_datetime(row.get(PARENT_DATETIME_COL))

    row_id_key = row_id if row_id is not None else -1
    dt_key = dt.timestamp() if dt is not None else 0.0

    return row_id_key, dt_key, original_index


def keep_latest_parent_rows_by_task_id(
        rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Во внешней таблице заданий может быть несколько строк с одним taskId.
    Оставляем только последнюю строку по каждому 'Номер задания'.

    Дубли ищем только по taskId.
    """
    latest_by_task_id: dict[int, tuple[tuple[int, float, int], int, dict[str, Any]]] = {}
    rows_without_task_id: list[tuple[int, dict[str, Any]]] = []

    for index, row in enumerate(rows):
        task_id = normalize_task_id(row.get(PARENT_TASK_ID_COL))

        if task_id is None:
            rows_without_task_id.append((index, dict(row)))
            continue

        sort_key = external_parent_latest_sort_key(row, index)

        old = latest_by_task_id.get(task_id)

        if old is None or sort_key > old[0]:
            latest_by_task_id[task_id] = (sort_key, index, dict(row))

    selected: list[tuple[int, dict[str, Any]]] = [
        (index, row)
        for _, index, row in latest_by_task_id.values()
    ]

    selected.extend(rows_without_task_id)
    selected.sort(key=lambda item: item[0])

    return [row for _, row in selected]


def filter_parent_rows(
        rows: list[dict[str, Any]],
        project_number: str,
) -> list[dict[str, Any]]:
    """
    Родительскую таблицу фильтруем по колонке 'Код проекта'.

    Пример:
        project_number = '25-F218'
        подходит '25-F218-SRU-204-GS1'
    """
    result: list[dict[str, Any]] = []

    for row in rows:
        sample_code = normalize_text(row.get(PARENT_SAMPLE_CODE_COL))

        if sample_code.startswith(project_number):
            result.append(dict(row))

    return result


def filter_child_rows(
        rows: list[dict[str, Any]],
        task_ids: set[int],
) -> list[dict[str, Any]]:
    """
    Дочернюю таблицу фильтруем по 'Номер ГТИ',
    который соответствует taskId / Номер задания.
    """
    result: list[dict[str, Any]] = []

    for row in rows:
        task_id = normalize_task_id(row.get(CHILD_TASK_ID_COL))

        if task_id in task_ids:
            result.append(dict(row))

    return result


# =============================================================================
# SQLite: последние статусы
# =============================================================================

def fetch_latest_statuses(
        db_path: str,
        task_ids: set[int],
) -> dict[int, str]:
    """
    Возвращает последний статус из SQLite TaskStatus по taskid.

    taskId считается уникальным независимо от типа исследования.
    """
    if not task_ids:
        return {}

    task_ids_sorted = sorted(task_ids)
    placeholders = ", ".join("?" for _ in task_ids_sorted)

    sql = f"""
        SELECT
            taskid,
            status
        FROM TaskStatus
        WHERE taskid IN ({placeholders})
        ORDER BY statusid
    """

    result: dict[int, str] = {}

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row

        cur = conn.execute(sql, task_ids_sorted)

        # ORDER BY statusid ASC:
        # более поздний статус перезапишет ранний.
        for row in cur.fetchall():
            task_id = int(row["taskid"])
            status = "" if row["status"] is None else str(row["status"]).strip()

            if status:
                result[task_id] = status

    return result


def apply_sqlite_statuses(
        parent_rows: list[dict[str, Any]],
        sqlite_statuses: dict[int, str],
) -> tuple[int, int]:
    """
    Заменяет только колонку 'Статус' в родительских строках.

    Если в SQLite есть последний статус по taskId — ставим его.
    Если нет — оставляем статус из внешней Excel-книги.
    """
    sqlite_statuses_applied = 0
    external_statuses_kept = 0

    for row in parent_rows:
        task_id = normalize_task_id(row.get(PARENT_TASK_ID_COL))

        if task_id is None:
            external_statuses_kept += 1
            continue

        sqlite_status = sqlite_statuses.get(task_id)

        if sqlite_status:
            row[PARENT_STATUS_COL] = sqlite_status
            sqlite_statuses_applied += 1
        else:
            external_statuses_kept += 1

    return sqlite_statuses_applied, external_statuses_kept


# =============================================================================
# Главная логика
# =============================================================================

def _refresh_tasks_impl(
        *,
        workbook: str,
        db_path: str,
        source_workbook: str,
        project_number: str,
        source_parent_sheet: str,
        source_parent_table: str,
        source_child_sheet: str,
        source_child_table: str,
        target_parent_sheet: str,
        target_parent_table: str,
        target_child_sheet: str,
        target_child_table: str,
) -> RefreshTasksReport:

    progress(f"Ищу открытую книгу с формами: {Path(workbook).name}")

    stage_started_at = time_module.perf_counter()

    target_book = find_open_book(workbook)

    log_elapsed(
        "Поиск открытой книги с формами",
        stage_started_at,
        details=f"книга={Path(workbook).name}",
    )

    progress(
        "Получаю таблицы внешнего журнала заданий"
    )

    stage_started_at = time_module.perf_counter()

    (
        parent_headers,
        external_parent_rows_all,
        child_headers,
        external_child_rows_all,
        external_data_source,
    ) = load_external_task_tables_cached(
        source_workbook=source_workbook,
        source_parent_sheet=source_parent_sheet,
        source_parent_table=source_parent_table,
        source_child_sheet=source_child_sheet,
        source_child_table=source_child_table,
    )

    log_elapsed(
        "Получение таблиц внешнего журнала",
        stage_started_at,
        details=(
            f"источник={external_data_source}; "
            f"parent={len(external_parent_rows_all)}; "
            f"child={len(external_child_rows_all)}"
        ),
    )

    progress("Проверяю обязательные колонки родительской таблицы")

    stage_started_at = time_module.perf_counter()

    require_columns(
        parent_headers,
        [
            PARENT_TASK_ID_COL,
            PARENT_SAMPLE_CODE_COL,
            PARENT_TASK_TYPE_COL,
            PARENT_STATUS_COL,
        ],
        f"внешняя таблица {source_parent_table}",
    )

    log_elapsed(
        "Проверка колонок родительской таблицы",
        stage_started_at,
        details=f"колонок={len(parent_headers)}",
    )

    progress("Проверяю обязательные колонки дочерней таблицы")

    stage_started_at = time_module.perf_counter()

    require_columns(
        child_headers,
        [
            CHILD_TASK_ID_COL,
        ],
        f"внешняя таблица {source_child_table}",
    )

    log_elapsed(
        "Проверка колонок дочерней таблицы",
        stage_started_at,
        details=f"колонок={len(child_headers)}",
    )

    progress(f"Фильтрую задания по проекту: {project_number}")

    stage_started_at = time_module.perf_counter()

    parent_rows = filter_parent_rows(
        external_parent_rows_all,
        project_number,
    )

    log_elapsed(
        "Фильтрация родительской таблицы по проекту",
        stage_started_at,
        details=(
            f"проект={project_number}; "
            f"исходных строк={len(external_parent_rows_all)}; "
            f"найдено={len(parent_rows)}"
        ),
    )

    progress(
        f"Найдено строк родительской таблицы "
        f"по проекту: {len(parent_rows)}"
    )

    parent_rows_before_dedupe = len(parent_rows)

    progress(
        "Удаляю дубли родительской таблицы по taskId, "
        "оставляю последний статус"
    )

    stage_started_at = time_module.perf_counter()

    parent_rows = keep_latest_parent_rows_by_task_id(parent_rows)

    parent_duplicates_removed = (
        parent_rows_before_dedupe - len(parent_rows)
    )

    log_elapsed(
        "Удаление дублей родительской таблицы",
        stage_started_at,
        details=(
            f"до={parent_rows_before_dedupe}; "
            f"после={len(parent_rows)}; "
            f"удалено={parent_duplicates_removed}"
        ),
    )

    progress(
        f"После удаления дублей осталось строк: {len(parent_rows)}. "
        f"Удалено дублей: {parent_duplicates_removed}"
    )

    stage_started_at = time_module.perf_counter()

    task_ids: set[int] = set()

    for row in parent_rows:
        task_id = normalize_task_id(
            row.get(PARENT_TASK_ID_COL)
        )

        if task_id is None:
            continue

        task_ids.add(task_id)

    log_elapsed(
        "Сбор уникальных taskId",
        stage_started_at,
        details=f"уникальных taskId={len(task_ids)}",
    )

    progress(f"Собрано уникальных taskId: {len(task_ids)}")

    progress("Читаю последние статусы из SQLite TaskStatus")

    stage_started_at = time_module.perf_counter()

    sqlite_statuses = fetch_latest_statuses(
        db_path,
        task_ids,
    )

    log_elapsed(
        "Чтение последних статусов из SQLite",
        stage_started_at,
        details=(
            f"запрошено taskId={len(task_ids)}; "
            f"получено статусов={len(sqlite_statuses)}"
        ),
    )

    progress(
        f"Найдено статусов в SQLite: {len(sqlite_statuses)}"
    )

    progress(
        "Подмешиваю статусы SQLite в родительскую таблицу"
    )

    stage_started_at = time_module.perf_counter()

    sqlite_statuses_applied, external_statuses_kept = (
        apply_sqlite_statuses(
            parent_rows,
            sqlite_statuses,
        )
    )

    log_elapsed(
        "Подмешивание статусов SQLite",
        stage_started_at,
        details=(
            f"из SQLite={sqlite_statuses_applied}; "
            f"оставлено из внешнего Excel={external_statuses_kept}"
        ),
    )

    progress(
        f"Статусов взято из SQLite: "
        f"{sqlite_statuses_applied}; "
        f"оставлено из внешнего Excel: "
        f"{external_statuses_kept}"
    )

    progress(
        "Фильтрую дочернюю таблицу task_mix по taskId"
    )

    stage_started_at = time_module.perf_counter()

    child_rows = filter_child_rows(
        external_child_rows_all,
        task_ids,
    )

    log_elapsed(
        "Фильтрация дочерней таблицы task_mix",
        stage_started_at,
        details=(
            f"исходных строк={len(external_child_rows_all)}; "
            f"найдено={len(child_rows)}"
        ),
    )

    progress(
        f"Строк task_mix после фильтрации: {len(child_rows)}"
    )

    progress(
        f"Ищу целевую таблицу Excel: "
        f"{target_parent_sheet}/{target_parent_table}"
    )

    stage_started_at = time_module.perf_counter()

    target_parent_table_obj = get_list_object(
        target_book,
        target_parent_sheet,
        target_parent_table,
    )

    log_elapsed(
        "Поиск целевой таблицы Task",
        stage_started_at,
        details=(
            f"лист={target_parent_sheet}; "
            f"таблица={target_parent_table}"
        ),
    )

    progress(
        f"Ищу целевую таблицу Excel: "
        f"{target_child_sheet}/{target_child_table}"
    )

    stage_started_at = time_module.perf_counter()

    target_child_table_obj = get_list_object(
        target_book,
        target_child_sheet,
        target_child_table,
    )

    log_elapsed(
        "Поиск целевой таблицы Task_mix",
        stage_started_at,
        details=(
            f"лист={target_child_sheet}; "
            f"таблица={target_child_table}"
        ),
    )

    progress(
        f"Записываю родительскую таблицу "
        f"{target_parent_table}: {len(parent_rows)} строк"
    )

    stage_started_at = time_module.perf_counter()

    parent_written = replace_table_rows(
        target_parent_table_obj,
        parent_rows,
    )

    log_elapsed(
        "Запись родительской таблицы Task",
        stage_started_at,
        details=f"записано строк={parent_written}",
    )

    progress(
        f"Записываю дочернюю таблицу "
        f"{target_child_table}: {len(child_rows)} строк"
    )

    stage_started_at = time_module.perf_counter()

    child_written = replace_table_rows(
        target_child_table_obj,
        child_rows,
    )

    log_elapsed(
        "Запись дочерней таблицы Task_mix",
        stage_started_at,
        details=f"записано строк={child_written}",
    )

    progress("Сохраняю книгу с формами")

    stage_started_at = time_module.perf_counter()

    target_book.save()

    log_elapsed(
        "Сохранение книги с формами",
        stage_started_at,
        details=f"книга={target_book.name}",
    )

    return RefreshTasksReport(
        ok=True,
        message=(
            "Задания успешно обновлены. "
            f"Дубликатов по taskId удалено: "
            f"{parent_duplicates_removed}"
        ),
        project=project_number,
        parent_rows_read=len(external_parent_rows_all),
        parent_rows_written=parent_written,
        child_rows_read=len(external_child_rows_all),
        child_rows_written=child_written,
        sqlite_statuses_applied=sqlite_statuses_applied,
        external_statuses_kept=external_statuses_kept,
    )

def refresh_tasks(
        *,
        workbook: str,
        db_path: str,
        source_workbook: str,
        project_number: str,
        source_parent_sheet: str,
        source_parent_table: str,
        source_child_sheet: str,
        source_child_table: str,
        target_parent_sheet: str,
        target_parent_table: str,
        target_child_sheet: str,
        target_child_table: str,
) -> RefreshTasksReport:

    total_started_at = time_module.perf_counter()
    completed_ok = False

    TASK_LOGGER.info("=" * 100)
    TASK_LOGGER.info(
        "Начало refresh_tasks: project=%s; workbook=%s",
        project_number,
        workbook,
    )

    try:
        report = _refresh_tasks_impl(
            workbook=workbook,
            db_path=db_path,
            source_workbook=source_workbook,
            project_number=project_number,
            source_parent_sheet=source_parent_sheet,
            source_parent_table=source_parent_table,
            source_child_sheet=source_child_sheet,
            source_child_table=source_child_table,
            target_parent_sheet=target_parent_sheet,
            target_parent_table=target_parent_table,
            target_child_sheet=target_child_sheet,
            target_child_table=target_child_table,
        )

        completed_ok = True
        return report

    except Exception:
        TASK_LOGGER.exception(
            "Ошибка refresh_tasks: project=%s",
            project_number,
        )
        raise

    finally:
        total_elapsed = (
            time_module.perf_counter() - total_started_at
        )

        TASK_LOGGER.info(
            "Завершение refresh_tasks: project=%s; "
            "успешно=%s; общее время=%.3f с",
            project_number,
            completed_ok,
            total_elapsed,
        )

        # Принудительно сбрасываем буфер лог-файла,
        # чтобы записи сразу появились на диске.
        for handler in TASK_LOGGER.handlers:
            try:
                handler.flush()
            except Exception:
                pass

        gc.collect()

def require_columns(
        actual_headers: list[str],
        required_headers: list[str],
        source_name: str,
) -> None:
    missing = [
        header
        for header in required_headers
        if header not in actual_headers
    ]

    if missing:
        raise ValueError(
            f"Не найдены обязательные колонки в {source_name}: {missing}. "
            f"Фактические колонки: {actual_headers}"
        )


def run_json(**kwargs: Any) -> str:
    try:
        report = refresh_tasks(**kwargs)
        return json.dumps(asdict(report), ensure_ascii=False, indent=2)

    except Exception as exc:
        traceback_text = traceback.format_exc()
        error_type = type(exc).__name__
        message = str(exc)

        write_error_log(
            message=message,
            error_type=error_type,
            traceback_text=traceback_text,
        )

        report = RefreshTasksReport(
            ok=False,
            message=(
                f"{error_type}: {message}\n\n"
                f"Полный traceback записан в лог:\n{ERROR_LOG_PATH}"
            ),
            project=str(kwargs.get("project_number", "")),
            error_type=error_type,
            traceback=None,
        )

        return json.dumps(asdict(report), ensure_ascii=False, indent=2)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Обновление таблиц task/task_mix из внешней Excel-книги "
            "с подмешиванием последних статусов из SQLite"
        )
    )
    parser.add_argument("--progress-file", default=None)
    parser.add_argument(
        "--workbook",
        required=True,
        help="Текущая открытая книга с формами",
    )

    parser.add_argument(
        "--db",
        required=True,
        help="Путь к SQLite БД",
    )

    parser.add_argument(
        "--source-workbook",
        required=True,
        help="Внешняя книга Excel с заданиями",
    )

    parser.add_argument(
        "--project",
        required=True,
        help="Номер проекта, например 25-F218",
    )

    parser.add_argument("--source-parent-sheet", default="task")
    parser.add_argument("--source-parent-table", default="task")
    parser.add_argument("--source-child-sheet", default="task_mix")
    parser.add_argument("--source-child-table", default="task_mix")

    parser.add_argument("--target-parent-sheet", default="Task")
    parser.add_argument("--target-parent-table", default="Task")
    parser.add_argument("--target-child-sheet", default="Task_mix")
    parser.add_argument("--target-child-table", default="Task_mix")

    args = parser.parse_args(argv)
    set_progress_file(args.progress_file)

    json_result = run_json(
        workbook=args.workbook,
        db_path=args.db,
        source_workbook=args.source_workbook,
        project_number=args.project,
        source_parent_sheet=args.source_parent_sheet,
        source_parent_table=args.source_parent_table,
        source_child_sheet=args.source_child_sheet,
        source_child_table=args.source_child_table,
        target_parent_sheet=args.target_parent_sheet,
        target_parent_table=args.target_parent_table,
        target_child_sheet=args.target_child_sheet,
        target_child_table=args.target_child_table,
    )

    print(json_result)

    try:
        parsed = json.loads(json_result)
        return 0 if parsed.get("ok") else 1
    except Exception:
        return 1


if __name__ == "__main__":
    sys.exit(main())
