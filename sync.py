from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
import traceback
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_io import ExcelWorkbookGateway, ExcelTableData
from repository import SQLiteRepository, SaveTableResult
from schema import StudySchema, TableSchema, get_studies
from task_status import TaskStatusRepository


TEMP_ID_RE = re.compile(r"^temp_\d+$", re.IGNORECASE)
PROJECT_NUMBER_RE = re.compile(r"^\d{2}-F\d{3}$")

PROJECT_DIR = Path(__file__).resolve().parent
ERROR_LOG_PATH = PROJECT_DIR / "sync_errors.log"
TASK_STATUS_EXPORT_PATH = PROJECT_DIR / "Журнал заданий ГТИ.xlsx"

def write_error_log(
    *,
    mode: str,
    message: str,
    error_type: str,
    traceback_text: str,
) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with ERROR_LOG_PATH.open("a", encoding="utf-8") as f:
        f.write("\n" + "=" * 100 + "\n")
        f.write(f"datetime: {now}\n")
        f.write(f"mode: {mode}\n")
        f.write(f"error_type: {error_type}\n")
        f.write(f"message: {message}\n")
        f.write("-" * 100 + "\n")
        f.write(traceback_text)
        f.write("\n")

def progress(message: str) -> None:
    print(
        f"[{datetime.now().strftime('%H:%M:%S')}] {message}",
        file=sys.stderr,
        flush=True,
    )

@dataclass
class SyncTableReport:
    """Отчёт по одной таблице Excel/SQLite."""

    logical_name: str
    sheet_name: str
    db_table: str
    direction: str
    rows_read: int = 0
    rows_written: int = 0
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    deleted: int = 0
    temp_ids_replaced: int = 0


@dataclass
class SyncReport:
    """Общий отчёт, который скрипт возвращает в Excel в JSON-формате."""

    ok: bool
    message: str
    mode: str
    tables: list[SyncTableReport] = field(default_factory=list)
    error_type: str | None = None
    traceback: str | None = None


class StudySyncService:
    """
    Главный сервис синхронизации Excel <-> SQLite.

    Поддерживает два режима:
    - save: Excel -> SQLite -> запись реальных id обратно в Excel;
    - load: SQLite -> Excel, фильтр по номеру проекта.

    Принятый вариант временных связей:
    - VBA при сведении может записать во внешний ключ дочерней таблицы временный ключ
      вида temp_1, temp_2, temp_3;
    - такой же временный ключ должен стоять в id-колонке родительской строки;
    - Python при сохранении считает temp_* не настоящим id БД, а временной связью;
    - родительская строка вставляется как новая, получает настоящий id из SQLite;
    - дочерние строки с temp_* получают настоящий resultId* перед INSERT;
    - в Excel затем блочно перезаписываются реальные id.
    """

    def __init__(self, workbook: str | None, db_path: str):
        self.excel = ExcelWorkbookGateway(workbook) if workbook else None
        self.repo = SQLiteRepository(db_path)
        self.task_status = TaskStatusRepository(self.repo.conn)

    def close(self) -> None:
        """Закрывает соединение с БД."""
        self.repo.close()

    def save_to_db(
            self,
            studies: list[StudySchema],
            *,
            task_status: str = TaskStatusRepository.COMPLETED_STATUS,
    ) -> SyncReport:
        """
        Сохраняет выбранные исследования из Excel в SQLite.

        Для исследования без дочерней таблицы:
        - temp_* в id-колонке считается пустым id;
        - строка вставляется как новая;
        - реальный id записывается обратно в Excel.

        Для исследования с дочерней таблицей:
        - сначала сохраняется родительская таблица;
        - строится словарь temp_* -> настоящий resultId*;
        - во всех дочерних строках temp_* во fk_column заменяется на настоящий id;
        - сохраняется дочерняя таблица;
        - в Excel блочно перезаписываются и rowId*, и resultId* дочерней таблицы.
        """

        if self.excel is None:
            raise ValueError("Для этого режима нужна Excel-книга")

        progress("Проверяю возможность установки статуса...")

        # ВАЖНО: до начала сохранения и до любых write_ids в Excel.
        self._validate_all_studies_can_receive_status(
            studies=studies,
            status=task_status,
        )

        reports: list[SyncTableReport] = []
        progress("Начинаю транзакцию SQLite...")
        self.repo.begin()

        try:
            for study in studies:
                progress(f"Синхронизация исследования {study.code}...")
                if study.source_table is None:
                    reports.append(
                        self._save_single_table(
                            study.result_table,
                            task_status=task_status,
                        )
                    )
                else:
                    parent_report, child_report = self._save_parent_and_child(
                        study,
                        task_status=task_status,
                    )
                    reports.append(parent_report)
                    reports.append(child_report)
            progress("Фиксирую изменения в SQLite...")
            self.repo.commit()

            progress("Сохраняю Excel-книгу...")
            self.excel.save()
            return SyncReport(
                ok=True,
                mode="save",
                message="Сохранение Excel -> SQLite успешно выполнено",
                tables=reports,
            )
        except Exception:
            self.repo.rollback()
            raise

    def load_from_db(self, studies: list[StudySchema], project_number: str) -> SyncReport:
        """
        Загружает актуальные данные из SQLite в Excel.

        Родительские таблицы фильтруются по:
            sampleCode LIKE '<project_number>-%'

        Дочерние таблицы загружаются по id найденных родительских строк.
        Старое содержимое умных таблиц Excel заменяется актуальным содержимым из БД.
        """

        if self.excel is None:
            raise ValueError("Для этого режима нужна Excel-книга")

        self._validate_project_number(project_number)
        reports: list[SyncTableReport] = []

        for study in studies:
            progress(f"Загружаю {study.code}: читаю результаты из БД...")
            parent_rows = self.repo.fetch_results_by_project(study.result_table, project_number)

            progress(f"Загружаю {study.code}: записываю results в Excel...")
            parent_written = self.excel.replace_table_rows(study.result_table, parent_rows)
            reports.append(
                SyncTableReport(
                    logical_name=study.result_table.logical_name,
                    sheet_name=study.result_table.sheet_name,
                    db_table=study.result_table.db_table,
                    direction="db_to_excel",
                    rows_read=len(parent_rows),
                    rows_written=parent_written,
                )
            )

            if study.source_table is not None:
                parent_ids = self._extract_ids(study.result_table, parent_rows)
                progress(f"Загружаю {study.code}: читаю sourceData из БД...")
                child_rows = self.repo.fetch_child_rows_by_parent_ids(study.source_table, parent_ids)
                progress(f"Загружаю {study.code}: записываю sourceData в Excel...")
                child_written = self.excel.replace_table_rows(study.source_table, child_rows)
                reports.append(
                    SyncTableReport(
                        logical_name=study.source_table.logical_name,
                        sheet_name=study.source_table.sheet_name,
                        db_table=study.source_table.db_table,
                        direction="db_to_excel",
                        rows_read=len(child_rows),
                        rows_written=child_written,
                    )
                )

        self.excel.save()
        return SyncReport(
            ok=True,
            mode="load",
            message=f"Загрузка SQLite -> Excel по проекту {project_number} успешно выполнена",
            tables=reports,
        )

    def set_task_status(
            self,
            *,
            taskid: int,
            task_type: str,
            status: str,
            result_id: int | None = None,
            date_time: Any = None,
    ) -> SyncReport:

        progress(f"Устанавливаю статус задания: {taskid} / {task_type} -> {status}")
        self.repo.begin()

        try:
            created = self.task_status.ensure_latest_status_for_result(
                taskid=taskid,
                task_type=task_type,
                status=status,
                result_id=result_id,
                date_time=self._get_status_datetime(
                    status=status,
                    explicit_date_time=date_time,
                ),
            )

            self.repo.commit()

            return SyncReport(
                ok=True,
                mode="status",
                message=(
                    f"Статус '{status}' для задания {taskid} / {task_type} "
                    f"{'добавлен' if created else 'уже был последним'}"
                ),
                tables=[],
            )

        except Exception:
            self.repo.rollback()
            raise

    def _save_single_table(
            self,
            schema: TableSchema,
            *,
            task_status: str,
    ) -> SyncTableReport:
        """
        Сохраняет одиночную таблицу без дочерних данных.

        temp_* в primary key считается временным id, поэтому перед сохранением
        заменяется на None. После INSERT настоящий id записывается обратно в Excel.
        """
        progress(f"Читаю таблицу Excel: {schema.excel_table_name}")
        table_data = self.excel.read_table(schema)

        study_code = schema.logical_name.replace("_RESULTS", "")

        rows_to_process, row_numbers_to_process, skipped_status_rows = (
            self._filter_result_rows_for_status(
                study_code=study_code,
                rows=table_data.rows,
                row_numbers=table_data.row_numbers,
                status=task_status,
            )
        )

        prepared_rows, temp_count = self._prepare_rows_for_save(schema, rows_to_process)

        progress(f"Сохраняю строки в БД: {schema.db_table}, строк: {len(prepared_rows)}")
        save_result: SaveTableResult = self.repo.save_rows(schema, prepared_rows)

        progress(f"Обновляю TaskStatus для {schema.logical_name}")
        self._ensure_task_statuses_for_results(
            study_code=schema.logical_name.replace("_RESULTS", ""),
            result_schema=schema,
            rows=prepared_rows,
            saved_ids=save_result.ids or [],
            status=task_status,
        )

        progress(f"Записываю новые ID обратно в Excel: {schema.excel_table_name}")
        self.excel.write_ids(schema, row_numbers_to_process, save_result.ids or [])

        return SyncTableReport(
            logical_name=schema.logical_name,
            sheet_name=schema.sheet_name,
            db_table=schema.db_table,
            direction="excel_to_db",
            rows_read=len(table_data.rows),
            rows_written=len(save_result.ids or []),
            inserted=save_result.inserted,
            updated=save_result.updated,
            skipped=save_result.skipped + len(skipped_status_rows),
            temp_ids_replaced=temp_count,
        )

    def _save_parent_and_child(
            self,
            study: StudySchema,
            *,
            task_status: str,
    ) -> tuple[SyncTableReport, SyncTableReport]:
        """
        Сохраняет пару таблиц: родительскую *_results и дочернюю *_sourceData.

        Основная логика temp_* находится здесь:
        - temp_* из parent.pk используется как временный ключ;
        - после сохранения родителя строится temp_to_real_id;
        - child.fk_column заменяется по temp_to_real_id.
        """
        parent_schema = study.result_table
        child_schema = study.source_table
        if child_schema is None:
            raise ValueError("Для _save_parent_and_child нужна дочерняя таблица")
        if not child_schema.fk_column:
            raise ValueError(f"Для таблицы {child_schema.logical_name} не указан fk_column")

        progress(f"Читаю родительскую таблицу Excel: {parent_schema.excel_table_name}")
        parent_data = self.excel.read_table(parent_schema)

        parent_rows_to_process, parent_row_numbers_to_process, skipped_parent_rows = (
            self._filter_result_rows_for_status(
                study_code=study.code,
                rows=parent_data.rows,
                row_numbers=parent_data.row_numbers,
                status=task_status,
            )
        )

        parent_rows, parent_temp_count = self._prepare_rows_for_save(
            parent_schema,
            parent_rows_to_process,
        )

        progress(f"Сохраняю родительскую таблицу в БД: {parent_schema.db_table}")
        parent_save_result = self.repo.save_rows(parent_schema, parent_rows)
        parent_ids = parent_save_result.ids or []

        progress(f"Обновляю TaskStatus для {study.code}")
        self._ensure_task_statuses_for_results(
            study_code=study.code,
            result_schema=parent_schema,
            rows=parent_rows,
            saved_ids=parent_ids,
            status=task_status,
        )

        progress(f"Записываю parent ID обратно в Excel: {parent_schema.excel_table_name}")
        self.excel.write_ids(parent_schema, parent_row_numbers_to_process, parent_ids)
        temp_to_real_id = self._build_temp_to_real_parent_id_map(
            schema=parent_schema,
            original_rows=parent_rows_to_process,
            saved_ids=parent_ids,
        )

        progress(f"Читаю дочернюю таблицу Excel: {child_schema.excel_table_name}")
        child_data = self.excel.read_table(child_schema)
        child_rows_raw = child_data.rows
        child_row_numbers_raw = child_data.row_numbers

        skipped_parent_keys = self._collect_parent_keys(
            schema=parent_schema,
            rows=skipped_parent_rows,
        )

        if skipped_parent_keys:
            filtered_child_rows = []
            filtered_child_row_numbers = []

            for child_row, child_row_number in zip(child_rows_raw, child_row_numbers_raw):
                fk_value = child_row.get(child_schema.fk_column)
                fk_key = self._normalize_temp_id(fk_value) if self._is_temp_id(fk_value) else str(fk_value).strip()

                if fk_key in skipped_parent_keys:
                    continue

                filtered_child_rows.append(child_row)
                filtered_child_row_numbers.append(child_row_number)
        else:
            filtered_child_rows = child_rows_raw
            filtered_child_row_numbers = child_row_numbers_raw

        child_rows, child_pk_temp_count = self._prepare_rows_for_save(
            child_schema,
            filtered_child_rows,
        )

        child_fk_temp_count = self._replace_child_temp_fk_with_real_id(
            child_schema=child_schema,
            child_rows=child_rows,
            child_data=child_data,
            temp_to_real_id=temp_to_real_id,
        )

        rows_to_save, save_indices, ids_to_delete, delete_indices = (
            self._split_child_rows_for_save_delete(child_schema, child_rows)
        )

        progress(f"Удаляю очищенные дочерние строки из БД: {child_schema.db_table}")
        deleted_count = self.repo.delete_rows_by_pk(child_schema, ids_to_delete)

        progress(f"Сохраняю дочернюю таблицу в БД: {child_schema.db_table}")
        child_save_result = self.repo.save_rows(child_schema, rows_to_save)
        saved_child_ids = child_save_result.ids or []

        # id для записи обратно в Excel в исходном порядке child_rows
        child_ids_for_excel: list[Any] = [None for _ in child_rows]

        for original_index, saved_id in zip(save_indices, saved_child_ids):
            child_ids_for_excel[original_index] = saved_id

        for original_index in delete_indices:
            child_ids_for_excel[original_index] = None

        progress(f"Записываю child ID обратно в Excel: {child_schema.excel_table_name}")
        self.excel.write_ids(
            child_schema,
            filtered_child_row_numbers,
            child_ids_for_excel,
        )

        self._write_excel_column_values_block(
            schema=child_schema,
            column_db_name=child_schema.fk_column,
            row_numbers=filtered_child_row_numbers,
            values=[row.get(child_schema.fk_column) for row in child_rows],
        )

        parent_report = SyncTableReport(
            logical_name=parent_schema.logical_name,
            sheet_name=parent_schema.sheet_name,
            db_table=parent_schema.db_table,
            direction="excel_to_db",
            rows_read=len(parent_data.rows),
            rows_written=len(parent_ids),
            inserted=parent_save_result.inserted,
            updated=parent_save_result.updated,
            skipped=parent_save_result.skipped + len(skipped_parent_rows),
            temp_ids_replaced=parent_temp_count,
        )
        skipped_child_by_status = len(child_rows_raw) - len(filtered_child_rows)
        child_report = SyncTableReport(
            logical_name=child_schema.logical_name,
            sheet_name=child_schema.sheet_name,
            db_table=child_schema.db_table,
            direction="excel_to_db",
            rows_read=len(child_data.rows),
            rows_written=len(saved_child_ids),
            inserted=child_save_result.inserted,
            updated=child_save_result.updated,
            deleted=deleted_count,
            skipped=child_save_result.skipped + skipped_child_by_status,
            temp_ids_replaced=child_pk_temp_count + child_fk_temp_count,
        )

        return parent_report, child_report

    def _collect_parent_keys(
            self,
            *,
            schema: TableSchema,
            rows: list[dict[str, Any]],
    ) -> set[str]:
        pk_name = schema.pk.db_name
        keys: set[str] = set()

        for row in rows:
            value = row.get(pk_name)
            if value is None or str(value).strip() == "":
                continue

            keys.add(self._normalize_temp_id(value) if self._is_temp_id(value) else str(value).strip())

        return keys

    def _should_skip_result_row_for_status(
            self,
            *,
            study_code: str,
            row: dict[str, Any],
            status: str,
    ) -> bool:
        normalized_status = self.task_status._normalize_status(status)

        # Пропускаем строки только при обычном сохранении "Завершено".
        if normalized_status != TaskStatusRepository.COMPLETED_STATUS:
            return False

        task_id_raw = row.get("TaskId")

        # Исследования без задания сохраняем как обычно.
        if task_id_raw is None or str(task_id_raw).strip() == "":
            return False

        taskid = int(task_id_raw)

        latest_status = self.task_status.get_latest_status(
            taskid=taskid,
            task_type=study_code,
        )

        return latest_status == TaskStatusRepository.VALIDATED_STATUS

    def _filter_result_rows_for_status(
            self,
            *,
            study_code: str,
            rows: list[dict[str, Any]],
            row_numbers: list[int],
            status: str,
    ) -> tuple[list[dict[str, Any]], list[int], list[dict[str, Any]]]:
        filtered_rows: list[dict[str, Any]] = []
        filtered_row_numbers: list[int] = []
        skipped_rows: list[dict[str, Any]] = []

        for row, row_number in zip(rows, row_numbers):
            if self._should_skip_result_row_for_status(
                    study_code=study_code,
                    row=row,
                    status=status,
            ):
                skipped_rows.append(row)
                continue

            filtered_rows.append(row)
            filtered_row_numbers.append(row_number)

        return filtered_rows, filtered_row_numbers, skipped_rows

    def _prepare_rows_for_save(
        self,
        schema: TableSchema,
        rows: list[dict[str, Any]],
    ) -> tuple[list[dict[str, Any]], int]:
        """
        Готовит строки к сохранению в SQLite.

        Делает копии строк. Если primary key имеет вид temp_*, он заменяется на None,
        чтобы repository выполнил INSERT без попытки UPDATE по строковому id.
        """
        pk_name = schema.pk.db_name
        prepared_rows: list[dict[str, Any]] = []
        temp_count = 0

        for row in rows:
            new_row = dict(row)
            if self._is_temp_id(new_row.get(pk_name)):
                new_row[pk_name] = None
                temp_count += 1
            prepared_rows.append(new_row)

        return prepared_rows, temp_count

    def _build_temp_to_real_parent_id_map(
        self,
        schema: TableSchema,
        original_rows: list[dict[str, Any]],
        saved_ids: list[int],
    ) -> dict[str, int]:
        """
        Строит словарь temp_* -> настоящий id БД для родительской таблицы.

        original_rows и saved_ids должны идти в одном порядке. Это обеспечивается тем,
        что repository.save_rows возвращает ids в порядке входных строк.
        """
        if len(original_rows) != len(saved_ids):
            raise ValueError(
                f"Таблица {schema.logical_name}: количество исходных строк и сохранённых id не совпадает"
            )

        pk_name = schema.pk.db_name
        mapping: dict[str, int] = {}

        for row, real_id in zip(original_rows, saved_ids):
            raw_id = row.get(pk_name)
            if not self._is_temp_id(raw_id):
                continue

            temp_id = self._normalize_temp_id(raw_id)
            if temp_id in mapping:
                raise ValueError(
                    f"Таблица {schema.logical_name}: временный id '{temp_id}' встречается несколько раз "
                    f"в родительской таблице. Он должен быть уникален в пределах таблицы."
                )
            mapping[temp_id] = int(real_id)

        return mapping

    def _replace_child_temp_fk_with_real_id(
        self,
        child_schema: TableSchema,
        child_rows: list[dict[str, Any]],
        child_data: ExcelTableData,
        temp_to_real_id: dict[str, int],
    ) -> int:
        """
        Заменяет temp_* во внешнем ключе дочерней таблицы на настоящий id родителя.

        Если в дочерней строке указан temp_*, которого нет в родительской таблице,
        выбрасывается подробная ошибка с номером строки DataBodyRange.
        """
        fk_column = child_schema.fk_column
        if not fk_column:
            raise ValueError(f"Для таблицы {child_schema.logical_name} не указан fk_column")

        replaced = 0

        for index, row in enumerate(child_rows):
            fk_value = row.get(fk_column)
            if not self._is_temp_id(fk_value):
                continue

            temp_id = self._normalize_temp_id(fk_value)
            real_id = temp_to_real_id.get(temp_id)
            if real_id is None:
                excel_row_number = child_data.row_numbers[index]
                raise ValueError(
                    f"Таблица {child_schema.logical_name}, строка DataBodyRange #{excel_row_number}: "
                    f"в столбце {fk_column} указан временный ключ '{temp_id}', "
                    f"но такой ключ не найден в родительской таблице."
                )

            row[fk_column] = real_id
            replaced += 1

        return replaced

    def _write_excel_column_values_block(
        self,
        schema: TableSchema,
        column_db_name: str,
        row_numbers: list[int],
        values: list[Any],
    ) -> None:
        """
        Блочно записывает значения в произвольный столбец умной таблицы Excel.

        Здесь это используется для перезаписи дочернего resultId* после замены temp_*
        на настоящий id родительской строки. Запись идёт одним блоком в весь столбец,
        связь строк сохраняется через row_numbers.
        """
        if len(row_numbers) != len(values):
            raise ValueError("Количество строк Excel и количество значений для записи не совпадает")
        if not row_numbers:
            return

        table = self.excel._get_list_object(schema)
        excel_column_name = schema.db_to_excel[column_db_name]
        headers = [str(h).strip() if h is not None else "" for h in table.HeaderRowRange.Value[0]]

        try:
            column_index = headers.index(excel_column_name) + 1
        except ValueError as exc:
            raise ValueError(f"В таблице '{table.Name}' нет столбца '{excel_column_name}'") from exc

        body = table.DataBodyRange
        if body is None:
            return

        column_range = body.Columns(column_index)
        current_values = column_range.Value
        total_rows = body.Rows.Count

        if total_rows == 1:
            current_matrix = [[current_values]]
        else:
            current_matrix = [[row[0]] for row in current_values]

        for row_number, value in zip(row_numbers, values):
            current_matrix[row_number - 1][0] = value

        column_range.Value = current_matrix

    def _ensure_task_statuses_for_results(
            self,
            *,
            study_code: str,
            result_schema: TableSchema,
            rows: list[dict[str, Any]],
            saved_ids: list[int],
            status: str,
    ) -> int:
        """
        Создаёт записи TaskStatus для сохранённых строк результата.

        Статус создаётся только если:
        - в строке есть непустой TaskId;
        - для пары task_type + taskid ещё нет записи в TaskStatus.

        Возвращает количество созданных записей.
        """
        if len(rows) != len(saved_ids):
            raise ValueError(
                f"{result_schema.logical_name}: количество строк и id не совпадает "
                f"при создании TaskStatus"
            )

        created = 0

        for row, result_id in zip(rows, saved_ids):
            task_id_raw = row.get("TaskId")

            if task_id_raw is None or str(task_id_raw).strip() == "":
                continue

            taskid = int(task_id_raw)

            was_created = self.task_status.ensure_latest_status_for_result(
                taskid=taskid,
                task_type=study_code,
                status=status,
                result_id=int(result_id),
                date_time=self._get_status_datetime(
                    status=status,
                    result_schema=result_schema,
                    row=row,
                ),
            )

            if was_created:
                created += 1

        return created

    @staticmethod
    def _is_empty_value(value: Any) -> bool:
        if value is None:
            return True
        return str(value).strip() == ""

    def _child_row_has_data(self, schema: TableSchema, row: dict[str, Any]) -> bool:
        """
        Проверяет, есть ли в дочерней строке реальные данные.

        Для AP_sourceData смотрим только pressure/density,
        чтобы active=True не мешал определить удалённую строку.
        """
        check_columns = schema.delete_if_empty_columns

        if not check_columns:
            check_columns = tuple(
                c.db_name
                for c in schema.columns
                if not c.primary_key and c.db_name != schema.fk_column
            )

        return any(
            not self._is_empty_value(row.get(column_name))
            for column_name in check_columns
        )

    def _get_status_datetime(
            self,
            *,
            status: str,
            result_schema: TableSchema | None = None,
            row: dict[str, Any] | None = None,
            explicit_date_time: Any = None,
    ) -> str:
        normalized_status = self.task_status._normalize_status(status)

        if explicit_date_time is not None and str(explicit_date_time).strip():
            return str(explicit_date_time).strip()

        if normalized_status in (
                TaskStatusRepository.VALIDATED_STATUS,
                TaskStatusRepository.CANCELED_STATUS,
        ):
            return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if result_schema is not None and row is not None:
            return self._get_experiment_datetime(result_schema, row)

        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _validate_all_studies_can_receive_status(
            self,
            *,
            studies: list[StudySchema],
            status: str,
    ) -> None:
        normalized_status = self.task_status._normalize_status(status)

        if normalized_status != TaskStatusRepository.VALIDATED_STATUS:
            return

        checked: set[tuple[str, int]] = set()

        for study in studies:
            parent_schema = study.result_table

            table_data = self.excel.read_table(parent_schema)
            parent_rows, _ = self._prepare_rows_for_save(parent_schema, table_data.rows)

            for row in parent_rows:
                task_id_raw = row.get("TaskId")

                if task_id_raw is None or str(task_id_raw).strip() == "":
                    continue

                taskid = int(task_id_raw)
                key = (study.code, taskid)

                if key in checked:
                    continue

                self.task_status.ensure_can_set_status(
                    taskid=taskid,
                    task_type=study.code,
                    status=normalized_status,
                )

                checked.add(key)

    def _validate_tasks_can_receive_status(
            self,
            *,
            study_code: str,
            rows: list[dict[str, Any]],
            status: str,
    ) -> None:
        normalized_status = self.task_status._normalize_status(status)

        if normalized_status != TaskStatusRepository.VALIDATED_STATUS:
            return

        checked: set[int] = set()

        for row in rows:
            task_id_raw = row.get("TaskId")

            if task_id_raw is None or str(task_id_raw).strip() == "":
                continue

            taskid = int(task_id_raw)

            if taskid in checked:
                continue

            self.task_status.ensure_can_set_status(
                taskid=taskid,
                task_type=study_code,
                status=normalized_status,
            )

            checked.add(taskid)

    def _split_child_rows_for_save_delete(
            self,
            schema: TableSchema,
            rows: list[dict[str, Any]],
    ) -> tuple[list[dict[str, Any]], list[int], list[Any], list[int]]:
        """
        Делит дочерние строки на save и delete.

        Возвращает:
            rows_to_save
            save_indices
            ids_to_delete
            delete_indices
        """
        pk_name = schema.pk.db_name

        rows_to_save: list[dict[str, Any]] = []
        save_indices: list[int] = []
        ids_to_delete: list[Any] = []
        delete_indices: list[int] = []

        for index, row in enumerate(rows):
            pk_value = row.get(pk_name)
            has_pk = not self._is_empty_value(pk_value)
            has_data = self._child_row_has_data(schema, row)

            if has_data:
                rows_to_save.append(row)
                save_indices.append(index)
            elif has_pk:
                ids_to_delete.append(pk_value)
                delete_indices.append(index)
            else:
                # rowId пустой и данных нет — просто игнорируем.
                pass

        return rows_to_save, save_indices, ids_to_delete, delete_indices

    @staticmethod
    def _get_experiment_datetime(result_schema: TableSchema, row: dict[str, Any]) -> str:
        """
        Возвращает дату/время эксперимента для TaskStatus.

        Если в schema.py указано:
            status_datetime_fields=("date",)
        вернёт значение date.

        Если указано:
            status_datetime_fields=("dateStart", "timeStart")
        вернёт "dateStart timeStart".
        """
        fields = result_schema.status_datetime_fields

        if not fields:
            value = row.get("dateTimeSync")
            return "" if value is None else str(value)

        values: list[str] = []

        for field_name in fields:
            value = row.get(field_name)
            if value is None:
                continue

            text = str(value).strip()
            if text:
                values.append(text)

        return " ".join(values)

    @staticmethod
    def _is_temp_id(value: Any) -> bool:
        """True, если значение выглядит как временный ключ temp_1, temp_2 и т.п."""
        if value is None:
            return False
        return TEMP_ID_RE.match(str(value).strip()) is not None

    @staticmethod
    def _normalize_temp_id(value: Any) -> str:
        """Нормализует temp-id к нижнему регистру без пробелов по краям."""
        return str(value).strip().lower()

    @staticmethod
    def _extract_ids(schema: TableSchema, rows: list[dict[str, Any]]) -> list[int]:
        """Достаёт primary key родительских строк для загрузки дочерних данных."""
        pk_name = schema.pk.db_name
        ids: list[int] = []
        for row in rows:
            value = row.get(pk_name)
            if value is not None:
                ids.append(int(value))
        return ids

    @staticmethod
    def _validate_project_number(project_number: str) -> None:
        """Проверяет формат номера проекта: ##-F###."""
        if not PROJECT_NUMBER_RE.match(project_number):
            raise ValueError(
                f"Неверный номер проекта '{project_number}'. "
                f"Ожидается формат ##-F###, например 25-F123"
            )

def export_task_status_to_excel(
    db_path: str,
    output_path: str | Path = TASK_STATUS_EXPORT_PATH,
) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    headers = [
        "statusid",
        "status",
        "taskid",
        "task_type",
        "resultId",
        "dateTime",
    ]

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row

        rows = conn.execute(
            """
            SELECT
                statusid,
                status,
                taskid,
                task_type,
                resultId,
                dateTime
            FROM TaskStatus
            ORDER BY statusid
            """
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "TaskStatus"

    ws.append(headers)

    for row in rows:
        ws.append([row[header] for header in headers])

    last_row = max(1, len(rows) + 1)
    last_col = len(headers)

    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"

    table = Table(
        displayName="TaskStatus",
        ref=table_ref,
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    ws.freeze_panes = "A2"

    column_widths = {
        "A": 12,
        "B": 18,
        "C": 12,
        "D": 14,
        "E": 12,
        "F": 22,
    }

    for column_letter, width in column_widths.items():
        ws.column_dimensions[column_letter].width = width

    tmp_path = output_path.with_name(output_path.stem + "_tmp" + output_path.suffix)

    if tmp_path.exists():
        tmp_path.unlink()

    wb.save(tmp_path)

    tmp_path.replace(output_path)

    return str(output_path)

def run_save(
    workbook: str,
    db_path: str,
    study_codes: list[str] | None = None,
    task_status: str = TaskStatusRepository.COMPLETED_STATUS,
) -> SyncReport:
    progress(f"Сохранение исследований: {study_codes or 'ALL'}")
    service = StudySyncService(workbook=workbook, db_path=db_path)
    try:
        return service.save_to_db(
            get_studies(study_codes),
            task_status=task_status,
        )
    finally:
        service.close()

def run_status(
    db_path: str,
    *,
    taskid: int,
    task_type: str,
    status: str,
    result_id: int | None = None,
    date_time: Any = None,
) -> SyncReport:
    service = StudySyncService(workbook=None, db_path=db_path)
    try:
        return service.set_task_status(
            taskid=taskid,
            task_type=task_type,
            status=status,
            result_id=result_id,
            date_time=date_time,
        )
    finally:
        service.close()

def run_load(
    workbook: str,
    db_path: str,
    project_number: str,
    study_codes: list[str] | None = None,
) -> SyncReport:
    """Python API: загрузить SQLite -> Excel по номеру проекта."""
    service = StudySyncService(workbook=workbook, db_path=db_path)
    try:
        studies = get_studies(study_codes)
        return service.load_from_db(studies, project_number)
    finally:
        service.close()


def run_json(
    mode: str,
    workbook: str | None,
    db_path: str,
    study_codes: list[str] | None = None,
    project_number: str | None = None,
    task_status: str = TaskStatusRepository.COMPLETED_STATUS,
    taskid: int | None = None,
    task_type: str | None = None,
    result_id: int | None = None,
    date_time: str | None = None,
) -> str:
    """
    Возвращает результат JSON-строкой для VBA.
    """

    try:
        progress(f"Запуск режима: {mode}")
        if mode == "save":
            if not workbook:
                raise ValueError("Для режима save обязательно нужен --workbook")

            report = run_save(
                workbook,
                db_path,
                study_codes,
                task_status=task_status,
            )

        elif mode == "load":
            if not workbook:
                raise ValueError("Для режима load обязательно нужен --workbook")

            if not project_number:
                raise ValueError("Для режима load обязательно нужен --project")

            report = run_load(workbook, db_path, project_number, study_codes)

        elif mode == "status":
            if taskid is None:
                raise ValueError("Для режима status обязательно нужен --task-id")

            if not task_type:
                raise ValueError("Для режима status обязательно нужен --task-type")

            report = run_status(
                db_path,
                taskid=taskid,
                task_type=task_type,
                status=task_status,
                result_id=result_id,
                date_time=date_time,
            )

        else:
            raise ValueError("mode должен быть 'save', 'load' или 'status'")

        try:
            progress("Выгружаю TaskStatus в Excel-файл...")
            export_path = export_task_status_to_excel(db_path)
            progress(f"TaskStatus выгружена: {export_path}")
            report.message = report.message + f"\nTaskStatus выгружена в файл: {export_path}"

        except Exception as export_exc:
            traceback_text = traceback.format_exc()

            write_error_log(
                mode=mode,
                message=f"Синхронизация выполнена, но не удалось выгрузить TaskStatus: {export_exc}",
                error_type=type(export_exc).__name__,
                traceback_text=traceback_text,
            )

            report.message = (
                report.message
                + "\n\nВНИМАНИЕ: синхронизация выполнена успешно, "
                + f"но не удалось выгрузить TaskStatus в Excel. Подробности в логе: {ERROR_LOG_PATH}"
            )

        return json.dumps(
            asdict(report),
            ensure_ascii=False,
            indent=2,
        )

    except Exception as exc:
        traceback_text = traceback.format_exc()
        error_type = type(exc).__name__
        message = str(exc)

        write_error_log(
            mode=mode,
            message=message,
            error_type=error_type,
            traceback_text=traceback_text,
        )

        report = SyncReport(
            ok=False,
            message=(
                f"{error_type}: {message}\n\n"
                f"Полный traceback записан в лог:\n{ERROR_LOG_PATH}"
            ),
            mode=mode,
            tables=[],
            error_type=error_type,
            traceback=None,
        )

        return json.dumps(
            asdict(report),
            ensure_ascii=False,
            indent=2,
        )


def main(argv: list[str] | None = None) -> int:
    """
    CLI-вход для запуска из VBA через WScript.Shell.Exec.

    Сохранить Excel -> SQLite:
        python sync.py --mode save --workbook "C:\\data\\book.xlsx" --db "C:\\data\\pvt.sqlite" --studies AP GOR

    Загрузить SQLite -> Excel:
        python sync.py --mode load --workbook "C:\\data\\book.xlsx" --db "C:\\data\\pvt.sqlite" --project 25-F123 --studies AP GOR
    """
    parser = argparse.ArgumentParser(description="Синхронизация Excel-таблиц исследований и SQLite")
    parser.add_argument(
        "--mode",
        required=True,
        choices=["save", "load", "status"],
        help="save: Excel -> SQLite; load: SQLite -> Excel; status: только запись TaskStatus",
    )
    parser.add_argument(
        "--workbook",
        default=None,
        help="Полный путь к уже открытой книге Excel. Не нужен для mode=status",
    )
    parser.add_argument("--db", required=True, help="Путь к SQLite-файлу")
    parser.add_argument("--project", default=None, help="Номер проекта для режима load, например 25-F123")
    parser.add_argument("--studies", nargs="*", default=None, help="Коды исследований: OP AP GC SSF GOR BP REC EMV. Если не указано — все")
    parser.add_argument(
        "--status",
        default=TaskStatusRepository.COMPLETED_STATUS,
        help="Статус TaskStatus: Завершено, Валидировано, Отменено",
    )
    parser.add_argument(
        "--task-id",
        type=int,
        default=None,
        help="Номер задания для mode=status",
    )
    parser.add_argument(
        "--task-type",
        default=None,
        help="Тип исследования для mode=status: OP AP GC SSF GOR BP REC EMV",
    )
    parser.add_argument(
        "--result-id",
        type=int,
        default=None,
        help="resultId для mode=status, если есть",
    )
    parser.add_argument(
        "--date-time",
        default=None,
        help="Дата/время статуса для mode=status",
    )
    args = parser.parse_args(argv)

    json_result = run_json(
        mode=args.mode,
        workbook=args.workbook,
        db_path=args.db,
        study_codes=args.studies,
        project_number=args.project,
        task_status=args.status,
        taskid=args.task_id,
        task_type=args.task_type,
        result_id=args.result_id,
        date_time=args.date_time,
    )

    print(json_result)

    try:
        parsed = json.loads(json_result)
        return 0 if parsed.get("ok") else 1
    except Exception:
        return 1


if __name__ == "__main__":
    sys.exit(main())
