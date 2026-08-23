from __future__ import annotations

import argparse
import ctypes
import json
import logging
import os
import re
import shutil
import sys
import threading
import time
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterable
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import xlwings as xw
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =============================================================================
# Настройки
# =============================================================================

PROJECT_DIR = Path(__file__).resolve().parent

MIGRATION_YEAR_DEFAULT = 2026

BASE_PROJECTS_DIR = Path(r"L:\LRC\common_data\ФЛЮИДЫ\ГТИ\Работа")
DB_PATH = Path(r"L:\LRC\common_data\ФЛЮИДЫ\ГТИ\sqlite-excel\sqlite\results.db")

# Старая надстройка — именно тот путь, который уже зарегистрирован в Excel AddIns.
# В save-v22 используем её как legacy. В create-v23 новая надстройка копируется
# на этот же путь, чтобы Excel продолжал подключать тот же AddIn и не возникала
# проблема mapped/UNC путей.
INSTALLED_ADDIN_PATH = Path(
    r"L:\LRC\common_data\ФЛЮИДЫ\ГТИ\sqlite-excel\надстройка новая ribbon.xlam"
)

# Исходник новой v23-надстройки.
NEW_ADDIN_SOURCE_PATH = Path(
    r"L:\LRC\common_data\ФЛЮИДЫ\ГТИ\repository_SQlite_Project\excel\Addin\PVT_Addin.xlam"
)

VERSION_FILE = PROJECT_DIR / "project_version.json"
CLEAN_TEMPLATE_GLOB = "*_Форма_v*.xlsx"

TASKS_WORKBOOK = Path(r"L:\LRC\exchange\КСП Лайт\Журнал_заданий_флюиды.xlsx")

TASKS_PARENT_SHEET = "ГТИ"
TASKS_PARENT_TABLE = "Журнал_ГТИ"

REFRESH_TASKS_SOURCE_PARENT_SHEET = "ГТИ"
REFRESH_TASKS_SOURCE_PARENT_TABLE = "Журнал_ГТИ"
REFRESH_TASKS_SOURCE_CHILD_SHEET = "Смешение"
REFRESH_TASKS_SOURCE_CHILD_TABLE = "ЖУрнал_объединения"

REFRESH_TASKS_TARGET_PARENT_SHEET = "Task"
REFRESH_TASKS_TARGET_PARENT_TABLE = "Task"
REFRESH_TASKS_TARGET_CHILD_SHEET = "Task_mix"
REFRESH_TASKS_TARGET_CHILD_TABLE = "Task_mix"

PROJECT_CELL_SHEET = "OP"
PROJECT_CELL_ADDRESS = "B6"

LEGACY_SAVE_ALL_MACRO = "wrappers.Silent_Save_All_wrap"

LEGACY_REMOVE_PATHS_MACRO = "broken.FixExternalLinksInFormulasAndAssignedMacros_ActiveWorkbook"

LOAD_FORMS_FROM_TABLES_MACRO = "wrappers.Load_All_results_to_forms_Auto"
SET_ALL_BASELINES_MACRO = "modMetadata.SetAllStudyBaselinesAfterLoad"
AFTER_REFRESH_MACRO = "PowerQuery.silentRefresh_Project"

FORM_VERSION_PROPERTY = "PVTFormVersion"
ADDIN_VERSION_PROPERTY = "PVTAddinVersion"

FORM_VERSION_KEYS = ("form", "form_version", "PVTFormVersion")
ADDIN_VERSION_KEYS = ("addin", "add_in", "addin_version", "PVTAddinVersion")

STUDY_STATE_PROPERTIES: dict[str, str] = {
    "OP": "OPDataState",
    "OPOH": "OPOHDataState",
    "AP": "APDataState",
    "BP": "BPDataState",
    "GC": "GCDataState",
    "EMV": "EMVDataState",
    "GOR": "GORDataState",
    "SSF": "SSFDataState",
    "CCE": "CCEDataState",
    "CVD": "CVDDataState",
    "REC": "RECDataState",
    "JOIN": "JOINDataState",
}

LEGACY_SAVE_STUDIES = tuple(STUDY_STATE_PROPERTIES.keys())

LEGACY_SAVE_RESULT_VALUES = {
    "NO_RESULT",
    "FAILED",
    "NO_CHANGES",
    "CHANGED",
}

SAFE_V23_STATES = {"empty", "synced"}

UNPROTECT_PASSWORDS = ("1984", "9184", "")
PROTECT_PASSWORD = "1984"

ARCHIVE_DIR_NAME = "архив"

STATE_PATH = PROJECT_DIR / "migration_v22_v23_state.json"
STATUS_REPORT_PATH = PROJECT_DIR / "migration_v22_v23_status.xlsx"
LOG_PATH = PROJECT_DIR / "migration_v22_v23.log"

MODES = {"scan-v22", "save-v22", "create-v23"}

# GOR / SSF / CVD в старой форме могут запросить интерактивное подтверждение
# шифров полученных проб. Такие формы заранее выводим в MANUAL_REVIEW.
LEGACY_INTERACTIVE_STUDIES: dict[str, tuple[str, str]] = {
    "GOR": ("sampleCode", "resultIdGOR"),
    "SSF": ("sampleCode", "resultIdSSF"),
    "CVD": ("sampleCode", "resultIdCVD"),
}

logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    filemode="a",
    encoding="utf-8",
)
logger = logging.getLogger(__name__)


# =============================================================================
# Dataclasses
# =============================================================================

@dataclass
class ManualFormIssue:
    study: str
    form_index: int
    sample_code: str
    result_id: str
    reason: str


@dataclass
class V22Inspection:
    folder: str
    project: str
    v22_candidates: list[str] = field(default_factory=list)
    selected_v22: str = ""
    lock_present: bool = False
    newer_workbooks: list[str] = field(default_factory=list)
    manual_forms: list[ManualFormIssue] = field(default_factory=list)
    status: str = ""
    message: str = ""

    @property
    def ready_for_save(self) -> bool:
        return self.status in {"READY", "READY_AUTO_CODES"}


@dataclass
class ReleaseInfo:
    form_version: str
    addin_version: str
    form_major: int
    template_path: str
    addin_source_path: str


@dataclass
class UiOptions:
    mode: str
    year: int
    auto_sample_codes: bool = False
    projects_text: str = ""
    cancelled: bool = False


@dataclass
class RunSummary:
    ok: bool
    mode: str
    year: int
    processed: int
    succeeded: int
    manual_review: int
    blocked: int
    ignored: int
    message: str
    state_file: str
    status_report: str


# =============================================================================
# Общие helpers
# =============================================================================

def progress(message: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def timestamp_text() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_empty_or_zero(value: Any) -> bool:
    text = normalize_text(value)
    if text == "":
        return True
    try:
        return float(text.replace(",", ".")) == 0
    except ValueError:
        return False


def project_from_folder_name(folder_name: str) -> str:
    match = re.match(r"^(\d{2})-F(\d{3})(?:-|$)", folder_name, re.IGNORECASE)
    if not match:
        return ""
    return f"{match.group(1)}-F{match.group(2)}".upper()



def normalize_project_number(value: str) -> str:
    text = normalize_text(value).upper()
    match = re.fullmatch(r"(\d{2})-F(\d+)", text)

    if not match:
        raise ValueError(f"Некорректный номер проекта: {value}")

    year_part, number_part = match.groups()
    return f"{year_part}-F{int(number_part):03d}"


def expand_project_range(start: str, end: str) -> list[str]:
    start_norm = normalize_project_number(start)
    end_norm = normalize_project_number(end)

    m1 = re.fullmatch(r"(\d{2}-F)(\d{3})", start_norm)
    m2 = re.fullmatch(r"(\d{2}-F)(\d{3})", end_norm)

    if not m1 or not m2:
        raise ValueError(f"Некорректный диапазон проектов: {start}...{end}")

    prefix1, number1 = m1.groups()
    prefix2, number2 = m2.groups()

    if prefix1 != prefix2:
        raise ValueError(
            f"Диапазон должен быть внутри одного года: {start}...{end}"
        )

    first = int(number1)
    last = int(number2)

    if last < first:
        raise ValueError(
            f"Конец диапазона меньше начала: {start}...{end}"
        )

    return [f"{prefix1}{number:03d}" for number in range(first, last + 1)]


def parse_projects_input(text: str, *, year: int) -> list[str]:
    """
    Разбирает выбор проектов для save-v22.

    Примеры:
        26-F001
        26-F001...26-F050
        26-F001, 26-F005, 26-F010...26-F020

    Пустая строка означает весь выбранный год.
    """
    raw = normalize_text(text)

    if not raw:
        return []

    raw = re.sub(r"\s*\.{2,3}\s*", "...", raw)
    parts = re.split(r"[,;\s]+", raw)

    result: list[str] = []

    for part in parts:
        part = part.strip()
        if not part:
            continue

        if "..." in part:
            start, end = part.split("...", 1)
            result.extend(expand_project_range(start, end))
        else:
            result.append(normalize_project_number(part))

    expected_prefix = f"{year % 100:02d}-F"
    wrong_year = [
        project
        for project in result
        if not project.startswith(expected_prefix)
    ]

    if wrong_year:
        raise ValueError(
            f"Для года {year} указаны проекты другого года: "
            + ", ".join(wrong_year)
        )

    seen: set[str] = set()
    unique: list[str] = []

    for project in result:
        if project in seen:
            continue
        seen.add(project)
        unique.append(project)

    return unique


def year_folder(year: int) -> Path:
    return BASE_PROJECTS_DIR / str(year)


def list_project_folders(year: int) -> list[Path]:
    root = year_folder(year)
    if not root.exists():
        raise FileNotFoundError(f"Папка года не найдена: {root}")

    result = [
        path
        for path in root.iterdir()
        if path.is_dir() and path.name.casefold() != ARCHIVE_DIR_NAME.casefold()
    ]
    return sorted(result, key=lambda path: path.name.casefold())


def is_ignored_workbook_name(name: str) -> bool:
    lower = name.casefold()
    return (
        lower.startswith("~$")
        or "__new__" in lower
        or "_backup_" in lower
    )


def looks_like_v22_workbook(path: Path) -> bool:
    if not path.is_file() or path.suffix.casefold() != ".xlsx":
        return False
    if is_ignored_workbook_name(path.name):
        return False
    return re.search(r"_v22(?:\D|$)", path.stem, re.IGNORECASE) is not None


def looks_like_v23_or_newer_workbook(path: Path) -> bool:
    if not path.is_file() or path.suffix.casefold() != ".xlsx":
        return False
    if is_ignored_workbook_name(path.name):
        return False

    match = re.search(r"_v(\d+)", path.stem, re.IGNORECASE)
    if not match:
        return False

    return int(match.group(1)) >= 23


def find_v22_workbooks(folder: Path) -> list[Path]:
    return sorted(
        [path for path in folder.iterdir() if looks_like_v22_workbook(path)],
        key=lambda path: path.name.casefold(),
    )


def find_v23_or_newer_workbooks(folder: Path) -> list[Path]:
    return sorted(
        [path for path in folder.iterdir() if looks_like_v23_or_newer_workbook(path)],
        key=lambda path: path.name.casefold(),
    )


def excel_lock_present(workbook_path: Path) -> bool:
    return workbook_path.with_name("~$" + workbook_path.name).exists()


def safe_unlink(path: Path) -> None:
    if not path.exists():
        return
    try:
        path.unlink()
    except Exception:
        logger.exception("Не удалось удалить файл %s", path)


# =============================================================================
# State JSON
# =============================================================================

def empty_state() -> dict[str, Any]:
    return {
        "schema_version": 1,
        "migration": "v22_to_v23",
        "updated_at": now_text(),
        "addin_switch": {},
        "projects": {},
    }


def load_state() -> dict[str, Any]:
    if not STATE_PATH.exists():
        return empty_state()

    with STATE_PATH.open("r", encoding="utf-8-sig") as file:
        data = json.load(file)

    if not isinstance(data, dict):
        raise TypeError(f"Некорректный state-файл: {STATE_PATH}")

    data.setdefault("schema_version", 1)
    data.setdefault("migration", "v22_to_v23")
    data.setdefault("updated_at", now_text())
    data.setdefault("addin_switch", {})
    data.setdefault("projects", {})
    return data


def save_state(state: dict[str, Any]) -> None:
    state["updated_at"] = now_text()

    temp_path = STATE_PATH.with_suffix(".json.tmp")
    with temp_path.open("w", encoding="utf-8") as file:
        json.dump(state, file, ensure_ascii=False, indent=2)

    os.replace(temp_path, STATE_PATH)


def state_key(folder: Path) -> str:
    # Не вызываем resolve(), чтобы не превращать mapped path в UNC.
    return str(folder)


def project_state_entry(state: dict[str, Any], folder: Path) -> dict[str, Any]:
    key = state_key(folder)
    projects = state.setdefault("projects", {})

    entry = projects.setdefault(
        key,
        {
            "folder": str(folder),
            "project": project_from_folder_name(folder.name),
            "year": int(folder.parent.name) if folder.parent.name.isdigit() else None,
            "v22_workbook": "",
            "scan": {},
            "save_v22": {},
            "create_v23": {},
        },
    )

    entry["folder"] = str(folder)
    entry["project"] = project_from_folder_name(folder.name)
    if folder.parent.name.isdigit():
        entry["year"] = int(folder.parent.name)

    entry.setdefault("scan", {})
    entry.setdefault("save_v22", {})
    entry.setdefault("create_v23", {})
    return entry


# =============================================================================
# Чтение Office custom metadata без Excel
# =============================================================================

def read_office_custom_properties(path: Path) -> dict[str, str]:
    if not path.exists():
        raise FileNotFoundError(path)

    try:
        with zipfile.ZipFile(path, "r") as archive:
            try:
                xml_data = archive.read("docProps/custom.xml")
            except KeyError:
                return {}
    except zipfile.BadZipFile as exc:
        raise RuntimeError(f"Файл не является корректным Office-файлом: {path}") from exc

    root = ET.fromstring(xml_data)
    result: dict[str, str] = {}

    for prop in root:
        name = prop.attrib.get("name", "").strip()
        if not name:
            continue

        children = list(prop)
        result[name] = (
            ""
            if not children or children[0].text is None
            else str(children[0].text).strip()
        )

    return result


def parse_version_tuple(version: str) -> tuple[int, ...]:
    text = str(version).strip()
    if not re.fullmatch(r"\d+(?:\.\d+)*", text):
        raise ValueError(f"Некорректная версия: {version}")
    return tuple(int(part) for part in text.split("."))


def version_major(version: str) -> int:
    return parse_version_tuple(version)[0]


def load_version_manifest() -> dict[str, Any]:
    if not VERSION_FILE.exists():
        raise FileNotFoundError(f"Не найден project_version.json: {VERSION_FILE}")

    with VERSION_FILE.open("r", encoding="utf-8-sig") as file:
        data = json.load(file)

    if not isinstance(data, dict):
        raise TypeError("project_version.json должен содержать JSON-объект")

    return data


def extract_manifest_version(
    manifest: dict[str, Any],
    keys: Iterable[str],
    caption: str,
) -> str:
    for key in keys:
        if key not in manifest:
            continue

        value = manifest[key]

        if isinstance(value, str):
            version = value.strip()
        elif isinstance(value, dict):
            raw = value.get("version")
            version = "" if raw is None else str(raw).strip()
        else:
            version = ""

        if version:
            parse_version_tuple(version)
            return version

    raise ValueError(
        f"В project_version.json не найдена версия компонента '{caption}'. "
        f"Проверены ключи: {', '.join(keys)}"
    )


def locate_current_template(expected_version: str) -> Path:
    candidates = [
        path
        for path in PROJECT_DIR.glob(CLEAN_TEMPLATE_GLOB)
        if path.is_file() and not path.name.startswith("~$")
    ]

    matching: list[Path] = []

    for path in candidates:
        try:
            props = read_office_custom_properties(path)
        except Exception as exc:
            logger.warning("Не удалось прочитать metadata шаблона %s: %s", path, exc)
            continue

        if props.get(FORM_VERSION_PROPERTY, "").strip() == expected_version:
            matching.append(path)

    if not matching:
        raise FileNotFoundError(
            f"Не найден шаблон с {FORM_VERSION_PROPERTY}={expected_version}. "
            f"Маска: {PROJECT_DIR / CLEAN_TEMPLATE_GLOB}"
        )

    if len(matching) > 1:
        raise RuntimeError(
            f"Найдено несколько шаблонов версии {expected_version}: "
            + "; ".join(str(path) for path in matching)
        )

    return matching[0]


def validate_new_addin_source(expected_version: str) -> None:
    if not NEW_ADDIN_SOURCE_PATH.exists():
        raise FileNotFoundError(
            f"Исходник новой надстройки не найден: {NEW_ADDIN_SOURCE_PATH}"
        )

    props = read_office_custom_properties(NEW_ADDIN_SOURCE_PATH)
    actual = props.get(ADDIN_VERSION_PROPERTY, "").strip()

    if actual != expected_version:
        raise RuntimeError(
            "Версия новой надстройки не соответствует project_version.json. "
            f"Ожидается {expected_version}, найдено {actual or '<нет>'}. "
            f"Файл: {NEW_ADDIN_SOURCE_PATH}"
        )


def preflight_v23_release() -> ReleaseInfo:
    progress("Проверяю project_version.json, шаблон v23 и новую надстройку")

    manifest = load_version_manifest()
    form_version = extract_manifest_version(manifest, FORM_VERSION_KEYS, "form")
    addin_version = extract_manifest_version(manifest, ADDIN_VERSION_KEYS, "addin")

    if version_major(form_version) != 23:
        raise RuntimeError(
            f"Мигратор v22->v23 ожидает form major 23, в JSON указано {form_version}"
        )

    template_path = locate_current_template(form_version)
    validate_new_addin_source(addin_version)

    progress(
        f"Preflight v23 OK: form={form_version}, addin={addin_version}, "
        f"template={template_path.name}"
    )

    return ReleaseInfo(
        form_version=form_version,
        addin_version=addin_version,
        form_major=23,
        template_path=str(template_path),
        addin_source_path=str(NEW_ADDIN_SOURCE_PATH),
    )


# =============================================================================
# Анализ старых v22-книг без Excel
# =============================================================================

def defined_name_lookup(workbook: openpyxl.Workbook) -> dict[str, Any]:
    result: dict[str, Any] = {}

    for name, defined_name in workbook.defined_names.items():
        result[str(name).casefold()] = defined_name

    # Обычно имена формы workbook-level, но подхватываем и local names.
    for worksheet in workbook.worksheets:
        try:
            items = worksheet.defined_names.items()
        except AttributeError:
            continue

        for name, defined_name in items:
            result.setdefault(str(name).casefold(), defined_name)

    return result


def read_defined_name_value(
    workbook: openpyxl.Workbook,
    names: dict[str, Any],
    range_name: str,
) -> tuple[bool, Any]:
    defined_name = names.get(range_name.casefold())
    if defined_name is None:
        return False, None

    try:
        destinations = list(defined_name.destinations)
    except Exception:
        return True, None

    if not destinations:
        return True, None

    sheet_name, cell_ref = destinations[0]

    # Для наших sampleCode/resultId ожидается одиночная ячейка.
    if ":" in cell_ref:
        cell_ref = cell_ref.split(":", 1)[0]

    cell_ref = cell_ref.replace("$", "")

    try:
        return True, workbook[sheet_name][cell_ref].value
    except Exception:
        return True, None


def detect_legacy_manual_forms(workbook_path: Path) -> list[ManualFormIssue]:
    issues: list[ManualFormIssue] = []

    workbook = openpyxl.load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )

    try:
        names = defined_name_lookup(workbook)

        for study, (sample_field, result_field) in LEGACY_INTERACTIVE_STUDIES.items():
            pattern = re.compile(
                rf"^{re.escape(study)}_(\d+)_{re.escape(sample_field)}$",
                re.IGNORECASE,
            )

            form_indexes: set[int] = set()

            for name in names:
                match = pattern.fullmatch(name)
                if match:
                    form_indexes.add(int(match.group(1)))

            for form_index in sorted(form_indexes):
                sample_name = f"{study}_{form_index}_{sample_field}"
                result_name = f"{study}_{form_index}_{result_field}"

                _sample_exists, sample_value = read_defined_name_value(
                    workbook, names, sample_name
                )
                result_exists, result_value = read_defined_name_value(
                    workbook, names, result_name
                )

                sample_text = normalize_text(sample_value)
                result_text = normalize_text(result_value)

                if not sample_text:
                    continue

                # Консервативная проверка:
                # заполненная форма + нет реального resultId => старый Save может
                # открыть окно генерации/подтверждения полученных шифров.
                if not result_exists or is_empty_or_zero(result_value):
                    issues.append(
                        ManualFormIssue(
                            study=study,
                            form_index=form_index,
                            sample_code=sample_text,
                            result_id=result_text,
                            reason=(
                                f"{study}_{form_index}: заполнен sampleCode, "
                                f"но {result_field} отсутствует/пустой. "
                                "Сохранить вручную и подтвердить шифры полученных проб."
                            ),
                        )
                    )

    finally:
        workbook.close()

    return issues


def inspect_v22_folder(
    folder: Path,
    *,
    auto_sample_codes: bool = False,
) -> V22Inspection | None:
    project = project_from_folder_name(folder.name)

    v22_candidates = find_v22_workbooks(folder)

    # v21 и любые папки без v22 для этой миграции полностью игнорируются.
    if not v22_candidates:
        return None

    inspection = V22Inspection(
        folder=str(folder),
        project=project,
        v22_candidates=[str(path) for path in v22_candidates],
    )

    newer = find_v23_or_newer_workbooks(folder)
    inspection.newer_workbooks = [str(path) for path in newer]

    if newer:
        inspection.status = "BLOCKED_V23_OR_NEWER_PRESENT"
        inspection.message = (
            "В папке одновременно с v22 уже есть рабочая книга v23 или новее: "
            + ", ".join(path.name for path in newer)
        )
        return inspection

    if len(v22_candidates) > 1:
        inspection.status = "BLOCKED_MULTIPLE_V22"
        inspection.message = (
            "Найдено несколько v22-книг. Скрипт их не открывает: "
            + ", ".join(path.name for path in v22_candidates)
        )
        return inspection

    workbook_path = v22_candidates[0]
    inspection.selected_v22 = str(workbook_path)
    inspection.lock_present = excel_lock_present(workbook_path)

    if inspection.lock_present:
        inspection.status = "BLOCKED_WORKBOOK_OPEN"
        inspection.message = "Обнаружен Excel lock-файл; книга, вероятно, открыта."
        return inspection

    try:
        inspection.manual_forms = detect_legacy_manual_forms(workbook_path)
    except Exception as exc:
        inspection.status = "MANUAL_REVIEW_SCAN_ERROR"
        inspection.message = f"Не удалось проверить GOR/SSF/CVD: {exc}"
        return inspection

    if inspection.manual_forms:
        if auto_sample_codes:
            inspection.status = "READY_AUTO_CODES"
            inspection.message = (
                "Есть GOR/SSF/CVD формы без resultId. Разрешена автоматическая "
                "генерация/подтверждение шифров через silent VBA."
            )
        else:
            inspection.status = "MANUAL_REVIEW"
            inspection.message = (
                "Есть GOR/SSF/CVD формы без resultId. Автоматическая генерация "
                "шифров отключена — требуется ручная проверка."
            )
        return inspection

    inspection.status = "READY"
    inspection.message = "Готово к автоматическому Save_All v22."
    return inspection


def inspect_all_v22(
    year: int,
    *,
    auto_sample_codes: bool = False,
) -> list[V22Inspection]:
    folders = list_project_folders(year)
    inspections: list[V22Inspection] = []

    progress(f"Сканирую папки {year}: {len(folders)}")

    for index, folder in enumerate(folders, start=1):
        if index == 1 or index % 25 == 0 or index == len(folders):
            progress(f"SCAN v22: {index}/{len(folders)}")

        inspection = inspect_v22_folder(
            folder,
            auto_sample_codes=auto_sample_codes,
        )

        # Нет v22 => полностью игнорируем, в том числе v21-only.
        if inspection is None:
            continue

        inspections.append(inspection)

    return inspections


def write_inspections_to_state(
    state: dict[str, Any],
    inspections: list[V22Inspection],
) -> None:
    for inspection in inspections:
        folder = Path(inspection.folder)
        entry = project_state_entry(state, folder)

        entry["v22_workbook"] = inspection.selected_v22
        entry["scan"] = {
            "at": now_text(),
            "status": inspection.status,
            "message": inspection.message,
            "v22_candidates": inspection.v22_candidates,
            "newer_workbooks": inspection.newer_workbooks,
            "lock_present": inspection.lock_present,
            "manual_forms": [asdict(item) for item in inspection.manual_forms],
            "auto_sample_codes": inspection.status == "READY_AUTO_CODES",
        }


# =============================================================================
# Отчёт Excel
# =============================================================================

def autosize_sheet(ws, *, max_width: int = 80) -> None:
    for column_cells in ws.columns:
        width = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(
                width,
                max((len(line) for line in value.splitlines()), default=0),
            )

        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def add_excel_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_status_report(state: dict[str, Any], year: int) -> Path:
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Projects"

    columns = (
        "Project",
        "Folder",
        "V22Workbook",
        "ScanStatus",
        "ScanMessage",
        "SaveV22Status",
        "SaveV22Message",

        "Save_OP",
        "Save_OPOH",
        "Save_AP",
        "Save_BP",
        "Save_GC",
        "Save_EMV",
        "Save_GOR",
        "Save_SSF",
        "Save_CCE",
        "Save_CVD",
        "Save_REC",
        "Save_JOIN",

        "CreateV23Status",
        "CreateV23Message",
        "V23Workbook",
        "V22Backup",
    )
    ws.append(columns)

    project_entries = []

    for entry in state.get("projects", {}).values():
        if entry.get("year") != year:
            continue

        # В отчёт миграции попадают только записи, где когда-либо был обнаружен v22.
        scan = entry.get("scan", {})
        v22_workbook = entry.get("v22_workbook", "")
        if not v22_workbook and not scan.get("v22_candidates"):
            continue

        project_entries.append(entry)

    project_entries.sort(
        key=lambda entry: str(entry.get("folder", "")).casefold()
    )

    for entry in project_entries:
        scan = entry.get("scan", {})
        save_info = entry.get("save_v22", {})
        create_info = entry.get("create_v23", {})
        study_results = save_info.get("studies", {})
        ws.append(
            (
                entry.get("project", ""),
                entry.get("folder", ""),
                entry.get("v22_workbook", ""),
                scan.get("status", ""),
                scan.get("message", ""),
                save_info.get("status", ""),
                save_info.get("message", ""),

                study_results.get("OP", ""),
                study_results.get("OPOH", ""),
                study_results.get("AP", ""),
                study_results.get("BP", ""),
                study_results.get("GC", ""),
                study_results.get("EMV", ""),
                study_results.get("GOR", ""),
                study_results.get("SSF", ""),
                study_results.get("CCE", ""),
                study_results.get("CVD", ""),
                study_results.get("REC", ""),
                study_results.get("JOIN", ""),

                create_info.get("status", ""),
                create_info.get("message", ""),
                create_info.get("v23_workbook", ""),
                create_info.get("v22_backup", ""),
            )
        )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    add_excel_table(ws, "MigrationProjects")
    autosize_sheet(ws)

    manual_ws = workbook.create_sheet("ManualForms")
    manual_ws.append(
        (
            "Project",
            "Folder",
            "Study",
            "FormIndex",
            "SampleCode",
            "ResultId",
            "Reason",
        )
    )

    for entry in project_entries:
        scan = entry.get("scan", {})
        for issue in scan.get("manual_forms", []):
            manual_ws.append(
                (
                    entry.get("project", ""),
                    entry.get("folder", ""),
                    issue.get("study", ""),
                    issue.get("form_index", ""),
                    issue.get("sample_code", ""),
                    issue.get("result_id", ""),
                    issue.get("reason", ""),
                )
            )

    for cell in manual_ws[1]:
        cell.font = Font(bold=True)

    manual_ws.freeze_panes = "A2"
    add_excel_table(manual_ws, "ManualForms")
    autosize_sheet(manual_ws)

    info_ws = workbook.create_sheet("Info")
    info_ws.append(("Parameter", "Value"))
    info_ws.append(("Migration", "v22 -> v23"))
    info_ws.append(("Year", year))
    info_ws.append(("GeneratedAt", now_text()))
    info_ws.append(("StateFile", str(STATE_PATH)))
    info_ws.append(("InstalledAddinPath", str(INSTALLED_ADDIN_PATH)))
    info_ws.append(("NewAddinSource", str(NEW_ADDIN_SOURCE_PATH)))

    addin_switch = state.get("addin_switch", {})
    info_ws.append(("AddinSwitchStatus", addin_switch.get("status", "")))
    info_ws.append(("LegacyAddinBackup", addin_switch.get("legacy_backup", "")))
    info_ws.append(("CurrentAddinVersion", addin_switch.get("current_version", "")))

    for cell in info_ws[1]:
        cell.font = Font(bold=True)

    add_excel_table(info_ws, "MigrationInfo")
    autosize_sheet(info_ws)

    workbook.save(STATUS_REPORT_PATH)
    workbook.close()
    return STATUS_REPORT_PATH


# =============================================================================
# Excel dialog watchdog для save-v22
# =============================================================================

class ExcelDialogWatchdog:
    """
    Закрывает неожиданные стандартные Excel/VBA диалоги (#32770), чтобы ночной
    save-v22 не завис навсегда. Мы НЕ нажимаем OK автоматически: окно закрывается,
    а проект помечается MANUAL_REVIEW.
    """

    WM_CLOSE = 0x0010

    def __init__(self, excel_hwnd: int, *, interval: float = 0.25) -> None:
        self.excel_hwnd = int(excel_hwnd)
        self.interval = interval
        self.detected: list[str] = []
        self._stop_event = threading.Event()
        self._thread: threading.Thread | None = None

        self.user32 = ctypes.windll.user32
        self.excel_pid = self._pid_from_hwnd(self.excel_hwnd)

    def _pid_from_hwnd(self, hwnd: int) -> int:
        pid = ctypes.c_ulong()
        self.user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
        return int(pid.value)

    def _window_text(self, hwnd: int) -> str:
        length = self.user32.GetWindowTextLengthW(hwnd)
        if length <= 0:
            return ""
        buffer = ctypes.create_unicode_buffer(length + 1)
        self.user32.GetWindowTextW(hwnd, buffer, length + 1)
        return buffer.value

    def _class_name(self, hwnd: int) -> str:
        buffer = ctypes.create_unicode_buffer(256)
        self.user32.GetClassNameW(hwnd, buffer, 256)
        return buffer.value

    def _enum_dialogs(self) -> list[int]:
        windows: list[int] = []

        @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)
        def callback(hwnd: int, _lparam: int) -> bool:
            if not self.user32.IsWindowVisible(hwnd):
                return True

            pid = ctypes.c_ulong()
            self.user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))

            if int(pid.value) != self.excel_pid:
                return True

            if self._class_name(hwnd) != "#32770":
                return True

            windows.append(int(hwnd))
            return True

        self.user32.EnumWindows(callback, 0)
        return windows

    def _run(self) -> None:
        seen: set[int] = set()

        while not self._stop_event.is_set():
            try:
                for hwnd in self._enum_dialogs():
                    if hwnd in seen:
                        continue

                    seen.add(hwnd)
                    title = self._window_text(hwnd) or "<без заголовка>"
                    self.detected.append(title)

                    logger.warning(
                        "Обнаружен неожиданный Excel/VBA диалог '%s'; закрываю его "
                        "и помечаю проект MANUAL_REVIEW",
                        title,
                    )

                    self.user32.PostMessageW(hwnd, self.WM_CLOSE, 0, 0)

            except Exception:
                logger.exception("Ошибка ExcelDialogWatchdog")

            self._stop_event.wait(self.interval)

    def start(self) -> None:
        self._thread = threading.Thread(
            target=self._run,
            name="ExcelDialogWatchdog",
            daemon=True,
        )
        self._thread.start()

    def stop(self) -> None:
        self._stop_event.set()
        if self._thread is not None:
            self._thread.join(timeout=2)


# =============================================================================
# Excel / AddIn helpers
# =============================================================================

def disable_excel_prompts(app: xw.App) -> None:
    app.display_alerts = False
    app.api.DisplayAlerts = False
    app.api.AskToUpdateLinks = False
    app.api.AlertBeforeOverwriting = False
    app.api.EnableEvents = False

def load_installed_addin_by_registered_name(
    app: xw.App,
    configured_path: Path,
) -> Any:

    addin_name = configured_path.name.lower()

    for i in range(1, app.api.AddIns.Count + 1):

        addin = app.api.AddIns.Item(i)

        if str(addin.Name).lower() != addin_name:
            continue

        progress(f"AddIn Name: {addin.Name}")
        progress(f"AddIn FullName: {addin.FullName}")
        progress(f"AddIn Installed до: {addin.Installed}")

        if addin.Installed:
            addin.Installed = False

        addin.Installed = True

        progress(f"AddIn Installed после: {addin.Installed}")

        return addin

    raise RuntimeError(
        f"Установленная надстройка "
        f"'{configured_path.name}' не найдена"
    )

def run_addin_macro(
    *,
    app: xw.App,
    workbook: xw.Book,
    addin_name: str,
    macro_name: str,
    args: tuple[Any, ...] = (),
) -> Any:
    disable_excel_prompts(app)
    workbook.activate()

    macro_ref = f"'{addin_name}'!{macro_name}"

    if args:
        result = app.api.Run(macro_ref, *args)
    else:
        result = app.api.Run(macro_ref)

    disable_excel_prompts(app)
    return result


def unprotect_all_sheets(book: xw.Book) -> None:
    for sheet in book.sheets:
        for password in UNPROTECT_PASSWORDS:
            try:
                sheet.api.Unprotect(Password=password)
                break
            except Exception:
                pass


def protect_all_sheets(book: xw.Book) -> None:
    for sheet in book.sheets:
        try:
            sheet.api.Protect(
                Password=PROTECT_PASSWORD,
                DrawingObjects=True,
                Contents=True,
                Scenarios=True,
                UserInterfaceOnly=True,
                AllowFiltering=True,
            )
        except Exception as exc:
            logger.warning("Не удалось защитить лист %s: %s", sheet.name, exc)


# =============================================================================
# Режим scan-v22
# =============================================================================

def run_scan_v22(
    year: int,
    *,
    auto_sample_codes: bool = False,
) -> RunSummary:
    state = load_state()
    inspections = inspect_all_v22(
        year,
        auto_sample_codes=auto_sample_codes,
    )

    write_inspections_to_state(state, inspections)
    save_state(state)
    report_path = write_status_report(state, year)

    ready = sum(item.ready_for_save for item in inspections)
    manual = sum(item.status.startswith("MANUAL_REVIEW") for item in inspections)
    blocked = len(inspections) - ready - manual

    return RunSummary(
        ok=True,
        mode="scan-v22",
        year=year,
        processed=len(inspections),
        succeeded=ready,
        manual_review=manual,
        blocked=blocked,
        ignored=0,
        message=(
            f"SCAN v22 завершён. Найдено v22-проектов: {len(inspections)}; "
            f"READY: {ready}; MANUAL_REVIEW: {manual}; BLOCKED: {blocked}; "
            f"AUTO_CODES: {'ON' if auto_sample_codes else 'OFF'}"
        ),
        state_file=str(STATE_PATH),
        status_report=str(report_path),
    )


# =============================================================================
# Режим save-v22
# =============================================================================

def validate_legacy_addin_before_save() -> None:
    if not INSTALLED_ADDIN_PATH.exists():
        raise FileNotFoundError(
            f"Legacy-надстройка не найдена: {INSTALLED_ADDIN_PATH}"
        )

    props = read_office_custom_properties(INSTALLED_ADDIN_PATH)

    # У v22 этой metadata нет. Если появилась PVTAddinVersion — скорее всего
    # create-v23 уже заменил установленный файл новой надстройкой.
    if props.get(ADDIN_VERSION_PROPERTY, "").strip():
        raise RuntimeError(
            "По пути установленной legacy-надстройки уже находится файл с "
            f"{ADDIN_VERSION_PROPERTY}. Save-v22 после переключения на v23 "
            "запускать нельзя без ручного восстановления старой надстройки."
        )


def legacy_macro_failed(result: Any) -> bool:
    """
    Если wrapper возвращает macroResult, стандартный первый элемент enum
    MACRO_FAILED обычно равен 0. Если wrapper является Sub и возвращает None,
    успешность определяется завершением COM-вызова и отсутствием диалогов.
    """
    if result is None:
        return False

    if isinstance(result, bool):
        return result is False

    if isinstance(result, (int, float)):
        return int(result) == 0

    text = normalize_text(result).casefold()
    return text in {"0", "false", "failed", "macro_failed"}

def parse_legacy_save_report(result: Any) -> dict[str, str]:
    """
    Разбирает VBA-ответ вида:

        OP=CHANGED
        OPOH=FAILED
        AP=NO_CHANGES
        ...

    Возвращает:
        {
            "OP": "CHANGED",
            "OPOH": "FAILED",
            ...
        }
    """

    text = normalize_text(result)

    if not text:
        raise RuntimeError(
            "Silent_Save_All_wrap не вернул отчёт по исследованиям."
        )

    if text.upper().startswith("FATAL="):
        raise RuntimeError(
            f"VBA завершил покомпонентное сохранение с FATAL: {text}"
        )

    study_results: dict[str, str] = {}

    for raw_line in text.splitlines():
        line = raw_line.strip()

        if not line:
            continue

        if "=" not in line:
            raise RuntimeError(
                f"Неизвестная строка в VBA-отчёте: {line!r}"
            )

        study, save_result = line.split("=", 1)

        study = study.strip().upper()
        save_result = save_result.strip().upper()

        if study not in LEGACY_SAVE_STUDIES:
            raise RuntimeError(
                f"VBA вернул неизвестное исследование: {study!r}"
            )

        if save_result not in LEGACY_SAVE_RESULT_VALUES:
            raise RuntimeError(
                f"VBA вернул неизвестный результат для {study}: "
                f"{save_result!r}"
            )

        if study in study_results:
            raise RuntimeError(
                f"VBA дважды вернул результат для исследования {study}."
            )

        study_results[study] = save_result

    missing = [
        study
        for study in LEGACY_SAVE_STUDIES
        if study not in study_results
    ]

    if missing:
        raise RuntimeError(
            "VBA не вернул результаты для исследований: "
            + ", ".join(missing)
        )

    return study_results

def classify_legacy_save_report(
    study_results: dict[str, str],
) -> tuple[str, str]:
    """
    Определяет общий статус save-v22.

    Особое правило OP/OPOH:
    проект начинается либо с OP, либо с OPOH.

    Если OP реально сохранился (CHANGED), ошибку OPOH игнорируем.
    Если OPOH реально сохранился (CHANGED), ошибку OP игнорируем.

    Ошибка любого другого исследования блокирует create-v23.
    """

    ignored_failures: list[str] = []

    op_result = study_results["OP"]
    opoh_result = study_results["OPOH"]

    # OP-проект: OPOH нас не интересует.
    if op_result == "CHANGED" and opoh_result in {
        "FAILED",
        "NO_RESULT",
    }:
        ignored_failures.append("OPOH")

    # OPOH-проект: OP нас не интересует.
    if opoh_result == "CHANGED" and op_result in {
        "FAILED",
        "NO_RESULT",
    }:
        ignored_failures.append("OP")

    blocking_failures = [
        study
        for study, result in study_results.items()
        if result in {"FAILED", "NO_RESULT"}
        and study not in ignored_failures
    ]

    changed = [
        study
        for study, result in study_results.items()
        if result == "CHANGED"
    ]

    no_changes = [
        study
        for study, result in study_results.items()
        if result == "NO_CHANGES"
    ]

    message_parts = []

    if changed:
        message_parts.append(
            "Сохранено: " + ", ".join(changed)
        )

    if no_changes:
        message_parts.append(
            "Без изменений: " + ", ".join(no_changes)
        )

    if ignored_failures:
        message_parts.append(
            "Игнорируются ошибки неактивной OP/OPOH-ветки: "
            + ", ".join(
                f"{study}={study_results[study]}"
                for study in ignored_failures
            )
        )

    if blocking_failures:
        message_parts.append(
            "Не удалось обработать: "
            + ", ".join(
                f"{study}={study_results[study]}"
                for study in blocking_failures
            )
        )

        return "PARTIAL", "; ".join(message_parts)

    return "OK", "; ".join(message_parts)

def save_one_v22_workbook(
    *,
    app: xw.App,
    addin_name: str,
    workbook_path: Path,
) -> tuple[str, str, dict[str, str]]:
    workbook: xw.Book | None = None

    try:
        disable_excel_prompts(app)

        workbook = app.books.open(
            str(workbook_path),
            update_links=False,
            read_only=False,
            ignore_read_only_recommended=True,
        )

        disable_excel_prompts(app)
        workbook.activate()

        # 1. До любых VBA-макросов разблокируем ВСЕ листы книги.
        unprotect_all_sheets(workbook)

        # 2. До Save_All обязательно убираем внешние/UNC пути из формул.
        remove_paths_macro = normalize_text(LEGACY_REMOVE_PATHS_MACRO)

        if not remove_paths_macro:
            raise RuntimeError(
                "Не задан LEGACY_REMOVE_PATHS_MACRO. "
                "Укажи точное имя существующего VBA-макроса очистки путей "
                "в настройках скрипта."
            )

        progress(
            f"Убираю внешние пути из формул VBA-макросом: "
            f"{remove_paths_macro}"
        )

        run_addin_macro(
            app=app,
            workbook=workbook,
            addin_name=addin_name,
            macro_name=remove_paths_macro,
        )

        disable_excel_prompts(app)
        workbook.activate()

        # 3. Только теперь запускаем покомпонентное silent-сохранение.
        watchdog = ExcelDialogWatchdog(int(app.api.Hwnd))
        watchdog.start()

        try:
            result = run_addin_macro(
                app=app,
                workbook=workbook,
                addin_name=addin_name,
                macro_name=LEGACY_SAVE_ALL_MACRO,
            )
        finally:
            time.sleep(0.2)
            watchdog.stop()

        if watchdog.detected:
            return (
                "MANUAL_REVIEW",
                "Во время Save_All появилось интерактивное окно Excel/VBA: "
                + "; ".join(watchdog.detected),
                {},
            )

        study_results = parse_legacy_save_report(result)

        status, message = classify_legacy_save_report(
            study_results
        )

        disable_excel_prompts(app)
        workbook.save()

        return status, message, study_results

    except Exception as exc:
        logger.exception(
            "Ошибка save-v22 для %s",
            workbook_path,
        )
        return (
            "FAILED",
            f"{type(exc).__name__}: {exc}",
            {},
        )

    finally:
        if workbook is not None:
            try:
                disable_excel_prompts(app)
                workbook.close()
            except Exception:
                logger.exception(
                    "Не удалось закрыть v22-книгу %s",
                    workbook_path,
                )


def run_save_v22(
    year: int,
    *,
    auto_sample_codes: bool = False,
    projects_text: str = "",
) -> RunSummary:
    validate_legacy_addin_before_save()

    state = load_state()
    inspections = inspect_all_v22(
        year,
        auto_sample_codes=auto_sample_codes,
    )
    write_inspections_to_state(state, inspections)
    save_state(state)

    requested_projects = parse_projects_input(
        projects_text,
        year=year,
    )
    requested_set = set(requested_projects)

    if requested_set:
        selected_inspections = [
            item
            for item in inspections
            if item.project in requested_set
        ]

        found_projects = {item.project for item in selected_inspections}
        missing_projects = [
            project
            for project in requested_projects
            if project not in found_projects
        ]

        progress(
            "SAVE v22: выбран диапазон/список проектов: "
            + ", ".join(requested_projects)
        )

        if missing_projects:
            progress(
                "SAVE v22: в выбранном диапазоне не найдены v22-проекты: "
                + ", ".join(missing_projects)
            )
    else:
        selected_inspections = inspections
        progress(
            "SAVE v22: проекты не указаны — обрабатываю весь выбранный год"
        )

    skipped_ok: list[V22Inspection] = []
    pending_inspections: list[V22Inspection] = []

    for inspection in selected_inspections:
        entry = project_state_entry(
            state,
            Path(inspection.folder),
        )

        previous_save_status = normalize_text(
            entry.get("save_v22", {}).get("status")
        ).upper()

        if previous_save_status == "OK":
            skipped_ok.append(inspection)
            continue

        pending_inspections.append(inspection)

    ready = [
        item
        for item in pending_inspections
        if item.ready_for_save
    ]

    manual_precheck = sum(
        item.status.startswith("MANUAL_REVIEW")
        for item in pending_inspections
    )

    blocked_precheck = (
        len(pending_inspections)
        - len(ready)
        - manual_precheck
    )

    skipped_ok_count = len(skipped_ok)

    progress(
        f"SAVE v22: SELECTED={len(selected_inspections)}, "
        f"SKIP_OK={skipped_ok_count}, READY={len(ready)}, "
        f"MANUAL={manual_precheck}, BLOCKED={blocked_precheck}"
    )

    if skipped_ok:
        progress(
            "SAVE v22: уже имеют save_v22=OK и будут пропущены: "
            + ", ".join(item.project for item in skipped_ok)
        )

    if not ready:
        report_path = write_status_report(state, year)

        return RunSummary(
            ok=blocked_precheck == 0,
            mode="save-v22",
            year=year,
            processed=len(selected_inspections),
            succeeded=0,
            manual_review=manual_precheck,
            blocked=blocked_precheck,
            ignored=skipped_ok_count,
            message=(
                "Нет выбранных v22-книг, требующих автоматического сохранения. "
                f"SKIP_OK={skipped_ok_count}; "
                f"MANUAL_REVIEW={manual_precheck}; "
                f"BLOCKED={blocked_precheck}."
            ),
            state_file=str(STATE_PATH),
            status_report=str(report_path),
        )

    if len(xw.apps) == 0:
        raise RuntimeError(
            "Excel не запущен. "
            "Сначала вручную запусти Excel и дождись загрузки надстройки."
        )

    if len(xw.apps) > 1:
        raise RuntimeError(
            "Запущено несколько экземпляров Excel. "
            "Закрой лишние экземпляры и оставь только тот, "
            "в котором штатно загружена старая надстройка."
        )

    app = xw.apps.active

    progress(
        f"Использую уже запущенный Excel: "
        f"HWND={app.api.Hwnd}"
    )

    save_ok = 0
    save_manual = manual_precheck
    save_failed = blocked_precheck

    try:
        addin_name = INSTALLED_ADDIN_PATH.name
        addin_found = False

        for i in range(1, app.api.AddIns.Count + 1):
            addin = app.api.AddIns.Item(i)

            if str(addin.Name).casefold() != addin_name.casefold():
                continue

            addin_found = True

            progress(
                f"Использую уже загруженную AddIn: "
                f"Name={addin.Name}; "
                f"FullName={addin.FullName}; "
                f"Installed={addin.Installed}"
            )

            if not bool(addin.Installed):
                raise RuntimeError(
                    f"Надстройка {addin_name!r} найдена, "
                    "но в текущем Excel она не активна. "
                    "Скрипт специально не будет её переподключать."
                )

            break

        if not addin_found:
            raise RuntimeError(
                f"В уже запущенном Excel не найдена надстройка "
                f"{addin_name!r}. "
                "Закрой Excel, запусти его обычным способом и дождись "
                "штатной загрузки надстройки."
            )

        disable_excel_prompts(app)
        app.screen_updating = False

        for index, inspection in enumerate(ready, start=1):
            folder = Path(inspection.folder)
            workbook_path = Path(inspection.selected_v22)

            progress("-" * 80)
            progress(
                f"SAVE v22: {index}/{len(ready)} "
                f"{folder.name} -> {workbook_path.name}"
            )

            status, message, study_results = save_one_v22_workbook(
                app=app,
                addin_name=addin_name,
                workbook_path=workbook_path,
            )

            entry = project_state_entry(state, folder)
            entry["v22_workbook"] = str(workbook_path)
            entry["save_v22"] = {
                "at": now_text(),
                "status": status,
                "message": message,
                "studies": study_results,
            }

            if status == "OK":
                save_ok += 1
            elif status == "MANUAL_REVIEW":
                save_manual += 1
            else:
                save_failed += 1

            save_state(state)

            progress(
                f"{folder.name}: {status} — {message}"
            )

    finally:
        try:
            app.screen_updating = True
            app.api.EnableEvents = True
            app.api.DisplayAlerts = True
            app.display_alerts = True
        except Exception:
            logger.exception(
                "Не удалось восстановить настройки Excel после save-v22"
            )

    report_path = write_status_report(state, year)

    return RunSummary(
        ok=save_failed == 0,
        mode="save-v22",
        year=year,
        processed=len(selected_inspections),
        succeeded=save_ok,
        manual_review=save_manual,
        blocked=save_failed,
        ignored=skipped_ok_count,
        message=(
            f"SAVE v22 завершён. OK={save_ok}; "
            f"SKIP_OK={skipped_ok_count}; "
            f"MANUAL_REVIEW={save_manual}; "
            f"BLOCKED/FAILED={save_failed}; "
            f"AUTO_CODES={'ON' if auto_sample_codes else 'OFF'}. "
            "create-v23 запускается только отдельным следующим запуском."
        ),
        state_file=str(STATE_PATH),
        status_report=str(report_path),
    )



# =============================================================================
# Переключение установленной надстройки v22 -> v23
# =============================================================================

def unique_addin_backup_path() -> Path:
    timestamp = timestamp_text()

    candidate = INSTALLED_ADDIN_PATH.with_name(
        f"{INSTALLED_ADDIN_PATH.stem}_v22_backup_{timestamp}"
        f"{INSTALLED_ADDIN_PATH.suffix}"
    )

    counter = 1
    while candidate.exists():
        candidate = INSTALLED_ADDIN_PATH.with_name(
            f"{INSTALLED_ADDIN_PATH.stem}_v22_backup_{timestamp}_{counter}"
            f"{INSTALLED_ADDIN_PATH.suffix}"
        )
        counter += 1

    return candidate


def switch_installed_addin_to_v23(
    *,
    release: ReleaseInfo,
    state: dict[str, Any],
) -> None:
    if not INSTALLED_ADDIN_PATH.exists():
        raise FileNotFoundError(
            f"Установленный AddIn-файл не найден: {INSTALLED_ADDIN_PATH}"
        )

    current_props = read_office_custom_properties(INSTALLED_ADDIN_PATH)
    current_version = current_props.get(ADDIN_VERSION_PROPERTY, "").strip()

    # Если этот путь уже содержит ровно текущую новую надстройку — повторно
    # legacy backup не создаём.
    if current_version == release.addin_version:
        progress(
            "Установленный путь уже содержит актуальную v23-надстройку; "
            "подмена не требуется"
        )
        state["addin_switch"] = {
            "at": now_text(),
            "status": "ALREADY_CURRENT",
            "legacy_backup": state.get("addin_switch", {}).get(
                "legacy_backup", ""
            ),
            "current_version": current_version,
        }
        save_state(state)
        return

    # У legacy v22 свойства PVTAddinVersion нет.
    if current_version:
        raise RuntimeError(
            "По установленному пути находится надстройка с неожиданной версией "
            f"{current_version}. Автоматическая подмена запрещена."
        )

    source_props = read_office_custom_properties(NEW_ADDIN_SOURCE_PATH)
    source_version = source_props.get(ADDIN_VERSION_PROPERTY, "").strip()

    if source_version != release.addin_version:
        raise RuntimeError(
            f"Исходная новая надстройка имеет версию {source_version or '<нет>'}, "
            f"ожидается {release.addin_version}"
        )

    backup_path = unique_addin_backup_path()
    temp_new = INSTALLED_ADDIN_PATH.with_name(
        f"{INSTALLED_ADDIN_PATH.stem}__v23_new__{timestamp_text()}"
        f"{INSTALLED_ADDIN_PATH.suffix}"
    )

    progress(f"Готовлю новую надстройку во временном файле: {temp_new.name}")
    shutil.copy2(NEW_ADDIN_SOURCE_PATH, temp_new)

    temp_props = read_office_custom_properties(temp_new)
    if temp_props.get(ADDIN_VERSION_PROPERTY, "").strip() != release.addin_version:
        safe_unlink(temp_new)
        raise RuntimeError("Проверка временной копии новой надстройки не пройдена")

    progress(f"Сохраняю legacy-надстройку рядом: {backup_path.name}")

    try:
        INSTALLED_ADDIN_PATH.replace(backup_path)

        try:
            temp_new.replace(INSTALLED_ADDIN_PATH)
        except Exception:
            if backup_path.exists() and not INSTALLED_ADDIN_PATH.exists():
                backup_path.replace(INSTALLED_ADDIN_PATH)
            raise

        installed_props = read_office_custom_properties(INSTALLED_ADDIN_PATH)
        installed_version = installed_props.get(
            ADDIN_VERSION_PROPERTY, ""
        ).strip()

        if installed_version != release.addin_version:
            raise RuntimeError(
                "После подмены версия установленной надстройки не совпала с ожидаемой"
            )

    except Exception:
        logger.exception("Ошибка подмены legacy-надстройки на v23")

        # Пытаемся откатить.
        try:
            if INSTALLED_ADDIN_PATH.exists():
                current_after_error = read_office_custom_properties(
                    INSTALLED_ADDIN_PATH
                ).get(ADDIN_VERSION_PROPERTY, "").strip()

                if current_after_error == release.addin_version:
                    INSTALLED_ADDIN_PATH.unlink()

            if backup_path.exists() and not INSTALLED_ADDIN_PATH.exists():
                backup_path.replace(INSTALLED_ADDIN_PATH)
        except Exception:
            logger.exception("Не удалось откатить legacy-надстройку")

        safe_unlink(temp_new)
        raise

    safe_unlink(temp_new)

    state["addin_switch"] = {
        "at": now_text(),
        "status": "SWITCHED_TO_V23",
        "legacy_backup": str(backup_path),
        "current_version": release.addin_version,
    }
    save_state(state)


def verify_loaded_v23_addin(
    addin: Any,
    expected_version: str,
) -> None:
    actual_path = Path(str(addin.FullName))

    if not actual_path.exists():
        raise RuntimeError(
            f"Excel сообщает недоступный путь надстройки: {actual_path}"
        )

    props = read_office_custom_properties(actual_path)
    actual_version = props.get(ADDIN_VERSION_PROPERTY, "").strip()

    if actual_version != expected_version:
        raise RuntimeError(
            "Реально подключённая Excel-надстройка имеет неверную версию. "
            f"Ожидается {expected_version}, найдено {actual_version or '<нет>'}. "
            f"Файл: {actual_path}"
        )


# =============================================================================
# create-v23 helpers
# =============================================================================

def run_refresh_tasks_for_book(
    *,
    workbook_path: Path,
    refresh_project: str,
) -> None:
    from refresh_tasks import refresh_tasks

    progress(f"Обновляю задания для {refresh_project}")

    refresh_tasks(
        workbook=str(workbook_path),
        db_path=str(DB_PATH),
        source_workbook=str(TASKS_WORKBOOK),
        project_number=refresh_project,
        source_parent_sheet=REFRESH_TASKS_SOURCE_PARENT_SHEET,
        source_parent_table=REFRESH_TASKS_SOURCE_PARENT_TABLE,
        source_child_sheet=REFRESH_TASKS_SOURCE_CHILD_SHEET,
        source_child_table=REFRESH_TASKS_SOURCE_CHILD_TABLE,
        target_parent_sheet=REFRESH_TASKS_TARGET_PARENT_SHEET,
        target_parent_table=REFRESH_TASKS_TARGET_PARENT_TABLE,
        target_child_sheet=REFRESH_TASKS_TARGET_CHILD_SHEET,
        target_child_table=REFRESH_TASKS_TARGET_CHILD_TABLE,
    )


def run_load_all_studies_for_book(
    *,
    workbook_path: Path,
    refresh_project: str,
) -> None:
    from sync import run_load

    progress(f"SQLite -> staging: {refresh_project}")

    report = run_load(
        workbook=str(workbook_path),
        db_path=str(DB_PATH),
        project_number=refresh_project,
        study_codes=None,
    )

    if not report.ok:
        raise RuntimeError(report.message)


def set_project_cell(book: xw.Book, refresh_project: str) -> None:
    book.sheets[PROJECT_CELL_SHEET].range(PROJECT_CELL_ADDRESS).value = (
        refresh_project
    )


def temporary_v23_path(folder: Path, final_name: str) -> Path:
    return folder / (
        f"{Path(final_name).stem}__new__"
        f"{datetime.now():%Y%m%d_%H%M%S_%f}.xlsx"
    )


def unique_v22_archive_path(
    archive_dir: Path,
    old_workbook: Path,
) -> Path:
    timestamp = timestamp_text()

    candidate = archive_dir / (
        f"{old_workbook.stem}_backup_{timestamp}{old_workbook.suffix}"
    )

    counter = 1
    while candidate.exists():
        candidate = archive_dir / (
            f"{old_workbook.stem}_backup_{timestamp}_{counter}"
            f"{old_workbook.suffix}"
        )
        counter += 1

    return candidate


def build_v23_workbook(
    *,
    app: xw.App,
    addin_name: str,
    template_path: Path,
    temp_path: Path,
    refresh_project: str,
) -> None:
    workbook: xw.Book | None = None

    shutil.copy2(template_path, temp_path)

    try:
        disable_excel_prompts(app)

        workbook = app.books.open(
            str(temp_path),
            update_links=True,
            read_only=False,
            ignore_read_only_recommended=True,
        )

        disable_excel_prompts(app)
        unprotect_all_sheets(workbook)

        progress(
            f"Записываю проект в {PROJECT_CELL_SHEET}!{PROJECT_CELL_ADDRESS}: "
            f"{refresh_project}"
        )
        set_project_cell(workbook, refresh_project)
        workbook.save()

        # Данные из БД загружаются напрямую Python-функцией sync.run_load.
        run_load_all_studies_for_book(
            workbook_path=temp_path,
            refresh_project=refresh_project,
        )

        run_addin_macro(
            app=app,
            workbook=workbook,
            addin_name=addin_name,
            macro_name=LOAD_FORMS_FROM_TABLES_MACRO,
        )

        run_refresh_tasks_for_book(
            workbook_path=temp_path,
            refresh_project=refresh_project,
        )

        run_addin_macro(
            app=app,
            workbook=workbook,
            addin_name=addin_name,
            macro_name=AFTER_REFRESH_MACRO,
        )

        # Хэш и статусы рассчитывает только VBA.
        run_addin_macro(
            app=app,
            workbook=workbook,
            addin_name=addin_name,
            macro_name=SET_ALL_BASELINES_MACRO,
            args=(workbook.api,),
        )

        protect_all_sheets(workbook)

        progress("Сохраняю временную v23-книгу")
        disable_excel_prompts(app)
        workbook.save()

    finally:
        if workbook is not None:
            try:
                disable_excel_prompts(app)
                workbook.close()
            except Exception:
                logger.exception("Не удалось закрыть временную v23-книгу %s", temp_path)


def validate_built_v23_workbook(
    path: Path,
    release: ReleaseInfo,
) -> None:
    props = read_office_custom_properties(path)

    actual_version = props.get(FORM_VERSION_PROPERTY, "").strip()

    if actual_version != release.form_version:
        raise RuntimeError(
            f"Созданная книга имеет {FORM_VERSION_PROPERTY}="
            f"{actual_version or '<нет>'}; ожидается {release.form_version}"
        )

    bad_states: list[str] = []

    for study, property_name in STUDY_STATE_PROPERTIES.items():
        state = props.get(property_name, "unknown").strip().lower() or "unknown"

        if state not in SAFE_V23_STATES:
            bad_states.append(f"{study}={state}")

    if bad_states:
        raise RuntimeError(
            "VBA не сформировал безопасный baseline новой v23: "
            + ", ".join(bad_states)
        )


def archive_v22_and_activate_v23(
    *,
    old_v22: Path,
    temp_v23: Path,
    final_v23: Path,
) -> Path:
    archive_dir = old_v22.parent / ARCHIVE_DIR_NAME
    archive_dir.mkdir(parents=True, exist_ok=True)

    backup_path = unique_v22_archive_path(archive_dir, old_v22)

    progress(f"Переношу v22 в архив: {backup_path.name}")
    old_v22.replace(backup_path)

    try:
        if final_v23.exists():
            raise FileExistsError(
                f"Финальный v23-файл уже существует: {final_v23}"
            )

        temp_v23.replace(final_v23)

    except Exception:
        # Возвращаем v22 на рабочее место.
        try:
            if backup_path.exists() and not old_v22.exists():
                backup_path.replace(old_v22)
        except Exception:
            logger.exception("Не удалось вернуть v22 из архива: %s", backup_path)

        raise

    return backup_path


# =============================================================================
# Режим create-v23
# =============================================================================

def eligible_saved_v22_entries(
    state: dict[str, Any],
    year: int,
    *,
    projects_text: str = "",
) -> list[dict[str, Any]]:
    requested_projects = parse_projects_input(
        projects_text,
        year=year,
    )
    requested_set = set(requested_projects)

    result: list[dict[str, Any]] = []

    for entry in state.get("projects", {}).values():
        if entry.get("year") != year:
            continue

        if entry.get("save_v22", {}).get("status") != "OK":
            continue

        # Если задан конкретный список/диапазон —
        # берём только эти проекты.
        if requested_set:
            project = normalize_text(
                entry.get("project")
            )

            if project not in requested_set:
                continue

        # Уже успешно созданные v23 повторно не трогаем.
        if entry.get("create_v23", {}).get("status") == "OK":
            continue

        result.append(entry)

    result.sort(
        key=lambda item: str(
            item.get("folder", "")
        ).casefold()
    )

    return result


def create_one_v23_from_saved_v22(
    *,
    app: xw.App,
    addin_name: str,
    release: ReleaseInfo,
    entry: dict[str, Any],
) -> tuple[str, str, str, str]:
    folder = Path(entry["folder"])
    project = normalize_text(entry.get("project"))

    if not project:
        raise RuntimeError(f"Не удалось определить номер проекта из папки {folder}")

    if not folder.exists():
        raise FileNotFoundError(f"Папка проекта исчезла: {folder}")

    # Повторно проверяем v22 непосредственно перед разрушительной операцией.
    v22_candidates = find_v22_workbooks(folder)

    if len(v22_candidates) != 1:
        raise RuntimeError(
            "Перед create-v23 количество v22-книг изменилось. Найдено: "
            f"{len(v22_candidates)}"
        )

    old_v22 = v22_candidates[0]

    saved_v22_path = normalize_text(entry.get("v22_workbook"))
    if saved_v22_path and str(old_v22) != saved_v22_path:
        raise RuntimeError(
            "v22-книга отличается от той, которая была успешно сохранена "
            f"на этапе save-v22. Было: {saved_v22_path}; сейчас: {old_v22}"
        )

    if excel_lock_present(old_v22):
        raise RuntimeError(f"v22-книга открыта в Excel: {old_v22}")

    newer = find_v23_or_newer_workbooks(folder)

    if newer:
        raise RuntimeError(
            "В рабочей папке уже есть v23 или более новая книга: "
            + ", ".join(path.name for path in newer)
        )

    final_v23 = folder / f"{folder.name}_v{release.form_version}.xlsx"
    temp_v23 = temporary_v23_path(folder, final_v23.name)

    try:
        build_v23_workbook(
            app=app,
            addin_name=addin_name,
            template_path=Path(release.template_path),
            temp_path=temp_v23,
            refresh_project=project,
        )

        validate_built_v23_workbook(temp_v23, release)

        backup_path = archive_v22_and_activate_v23(
            old_v22=old_v22,
            temp_v23=temp_v23,
            final_v23=final_v23,
        )

        return (
            "OK",
            "v23 создана; старая v22 перенесена в дочернюю папку 'архив'.",
            str(final_v23),
            str(backup_path),
        )

    except Exception:
        safe_unlink(temp_v23)
        raise


def run_create_v23(
    year: int,
    *,
    projects_text: str = "",
) -> RunSummary:
    state = load_state()

    entries = eligible_saved_v22_entries(
        state,
        year,
        projects_text=projects_text,
    )

    if not entries:
        raise RuntimeError(
            "Нет выбранных проектов, доступных для create-v23. "
            "Проект должен иметь save_v22=OK и ещё не иметь create_v23=OK."
        )

    # В create-v23 только здесь впервые проверяем новую release-конфигурацию.
    release = preflight_v23_release()

    # ВАЖНО: create-v23 — отдельный запуск. Старый Excel из save-v22 уже должен
    # быть закрыт. Подменяем файл по тому же зарегистрированному пути AddIn.
    switch_installed_addin_to_v23(
        release=release,
        state=state,
    )

    app = xw.App(visible=True, add_book=False)

    ok_count = 0
    failed_count = 0

    try:
        # Надстройке нужно дать нормально инициализироваться через события.
        app.display_alerts = False
        app.api.DisplayAlerts = False
        app.api.AskToUpdateLinks = False
        app.api.AlertBeforeOverwriting = False
        app.api.EnableEvents = True
        app.screen_updating = False

        progress("Подключаю уже зарегистрированную AddIn — теперь это новая v23")
        addin = load_installed_addin_by_registered_name(
            app,
            INSTALLED_ADDIN_PATH,
        )
        verify_loaded_v23_addin(addin, release.addin_version)
        addin_name = str(addin.Name)

        disable_excel_prompts(app)

        for index, entry in enumerate(entries, start=1):
            folder = Path(entry["folder"])

            progress("-" * 80)
            progress(
                f"CREATE v23: {index}/{len(entries)} {folder.name}"
            )

            try:
                status, message, v23_workbook, v22_backup = (
                    create_one_v23_from_saved_v22(
                        app=app,
                        addin_name=addin_name,
                        release=release,
                        entry=entry,
                    )
                )

                entry["create_v23"] = {
                    "at": now_text(),
                    "status": status,
                    "message": message,
                    "v23_workbook": v23_workbook,
                    "v22_backup": v22_backup,
                }

                ok_count += 1
                progress(f"{folder.name}: OK")

            except Exception as exc:
                logger.exception("Ошибка create-v23 для %s", folder)

                entry["create_v23"] = {
                    "at": now_text(),
                    "status": "FAILED",
                    "message": f"{type(exc).__name__}: {exc}",
                    "v23_workbook": "",
                    "v22_backup": "",
                }

                failed_count += 1
                progress(f"{folder.name}: FAILED — {exc}")

            save_state(state)

    finally:
        try:
            app.quit()
        except Exception:
            logger.exception("Не удалось закрыть Excel после create-v23")

    report_path = write_status_report(state, year)

    return RunSummary(
        ok=failed_count == 0,
        mode="create-v23",
        year=year,
        processed=len(entries),
        succeeded=ok_count,
        manual_review=0,
        blocked=failed_count,
        ignored=0,
        message=(
            f"CREATE v23 завершён. OK={ok_count}; FAILED/BLOCKED={failed_count}. "
            "Обработаны только проекты с save_v22=OK."
        ),
        state_file=str(STATE_PATH),
        status_report=str(report_path),
    )


# =============================================================================
# UI
# =============================================================================

def ask_ui_options() -> UiOptions:
    try:
        import tkinter as tk
        from tkinter import messagebox

        result = UiOptions(
            mode="scan-v22",
            year=MIGRATION_YEAR_DEFAULT,
            auto_sample_codes=False,
            projects_text="",
            cancelled=True,
        )

        root = tk.Tk()
        root.title("Миграция форм ГТИ v22 → v23")
        root.resizable(False, False)

        frame = tk.Frame(root, padx=14, pady=14)
        frame.pack(fill="both", expand=True)

        tk.Label(
            frame,
            text=(
                "Каждый режим запускается ОТДЕЛЬНО.\n"
                "Для save-v22 можно указать проект/диапазон; пусто = весь год.\n"
                "Файлы v21 игнорируются полностью."
            ),
            justify="left",
            anchor="w",
        ).pack(fill="x", pady=(0, 12))

        mode_frame = tk.LabelFrame(frame, text="Режим", padx=10, pady=8)
        mode_frame.pack(fill="x", pady=(0, 12))

        mode_var = tk.StringVar(value="scan-v22")

        tk.Radiobutton(
            mode_frame,
            text="scan-v22 — только анализ и Excel-отчёт, Excel не запускается",
            variable=mode_var,
            value="scan-v22",
        ).pack(anchor="w")

        tk.Radiobutton(
            mode_frame,
            text=(
                "save-v22 — сохранить выбранные безопасные v22 "
                "через уже запущенный Excel"
            ),
            variable=mode_var,
            value="save-v22",
        ).pack(anchor="w")

        tk.Radiobutton(
            mode_frame,
            text=(
                "create-v23 — отдельный этап: заменить AddIn и создать v23 "
                "только для save_v22=OK"
            ),
            variable=mode_var,
            value="create-v23",
        ).pack(anchor="w")

        year_frame = tk.Frame(frame)
        year_frame.pack(fill="x", pady=(0, 12))

        tk.Label(year_frame, text="Год:").pack(side="left")
        year_var = tk.StringVar(value=str(MIGRATION_YEAR_DEFAULT))
        tk.Entry(
            year_frame,
            textvariable=year_var,
            width=8,
        ).pack(side="left", padx=(8, 0))

        projects_frame = tk.Frame(frame)
        projects_frame.pack(fill="x", pady=(0, 8))

        tk.Label(
            projects_frame,
            text="Проекты для save-v22 / create-v23:",
        ).pack(side="left")

        projects_var = tk.StringVar(value="")
        tk.Entry(
            projects_frame,
            textvariable=projects_var,
            width=44,
        ).pack(side="left", padx=(8, 0))

        tk.Label(
            frame,
            text=(
                "Примеры: 26-F001 | 26-F001...26-F050 | "
                "26-F001, 26-F005, 26-F010...26-F020\n"
                "Для save-v22 пусто = весь выбранный год.\n"
                "Для create-v23 обязательно укажи проект, список или диапазон.\n"
                "В save-v22 уже сохранённые save_v22=OK пропускаются.\n"
                "В create-v23 обрабатываются только save_v22=OK, "
                "а create_v23=OK пропускаются."
            ),
            justify="left",
            anchor="w",
            wraplength=760,
        ).pack(fill="x", pady=(0, 12))

        auto_sample_codes_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            frame,
            text=(
                "Автоматически создать шифры SSF/ПЩК/СВМЭ "
                "(silent-режим GOR / SSF / CVD)"
            ),
            variable=auto_sample_codes_var,
        ).pack(anchor="w", pady=(0, 12))

        tk.Label(
            frame,
            text=(
                "По умолчанию GOR / SSF / CVD без resultId попадают в MANUAL_REVIEW.\n"
                "Если включить автоматическое создание шифров, такие проекты будут "
                "допущены к save-v22, а silent VBA создаст/примет шифры без окон.\n"
                "Неожиданное VBA-окно по-прежнему будет закрыто watchdog-ом, "
                "а проект получит MANUAL_REVIEW."
            ),
            justify="left",
            anchor="w",
            wraplength=760,
        ).pack(fill="x", pady=(0, 12))

        buttons = tk.Frame(frame)
        buttons.pack(fill="x")

        def submit() -> None:
            try:
                year = int(year_var.get().strip())
            except ValueError:
                messagebox.showerror("Ошибка", "Год должен быть целым числом")
                return

            mode = mode_var.get().strip().lower()
            projects_text = projects_var.get().strip()

            if mode in {"save-v22", "create-v23"} and projects_text:
                try:
                    parse_projects_input(
                        projects_text,
                        year=year,
                    )
                except ValueError as exc:
                    messagebox.showerror(
                        "Ошибка",
                        str(exc),
                    )
                    return
            
            if mode == "create-v23" and not projects_text:
                messagebox.showerror(
                    "Ошибка",
                    "Для create-v23 обязательно укажи "
                    "проект, список проектов или диапазон."
                )
                return

            if mode == "create-v23":
                answer = messagebox.askyesno(
                    "Подтверждение",
                    "create-v23 заменит файл установленной старой надстройки "
                    "новой v23 (старый файл будет сохранён рядом как backup) "
                    "и начнёт создавать новые книги.\n\n"
                    f"Выбранные проекты:\n{projects_text}\n\n"
                    "Продолжить?",
                )

                if not answer:
                    return

            result.mode = mode
            result.year = year
            result.auto_sample_codes = bool(auto_sample_codes_var.get())
            result.projects_text = projects_text
            result.cancelled = False
            root.destroy()

        def cancel() -> None:
            result.cancelled = True
            root.destroy()

        tk.Button(
            buttons,
            text="Запустить",
            width=14,
            command=submit,
        ).pack(side="left")

        tk.Button(
            buttons,
            text="Отмена",
            width=14,
            command=cancel,
        ).pack(side="left", padx=(8, 0))

        root.bind("<Return>", lambda _event: submit())
        root.bind("<Escape>", lambda _event: cancel())
        root.protocol("WM_DELETE_WINDOW", cancel)

        root.update_idletasks()
        width = root.winfo_reqwidth()
        height = root.winfo_reqheight()
        x = max((root.winfo_screenwidth() - width) // 2, 0)
        y = max((root.winfo_screenheight() - height) // 2, 0)
        root.geometry(f"{width}x{height}+{x}+{y}")

        root.mainloop()
        return result

    except Exception:
        print("Режимы: scan-v22 / save-v22 / create-v23")
        mode = input("Режим: ").strip().lower()

        if mode not in MODES:
            return UiOptions(
                mode="",
                year=MIGRATION_YEAR_DEFAULT,
                cancelled=True,
            )

        raw_year = input(f"Год [{MIGRATION_YEAR_DEFAULT}]: ").strip()
        year = int(raw_year) if raw_year else MIGRATION_YEAR_DEFAULT

        projects_text = ""

        if mode in {"save-v22", "create-v23"}:
            projects_text = input(
                "Проект/диапазон "
                "(например 26-F001...26-F050; "
                "для save-v22 пусто = весь год): "
            ).strip()

            if projects_text:
                parse_projects_input(
                    projects_text,
                    year=year,
                )

        if mode == "create-v23" and not projects_text:
            raise ValueError(
                "Для create-v23 обязательно укажи "
                "проект, список проектов или диапазон."
            )

        auto_raw = input(
            "Автоматически создать шифры SSF/ПЩК/СВМЭ? [y/N]: "
        ).strip().casefold()
        auto_sample_codes = auto_raw in {"y", "yes", "д", "да"}

        return UiOptions(
            mode=mode,
            year=year,
            auto_sample_codes=auto_sample_codes,
            projects_text=projects_text,
            cancelled=False,
        )



# =============================================================================
# CLI / main
# =============================================================================

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Пошаговая миграция Excel-проектов ГТИ с v22 на v23. "
            "scan-v22, save-v22 и create-v23 запускаются отдельно."
        )
    )

    parser.add_argument(
        "--mode",
        choices=sorted(MODES),
        help="scan-v22, save-v22 или create-v23",
    )

    parser.add_argument(
        "--year",
        type=int,
        default=MIGRATION_YEAR_DEFAULT,
        help=f"Год папок проектов. По умолчанию {MIGRATION_YEAR_DEFAULT}.",
    )

    parser.add_argument(
        "--projects",
        default="",
        help=(
            "Проект/диапазон для save-v22 или create-v23, например "
            "26-F001 или 26-F001...26-F050 или список через запятую. "
            "Для save-v22 пусто = весь год; "
            "для create-v23 параметр обязателен."
        ),
    )

    parser.add_argument(
        "--auto-sample-codes",
        action="store_true",
        help=(
            "Разрешить silent VBA автоматически создавать/принимать шифры "
            "GOR/SSF/CVD вместо MANUAL_REVIEW."
        ),
    )

    return parser


def run_mode(
    mode: str,
    year: int,
    *,
    auto_sample_codes: bool = False,
    projects_text: str = "",
) -> RunSummary:
    if mode == "scan-v22":
        return run_scan_v22(
            year,
            auto_sample_codes=auto_sample_codes,
        )

    if mode == "save-v22":
        return run_save_v22(
            year,
            auto_sample_codes=auto_sample_codes,
            projects_text=projects_text,
        )

    if mode == "create-v23":
        return run_create_v23(
            year,
            projects_text=projects_text,
        )

    raise ValueError(f"Неизвестный режим: {mode}")


def main(argv: list[str] | None = None) -> int:
    args_list = list(sys.argv[1:] if argv is None else argv)

    if not args_list:
        options = ask_ui_options()

        if options.cancelled:
            progress("Операция отменена пользователем")
            return 0

        mode = options.mode
        year = options.year
        auto_sample_codes = options.auto_sample_codes
        projects_text = options.projects_text

    else:
        parser = build_parser()
        args = parser.parse_args(args_list)

        if not args.mode:
            parser.error("--mode обязателен при запуске с параметрами")

        mode = args.mode
        year = args.year
        auto_sample_codes = bool(args.auto_sample_codes)
        projects_text = str(args.projects or "").strip()

        if mode in {"save-v22", "create-v23"} and projects_text:
            parse_projects_input(
                projects_text,
                year=year,
            )

        if mode == "create-v23" and not projects_text:
            parser.error(
                "--projects обязателен для create-v23"
            )

    try:
        summary = run_mode(
            mode,
            year,
            auto_sample_codes=auto_sample_codes,
            projects_text=projects_text,
        )
    except Exception as exc:
        logger.exception("Критическая ошибка мигратора")

        summary = RunSummary(
            ok=False,
            mode=mode,
            year=year,
            processed=0,
            succeeded=0,
            manual_review=0,
            blocked=1,
            ignored=0,
            message=(
                f"{type(exc).__name__}: {exc}. "
                f"Подробности в логе: {LOG_PATH}"
            ),
            state_file=str(STATE_PATH),
            status_report=str(STATUS_REPORT_PATH),
        )

    print(json.dumps(asdict(summary), ensure_ascii=False, indent=2))
    return 0 if summary.ok else 1


if __name__ == "__main__":
    sys.exit(main())