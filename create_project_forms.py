from __future__ import annotations

import argparse
import json
import logging
import re
import shutil
import sqlite3
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterable
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
import xlwings as xw
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =============================================================================
# Настройки
# =============================================================================

PROJECT_DIR = Path(__file__).resolve().parent
VERSION_FILE = PROJECT_DIR / "project_version.json"

DB_PATH = Path(r"L:\LRC\common_data\ФЛЮИДЫ\ГТИ\sqlite-excel\sqlite\results.db")
BASE_PROJECTS_DIR = Path(r"L:\LRC\common_data\ФЛЮИДЫ\ГТИ\Работа")
TASKS_WORKBOOK = Path(r"L:\LRC\exchange\КСП Лайт\Журнал_заданий_флюиды.xlsx")

# Чистый шаблон ищется рядом со скриптом. Версия определяется только по metadata.
CLEAN_TEMPLATE_GLOB = "*_Форма_v*.xlsx"

# Важно: xlam не открывается через app.books.open(). По этому пути берётся
# только имя установленной надстройки и проверяются metadata.
ADDIN_PATH = Path(r"L:\LRC\common_data\ФЛЮИДЫ\ГТИ\sqlite-excel\надстройка новая ribbon.xlam")

FORM_VERSION_PROPERTY = "PVTFormVersion"
ADDIN_VERSION_PROPERTY = "PVTAddinVersion"
FORM_VERSION_KEYS = ("form", "form_version", "PVTFormVersion")
ADDIN_VERSION_KEYS = ("addin", "add_in", "addin_version", "PVTAddinVersion")

UNPROTECT_PASSWORDS = ("1984", "9184", "")
PROTECT_PASSWORD = "1984"

TASKS_PARENT_SHEET = "ГТИ"
TASKS_PARENT_TABLE = "Журнал_ГТИ"
TASK_SAMPLE_CODE_COL = "Код проекта"
TASK_DATETIME_COL = "Дата и время"

REFRESH_TASKS_SOURCE_PARENT_SHEET = "ГТИ"
REFRESH_TASKS_SOURCE_PARENT_TABLE = "Журнал_ГТИ"
REFRESH_TASKS_SOURCE_CHILD_SHEET = "Смешение"
REFRESH_TASKS_SOURCE_CHILD_TABLE = "ЖУрнал_объединения"
REFRESH_TASKS_TARGET_PARENT_SHEET = "Task"
REFRESH_TASKS_TARGET_PARENT_TABLE = "Task"
REFRESH_TASKS_TARGET_CHILD_SHEET = "Task_mix"
REFRESH_TASKS_TARGET_CHILD_TABLE = "Task_mix"

AFTER_REFRESH_MACRO = "PowerQuery.silentRefresh_Project"
LOAD_FORMS_FROM_TABLES_MACRO = "wrappers.Load_All_results_to_forms_Auto"
SET_ALL_BASELINES_MACRO = "modMetadata.SetAllStudyBaselinesAfterLoad"

PROJECT_CELL_SHEET = "OP"
PROJECT_CELL_ADDRESS = "B6"
RECENT_PROJECT_LIMIT = 25
FOLDER_SEARCH_PARTS = 3
FOLDER_CREATE_PARTS = 4
REFRESH_PROJECT_PARTS = 2
ARCHIVE_DIR_NAME = "архив"
LOG_PATH = PROJECT_DIR / "project_forms.log"

STUDY_STATE_PROPERTIES: dict[str, str] = {
    "OP": "OPDataState",
    "OPOH": "OPOHDataState",
    "AP": "APDataState",
    "BP": "BPDataState",
    "GC": "GCDataState",
    "EMV": "EMVDataState",
    "GOR": "GORDataState",
    "SSF": "SSFDataState",
    "CCE": "CCEDataState",
    "CVD": "CVDDataState",
    "REC": "RECDataState",
    "JOIN": "JOINDataState",
}
SAFE_RECREATE_STATES = {"empty", "synced"}
MODES = {"scan", "update", "recreate"}

logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    filemode="a",
    encoding="utf-8",
)

logger = logging.getLogger(__name__)


# =============================================================================
# Dataclasses
# =============================================================================

@dataclass
class ReleaseInfo:
    form_version: str
    addin_version: str
    form_major: int
    template_path: str
    addin_path: str


@dataclass
class TaskProjectInfo:
    refresh_project: str
    latest_datetime: datetime | None = None
    latest_sample_code: str = ""
    new_folder_name: str = ""
    search_keys: tuple[str, ...] = ()

    @property
    def ambiguous_search_key(self) -> bool:
        return len(self.search_keys) > 1

    @property
    def search_key(self) -> str:
        return self.search_keys[0] if len(self.search_keys) == 1 else ""


@dataclass
class ProjectInspection:
    refresh_project: str
    expected_folder: str = ""
    folder_search_key: str = ""
    found_folders: list[str] = field(default_factory=list)
    selected_folder: str = ""
    workbook_candidates: list[str] = field(default_factory=list)
    selected_workbook: str = ""
    workbook_version: str = ""
    current_form_version: str = ""
    version_status: str = ""
    study_states: dict[str, str] = field(default_factory=dict)
    safe_to_recreate: bool = False
    blocking_studies: list[str] = field(default_factory=list)
    recommended_action: str = ""
    issue: str = ""
    excel_lock_present: bool = False


@dataclass
class ProjectRunResult:
    project: str
    ok: bool
    action: str
    message: str
    folder: str | None = None
    workbook: str | None = None
    backup: str | None = None
    error_type: str | None = None


@dataclass
class RunReport:
    ok: bool
    mode: str
    message: str
    form_version: str
    addin_version: str
    processed: int
    succeeded: int
    failed: int
    scan_report: str | None = None
    projects: list[ProjectRunResult] = field(default_factory=list)


@dataclass
class UiOptions:
    mode: str
    projects_text: str
    year: int
    cancelled: bool = False


# =============================================================================
# Общие helpers
# =============================================================================

def progress(message: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def canonicalize_f_part(value: str) -> str:
    text = str(value).strip().upper()
    match = re.fullmatch(r"F(\d{3})([A-ZА-Я]*)?", text)
    return f"F{match.group(1)}" if match else text


def derive_project_from_sample_code(sample_code: Any, parts_count: int) -> str:
    text = normalize_text(sample_code)
    if not text:
        return ""
    parts = text.split("-")
    if len(parts) < 2:
        return text
    parts[1] = canonicalize_f_part(parts[1])
    return "-".join(parts[:min(parts_count, len(parts))])


def normalize_project_number(value: str) -> str:
    value = value.strip().upper()
    match = re.fullmatch(r"(\d{2})-F(\d+)", value)
    if not match:
        raise ValueError(f"Некорректный номер проекта: {value}")
    year, number = match.groups()
    return f"{year}-F{int(number):03d}"


def project_from_folder_name(folder_name: str) -> str:
    match = re.match(r"^(\d{2})-F(\d{3})(?:-|$)", folder_name, re.IGNORECASE)
    if not match:
        return ""
    return f"{match.group(1)}-F{match.group(2)}".upper()


def project_year_folder(project: str) -> str:
    match = re.fullmatch(r"(\d{2})-F\d{3}", project, re.IGNORECASE)
    if not match:
        raise ValueError(f"Не удалось определить год проекта: {project}")
    return f"20{match.group(1)}"


def normalize_excel_date(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return None if value <= 0 else datetime(1899, 12, 30) + timedelta(days=float(value))  # noqa: DTZ001
    text = str(value).strip()
    if not text:
        return None
    for fmt in (
        "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, fmt)  # noqa: DTZ007
        except ValueError:
            pass
    return None


def parse_version_tuple(version: str) -> tuple[int, ...]:
    text = str(version).strip()
    if not re.fullmatch(r"\d+(?:\.\d+)*", text):
        raise ValueError(f"Некорректная версия: {version}")
    return tuple(int(part) for part in text.split("."))


def version_major(version: str) -> int:
    return parse_version_tuple(version)[0]


def compare_versions(left: str, right: str) -> int:
    a = list(parse_version_tuple(left))
    b = list(parse_version_tuple(right))
    size = max(len(a), len(b))
    a.extend([0] * (size - len(a)))
    b.extend([0] * (size - len(b)))
    return (a > b) - (a < b)


def expand_project_range(start: str, end: str) -> list[str]:
    start = normalize_project_number(start)
    end = normalize_project_number(end)
    m1 = re.fullmatch(r"(\d{2}-F)(\d{3})", start)
    m2 = re.fullmatch(r"(\d{2}-F)(\d{3})", end)
    if not m1 or not m2:
        raise ValueError(f"Некорректный диапазон: {start}...{end}")
    prefix1, n1 = m1.groups()
    prefix2, n2 = m2.groups()
    if prefix1 != prefix2:
        raise ValueError(f"Диапазон должен быть внутри одного года: {start}...{end}")
    n1_i, n2_i = int(n1), int(n2)
    if n2_i < n1_i:
        raise ValueError(f"Конец диапазона меньше начала: {start}...{end}")
    return [f"{prefix1}{i:03d}" for i in range(n1_i, n2_i + 1)]


def parse_projects_input(text: str) -> list[str]:
    text = text.strip()
    if not text:
        return []
    text = re.sub(r"\s*\.{2,3}\s*", "...", text)
    parts = re.split(r"[,\s;]+", text)
    result: list[str] = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if "..." in part:
            start, end = part.split("...", 1)
            result.extend(expand_project_range(start, end))
        else:
            result.append(normalize_project_number(part))
    return list(dict.fromkeys(result))


# =============================================================================
# Office custom metadata без открытия Excel
# =============================================================================

def read_office_custom_properties(path: Path) -> dict[str, str]:
    if not path.exists():
        raise FileNotFoundError(path)
    try:
        with zipfile.ZipFile(path, "r") as archive:
            try:
                xml_data = archive.read("docProps/custom.xml")
            except KeyError:
                return {}
    except zipfile.BadZipFile as exc:
        raise RuntimeError(f"Файл не является корректным Office-файлом: {path}") from exc

    root = ET.fromstring(xml_data)
    result: dict[str, str] = {}
    for prop in root:
        name = prop.attrib.get("name", "").strip()
        if not name:
            continue
        children = list(prop)
        result[name] = "" if not children or children[0].text is None else str(children[0].text).strip()
    return result


# =============================================================================
# project_version.json / release preflight
# =============================================================================

def load_version_manifest() -> dict[str, Any]:
    if not VERSION_FILE.exists():
        raise FileNotFoundError(f"Не найден project_version.json: {VERSION_FILE}")
    with VERSION_FILE.open("r", encoding="utf-8-sig") as file:
        data = json.load(file)
    if not isinstance(data, dict):
        raise TypeError("project_version.json должен содержать JSON-объект")
    return data


def extract_manifest_version(manifest: dict[str, Any], keys: Iterable[str], caption: str) -> str:
    for key in keys:
        if key not in manifest:
            continue
        value = manifest[key]
        if isinstance(value, str):
            version = value.strip()
        elif isinstance(value, dict):
            raw = value.get("version")
            version = "" if raw is None else str(raw).strip()
        else:
            version = ""
        if version:
            parse_version_tuple(version)
            return version
    raise ValueError(
        f"В project_version.json не найдена версия компонента '{caption}'. "
        f"Проверены ключи: {', '.join(keys)}"
    )


def locate_current_template(expected_version: str) -> Path:
    candidates = [
        path for path in PROJECT_DIR.glob(CLEAN_TEMPLATE_GLOB)
        if path.is_file() and not path.name.startswith("~$")
    ]
    matching: list[Path] = []
    for path in candidates:
        try:
            props = read_office_custom_properties(path)
        except Exception as exc:
            logger.warning("Не удалось прочитать metadata шаблона %s: %s", path, exc)
            continue
        if props.get(FORM_VERSION_PROPERTY, "").strip() == expected_version:
            matching.append(path)
    if not matching:
        raise FileNotFoundError(
            f"Не найден шаблон с {FORM_VERSION_PROPERTY}={expected_version}. "
            f"Маска: {PROJECT_DIR / CLEAN_TEMPLATE_GLOB}"
        )
    if len(matching) > 1:
        raise RuntimeError(
            f"Найдено несколько шаблонов версии {expected_version}: "
            + "; ".join(str(path) for path in matching)
        )
    return matching[0]


def validate_addin_file(path: Path, expected_version: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Файл надстройки не найден: {path}")
    props = read_office_custom_properties(path)
    actual = props.get(ADDIN_VERSION_PROPERTY, "").strip()
    if actual != expected_version:
        raise RuntimeError(
            f"Версия надстройки не соответствует project_version.json. "
            f"Ожидается {expected_version}, найдено {actual or '<нет>'}. Файл: {path}"
        )


def preflight_release() -> ReleaseInfo:
    progress("Проверяю project_version.json, шаблон формы и надстройку")
    manifest = load_version_manifest()
    form_version = extract_manifest_version(manifest, FORM_VERSION_KEYS, "form")
    addin_version = extract_manifest_version(manifest, ADDIN_VERSION_KEYS, "addin")
    template_path = locate_current_template(form_version)
    template_props = read_office_custom_properties(template_path)
    template_version = template_props.get(FORM_VERSION_PROPERTY, "").strip()
    if template_version != form_version:
        raise RuntimeError(
            f"Версия шаблона {template_version or '<нет>'} не совпадает с {form_version}"
        )
    validate_addin_file(ADDIN_PATH, addin_version)
    progress(
        f"Preflight OK: form={form_version}, addin={addin_version}, template={template_path.name}"
    )
    return ReleaseInfo(
        form_version=form_version,
        addin_version=addin_version,
        form_major=version_major(form_version),
        template_path=str(template_path),
        addin_path=str(ADDIN_PATH),
    )

# =============================================================================
# Чтение журнала заданий
# =============================================================================

def read_external_task_rows(workbook_path: Path, sheet_name: str, table_name: str) -> list[dict[str, Any]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Внешний журнал заданий не найден: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Во внешней книге нет листа '{sheet_name}'")
        ws = wb[sheet_name]

        if table_name in ws.tables:
            table = ws.tables[table_name]
            rows_values = [[cell.value for cell in row] for row in ws[table.ref]]
        else:
            rows_values = [
                list(row) for row in ws.iter_rows(values_only=True)
                if any(cell is not None for cell in row)
            ]

        if not rows_values:
            return []

        headers = [str(value).strip() if value is not None else "" for value in rows_values[0]]
        result: list[dict[str, Any]] = []
        for raw_row in rows_values[1:]:
            row = {headers[i]: raw_row[i] if i < len(raw_row) else "" for i in range(len(headers))}
            if all(normalize_text(value) == "" for value in row.values()):
                continue
            result.append(row)
        return result
    finally:
        wb.close()


def build_task_project_infos(rows: list[dict[str, Any]]) -> dict[str, TaskProjectInfo]:
    grouped: dict[str, dict[str, Any]] = {}

    for row in rows:
        sample_code = normalize_text(row.get(TASK_SAMPLE_CODE_COL))
        if not sample_code:
            continue

        refresh_project = derive_project_from_sample_code(sample_code, REFRESH_PROJECT_PARTS)
        search_key = derive_project_from_sample_code(sample_code, FOLDER_SEARCH_PARTS)
        new_folder_name = derive_project_from_sample_code(sample_code, FOLDER_CREATE_PARTS)

        try:
            refresh_project = normalize_project_number(refresh_project)
        except ValueError:
            continue

        dt = normalize_excel_date(row.get(TASK_DATETIME_COL))
        item = grouped.setdefault(
            refresh_project,
            {
                "latest_datetime": None,
                "latest_sample_code": "",
                "new_folder_name": "",
                "search_keys": set(),
            },
        )

        if search_key:
            item["search_keys"].add(search_key)

        old_dt = item["latest_datetime"]
        if old_dt is None or (dt is not None and dt > old_dt):
            item["latest_datetime"] = dt
            item["latest_sample_code"] = sample_code
            item["new_folder_name"] = new_folder_name

    return {
        project: TaskProjectInfo(
            refresh_project=project,
            latest_datetime=item["latest_datetime"],
            latest_sample_code=item["latest_sample_code"],
            new_folder_name=item["new_folder_name"],
            search_keys=tuple(sorted(item["search_keys"])),
        )
        for project, item in grouped.items()
    }


def latest_task_projects(infos: dict[str, TaskProjectInfo], limit: int) -> list[str]:
    sorted_infos = sorted(
        infos.values(),
        key=lambda item: item.latest_datetime or datetime.min,
        reverse=True,
    )
    return [item.refresh_project for item in sorted_infos[:limit]]


# =============================================================================
# Поиск папок и книг
# =============================================================================

def year_dir_for_project(project: str) -> Path:
    return BASE_PROJECTS_DIR / project_year_folder(project)


def list_project_ids_from_year_folder(year: int) -> set[str]:
    year_dir = BASE_PROJECTS_DIR / str(year)
    if not year_dir.exists():
        return set()

    result: set[str] = set()
    for path in year_dir.iterdir():
        if not path.is_dir() or path.name.casefold() == ARCHIVE_DIR_NAME.casefold():
            continue
        project = project_from_folder_name(path.name)
        if project:
            result.add(project)
    return result


def find_project_folders(
    refresh_project: str,
    task_info: TaskProjectInfo | None,
) -> tuple[list[Path], str, str]:
    """
    Возвращает matches, search_key, issue.

    При наличии данных журнала ищем строго по первым 3 фрагментам.
    Без данных журнала существующую папку можно найти fallback-поиском по 26-F083-...
    """
    year_dir = year_dir_for_project(refresh_project)
    search_key = task_info.search_key if task_info else ""

    if task_info is not None and task_info.ambiguous_search_key:
        return [], "", (
            "В журнале заданий для одного проекта найдены разные первые три фрагмента: "
            + ", ".join(task_info.search_keys)
        )

    if not year_dir.exists():
        return [], search_key, ""

    all_folders = [
        path for path in year_dir.iterdir()
        if path.is_dir() and path.name.casefold() != ARCHIVE_DIR_NAME.casefold()
    ]

    if search_key:
        matches = [
            path for path in all_folders
            if path.name.casefold() == search_key.casefold()
            or path.name.casefold().startswith(search_key.casefold() + "-")
        ]
        if matches:
            return sorted(matches), search_key, ""

        # Не создаём дубль молча, если папка того же 26-F083 уже есть,
        # но третий фрагмент отличается от актуального журнала.
        other_same_project = [
            path for path in all_folders
            if path.name.casefold().startswith(refresh_project.casefold() + "-")
        ]
        if other_same_project:
            return [], search_key, (
                f"Не найдена папка по ключу '{search_key}', но существуют папки того же проекта: "
                + ", ".join(path.name for path in other_same_project)
            )
        return [], search_key, ""

    matches = [
        path for path in all_folders
        if path.name.casefold().startswith(refresh_project.casefold() + "-")
    ]
    return sorted(matches), "", ""


def workbook_search_key(folder: Path, task_info: TaskProjectInfo | None) -> str:
    if task_info is not None and task_info.search_key:
        return task_info.search_key
    return derive_project_from_sample_code(folder.name, FOLDER_SEARCH_PARTS)


def is_ignored_workbook_name(name: str) -> bool:
    lower = name.casefold()
    return lower.startswith("~$") or "__new__" in lower or "_backup_" in lower


def find_project_workbooks(folder: Path, search_key: str) -> list[Path]:
    """
    Ищет только непосредственно внутри папки проекта.
    В дочернюю папку 'архив' (и любые другие подпапки) не заходит.
    Название — лишь фильтр кандидатов; версия берётся из metadata.
    """
    if not folder.exists():
        return []

    if not search_key:
        search_key = derive_project_from_sample_code(folder.name, FOLDER_SEARCH_PARTS)

    pattern = re.compile(rf"^{re.escape(search_key)}(?:-.+)?_v.+\.xlsx$", re.IGNORECASE)
    result: list[Path] = []
    for path in folder.iterdir():
        if not path.is_file() or path.suffix.casefold() != ".xlsx":
            continue
        if is_ignored_workbook_name(path.name):
            continue
        if pattern.fullmatch(path.name):
            result.append(path)
    return sorted(result)


def excel_lock_present(workbook_path: Path) -> bool:
    return workbook_path.with_name("~$" + workbook_path.name).exists()


# =============================================================================
# Инспекция проекта
# =============================================================================

def get_version_status(actual_version: str, release: ReleaseInfo) -> str:
    if not actual_version:
        return "MISSING_VERSION"
    try:
        actual_major = version_major(actual_version)
    except ValueError:
        return "INVALID_VERSION"
    if actual_major != release.form_major:
        return "INCOMPATIBLE_MAJOR"

    cmp_result = compare_versions(actual_version, release.form_version)
    if cmp_result == 0:
        return "CURRENT"
    if cmp_result < 0:
        return "OUTDATED_COMPATIBLE"
    return "AHEAD_OF_RELEASE"


def read_study_states(props: dict[str, str]) -> tuple[dict[str, str], list[str], bool]:
    states: dict[str, str] = {}
    blocking: list[str] = []
    for study, property_name in STUDY_STATE_PROPERTIES.items():
        state = props.get(property_name, "unknown").strip().lower() or "unknown"
        states[study] = state
        if state not in SAFE_RECREATE_STATES:
            blocking.append(f"{study}={state}")
    return states, blocking, not blocking


def inspect_project(
    refresh_project: str,
    task_info: TaskProjectInfo | None,
    release: ReleaseInfo,
) -> ProjectInspection:
    inspection = ProjectInspection(
        refresh_project=refresh_project,
        expected_folder=task_info.new_folder_name if task_info is not None else "",
        current_form_version=release.form_version,
    )

    folders, search_key, folder_issue = find_project_folders(refresh_project, task_info)
    inspection.folder_search_key = search_key
    inspection.found_folders = [str(path) for path in folders]

    if folder_issue:
        inspection.issue = folder_issue
        inspection.recommended_action = "BLOCKED_FOLDER_MISMATCH"
        return inspection
    if len(folders) > 1:
        inspection.issue = "Найдено несколько папок проекта"
        inspection.recommended_action = "BLOCKED_MULTIPLE_FOLDERS"
        return inspection
    if not folders:
        if task_info is None or not task_info.new_folder_name:
            inspection.issue = (
                "Папка проекта не найдена, а в журнале нет данных для создания имени новой папки"
            )
            inspection.recommended_action = "BLOCKED_NO_FOLDER_INFO"
        else:
            inspection.recommended_action = "CREATE"
        return inspection

    folder = folders[0]
    inspection.selected_folder = str(folder)
    search_key = workbook_search_key(folder, task_info)
    if not inspection.folder_search_key:
        inspection.folder_search_key = search_key

    workbooks = find_project_workbooks(folder, search_key)
    inspection.workbook_candidates = [str(path) for path in workbooks]

    if len(workbooks) > 1:
        inspection.issue = "Найдено несколько рабочих книг проекта"
        inspection.recommended_action = "BLOCKED_MULTIPLE_WORKBOOKS"
        return inspection
    if not workbooks:
        inspection.recommended_action = "CREATE"
        return inspection

    workbook = workbooks[0]
    inspection.selected_workbook = str(workbook)
    inspection.excel_lock_present = excel_lock_present(workbook)
    if inspection.excel_lock_present:
        inspection.issue = "Обнаружен Excel lock-файл; книга, вероятно, открыта"
        inspection.recommended_action = "BLOCKED_WORKBOOK_OPEN"
        return inspection

    try:
        props = read_office_custom_properties(workbook)
    except Exception as exc:
        inspection.issue = f"Не удалось прочитать metadata: {exc}"
        inspection.recommended_action = "BLOCKED_METADATA_ERROR"
        return inspection

    actual_version = props.get(FORM_VERSION_PROPERTY, "").strip()
    inspection.workbook_version = actual_version
    inspection.version_status = get_version_status(actual_version, release)

    if inspection.version_status in {"MISSING_VERSION", "INVALID_VERSION", "INCOMPATIBLE_MAJOR"}:
        inspection.recommended_action = "MIGRATION_REQUIRED"
        return inspection
    if inspection.version_status == "AHEAD_OF_RELEASE":
        inspection.issue = (
            "Версия книги выше текущей версии из project_version.json; автоматический downgrade запрещён"
        )
        inspection.recommended_action = "BLOCKED_VERSION_AHEAD"
        return inspection

    states, blocking, safe = read_study_states(props)
    inspection.study_states = states
    inspection.blocking_studies = blocking
    inspection.safe_to_recreate = safe

    if inspection.version_status == "OUTDATED_COMPATIBLE":
        inspection.recommended_action = "RECREATE_TO_CURRENT" if safe else "SAVE_REQUIRED_BEFORE_RECREATE"
    else:
        inspection.recommended_action = "OK" if safe else "SAVE_REQUIRED"
    return inspection


# =============================================================================
# Scan report
# =============================================================================

PROJECT_REPORT_COLUMNS = (
    "Project", "FolderSearchKey", "ExpectedFolder", "FoundFolders", "SelectedFolder",
    "WorkbookCandidates", "SelectedWorkbook", "WorkbookVersion", "CurrentFormVersion",
    "VersionStatus", "ExcelLock", "SafeToRecreate", "BlockingStudies",
    "RecommendedAction", "Issue",
)


def autosize_sheet(ws, *, max_width: int = 70) -> None:
    for column_cells in ws.columns:
        width = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, max((len(line) for line in value.splitlines()), default=0))
        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def add_excel_table(ws, name: str) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_scan_report(inspections: list[ProjectInspection], report_path: Path, release: ReleaseInfo) -> Path:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"
    ws.append(PROJECT_REPORT_COLUMNS)

    for item in inspections:
        ws.append((
            item.refresh_project,
            item.folder_search_key,
            item.expected_folder,
            "\n".join(item.found_folders),
            item.selected_folder,
            "\n".join(item.workbook_candidates),
            item.selected_workbook,
            item.workbook_version,
            release.form_version,
            item.version_status,
            "YES" if item.excel_lock_present else "NO",
            "YES" if item.safe_to_recreate else "NO",
            ", ".join(item.blocking_studies),
            item.recommended_action,
            item.issue,
        ))

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"
    add_excel_table(ws, "ProjectScan")
    autosize_sheet(ws)

    states_ws = wb.create_sheet("StudyStates")
    states_ws.append(("Project", "Study", "DataState"))
    for item in inspections:
        for study in STUDY_STATE_PROPERTIES:
            states_ws.append((item.refresh_project, study, item.study_states.get(study, "")))
    for cell in states_ws[1]:
        cell.font = Font(bold=True)
    states_ws.freeze_panes = "A2"
    add_excel_table(states_ws, "StudyStates")
    autosize_sheet(states_ws, max_width=35)

    info_ws = wb.create_sheet("Release")
    info_ws.append(("Parameter", "Value"))
    info_ws.append(("GeneratedAt", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    info_ws.append(("FormVersion", release.form_version))
    info_ws.append(("AddinVersion", release.addin_version))
    info_ws.append(("Template", release.template_path))
    info_ws.append(("BaseProjectsDir", str(BASE_PROJECTS_DIR)))
    for cell in info_ws[1]:
        cell.font = Font(bold=True)
    add_excel_table(info_ws, "ReleaseInfo")
    autosize_sheet(info_ws)

    wb.save(report_path)
    wb.close()
    return report_path


def project_ids_for_scan(
    requested_projects: list[str],
    year: int,
    task_infos: dict[str, TaskProjectInfo],
) -> list[str]:
    if requested_projects:
        return requested_projects
    prefix = f"{year % 100:02d}-F"
    from_tasks = {project for project in task_infos if project.startswith(prefix)}
    from_filesystem = list_project_ids_from_year_folder(year)
    return sorted(from_tasks | from_filesystem)


def run_scan(
    *,
    release: ReleaseInfo,
    projects_text: str,
    year: int,
    report_path: Path | None,
) -> RunReport:
    progress("SCAN: Excel запускаться не будет")
    task_rows = read_external_task_rows(TASKS_WORKBOOK, TASKS_PARENT_SHEET, TASKS_PARENT_TABLE)
    task_infos = build_task_project_infos(task_rows)
    requested = parse_projects_input(projects_text)
    projects = project_ids_for_scan(requested, year, task_infos)
    progress(f"SCAN: проектов к проверке: {len(projects)}")

    inspections: list[ProjectInspection] = []
    for index, project in enumerate(projects, start=1):
        if index == 1 or index % 50 == 0 or index == len(projects):
            progress(f"SCAN: {index}/{len(projects)}")
        inspections.append(inspect_project(project, task_infos.get(project), release))

    if report_path is None:
        report_path = PROJECT_DIR / f"project_scan_{year}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    written = write_scan_report(inspections, report_path, release)

    attention = sum(
        1 for item in inspections
        if item.recommended_action not in {"OK"}
    )
    progress(f"SCAN завершён. Отчёт: {written}")
    return RunReport(
        ok=True,
        mode="scan",
        message=f"Сканирование завершено. Проектов: {len(inspections)}, требуют внимания: {attention}",
        form_version=release.form_version,
        addin_version=release.addin_version,
        processed=len(inspections),
        succeeded=len(inspections),
        failed=0,
        scan_report=str(written),
    )

# =============================================================================
# DB
# =============================================================================

def get_projects_with_db_results() -> set[str]:
    """
    Бизнес-инвариант: любой начатый проект обязательно имеет OP или OPOH.
    Поэтому наличие сохранённых результатов определяется по OP_results/OPOH_results.
    """
    if not DB_PATH.exists():
        raise FileNotFoundError(f"База данных не найдена: {DB_PATH}")

    result: set[str] = set()
    with sqlite3.connect(DB_PATH) as conn:
        rows = conn.execute(
            """
            SELECT sampleCode
            FROM OP_results
            WHERE sampleCode IS NOT NULL AND TRIM(sampleCode) <> ''
            UNION
            SELECT sampleCode
            FROM OPOH_results
            WHERE sampleCode IS NOT NULL AND TRIM(sampleCode) <> ''
            """
        ).fetchall()

    for (sample_code,) in rows:
        project = derive_project_from_sample_code(sample_code, REFRESH_PROJECT_PARTS)
        try:
            project = normalize_project_number(project)
        except ValueError:
            continue
        result.add(project)
    return result


# =============================================================================
# Excel / xlwings
# =============================================================================

def disable_excel_prompts(app: xw.App) -> None:
    app.display_alerts = False
    app.api.DisplayAlerts = False
    app.api.AskToUpdateLinks = False
    app.api.AlertBeforeOverwriting = False
    app.api.EnableEvents = False


def unprotect_all_sheets(book: xw.Book) -> None:
    for sheet in book.sheets:
        for password in UNPROTECT_PASSWORDS:
            try:
                sheet.api.Unprotect(Password=password)
                break
            except Exception:
                pass


def protect_all_sheets(book: xw.Book) -> None:
    for sheet in book.sheets:
        try:
            sheet.api.Protect(
                Password=PROTECT_PASSWORD,
                DrawingObjects=True,
                Contents=True,
                Scenarios=True,
                UserInterfaceOnly=True,
                AllowFiltering=True,
            )
        except Exception as exc:
            logger.warning("Не удалось защитить лист %s: %s", sheet.name, exc)


def load_installed_addin(app: xw.App, configured_addin_path: Path, expected_version: str):
    """
    xlam НЕ открывается отдельной книгой. Ищем установленную AddIn по имени и
    включаем Installed=True. Это сохраняет рабочую схему без дублей mapped/UNC путей.
    """
    addin_name = configured_addin_path.name.casefold()

    for index in range(1, app.api.AddIns.Count + 1):
        addin = app.api.AddIns.Item(index)
        if str(addin.Name).casefold() != addin_name:
            continue

        progress(f"AddIn Name: {addin.Name}")
        progress(f"AddIn FullName: {addin.FullName}")
        progress(f"AddIn Installed до: {addin.Installed}")

        if addin.Installed:
            addin.Installed = False
        addin.Installed = True

        progress(f"AddIn Installed после: {addin.Installed}")

        actual_path = Path(str(addin.FullName))
        if not actual_path.exists():
            raise RuntimeError(f"Excel сообщает недоступный путь надстройки: {actual_path}")

        props = read_office_custom_properties(actual_path)
        actual_version = props.get(ADDIN_VERSION_PROPERTY, "").strip()
        if actual_version != expected_version:
            raise RuntimeError(
                "Версия реально подключённой Excel-надстройки не совпадает с project_version.json. "
                f"Ожидается {expected_version}, найдено {actual_version or '<нет>'}. Файл: {actual_path}"
            )
        return addin

    raise RuntimeError(
        f"Установленная надстройка '{configured_addin_path.name}' не найдена в Application.AddIns"
    )


def run_addin_macro(
    *,
    app: xw.App,
    workbook: xw.Book,
    addin_name: str,
    macro_name: str,
    args: tuple[Any, ...] = (),
) -> Any:
    disable_excel_prompts(app)
    workbook.activate()
    macro_ref = f"'{addin_name}'!{macro_name}"
    result = app.api.Run(macro_ref, *args) if args else app.api.Run(macro_ref)
    disable_excel_prompts(app)
    return result


def set_project_cell(book: xw.Book, refresh_project: str) -> None:
    book.sheets[PROJECT_CELL_SHEET].range(PROJECT_CELL_ADDRESS).value = refresh_project


def run_refresh_tasks_for_book(*, workbook_path: Path, refresh_project: str) -> None:
    from refresh_tasks import refresh_tasks

    progress(f"Обновляю задания для {refresh_project}")
    refresh_tasks(
        workbook=str(workbook_path),
        db_path=str(DB_PATH),
        source_workbook=str(TASKS_WORKBOOK),
        project_number=refresh_project,
        source_parent_sheet=REFRESH_TASKS_SOURCE_PARENT_SHEET,
        source_parent_table=REFRESH_TASKS_SOURCE_PARENT_TABLE,
        source_child_sheet=REFRESH_TASKS_SOURCE_CHILD_SHEET,
        source_child_table=REFRESH_TASKS_SOURCE_CHILD_TABLE,
        target_parent_sheet=REFRESH_TASKS_TARGET_PARENT_SHEET,
        target_parent_table=REFRESH_TASKS_TARGET_PARENT_TABLE,
        target_child_sheet=REFRESH_TASKS_TARGET_CHILD_SHEET,
        target_child_table=REFRESH_TASKS_TARGET_CHILD_TABLE,
    )


def run_load_all_studies_for_book(*, workbook_path: Path, refresh_project: str) -> None:
    """SQLite -> staging напрямую через sync.run_load, без запуска второго python.exe."""
    from sync import run_load

    progress(f"Загружаю из БД все исследования для {refresh_project}")
    report = run_load(
        workbook=str(workbook_path),
        db_path=str(DB_PATH),
        project_number=refresh_project,
        study_codes=None,
    )
    if not report.ok:
        raise RuntimeError(report.message)


def load_forms_from_tables(*, app: xw.App, workbook: xw.Book, addin_name: str) -> None:
    progress("Раскладываю *_results/*_sourceData по формам")
    run_addin_macro(
        app=app,
        workbook=workbook,
        addin_name=addin_name,
        macro_name=LOAD_FORMS_FROM_TABLES_MACRO,
    )


def set_all_study_baselines(*, app: xw.App, workbook: xw.Book, addin_name: str) -> None:
    progress("VBA рассчитывает DataHash/DataState для всех исследований")
    run_addin_macro(
        app=app,
        workbook=workbook,
        addin_name=addin_name,
        macro_name=SET_ALL_BASELINES_MACRO,
        args=(workbook.api,),
    )


def run_after_refresh_macro(*, app: xw.App, workbook: xw.Book, addin_name: str) -> None:
    progress(f"Запускаю макрос: {AFTER_REFRESH_MACRO}")
    run_addin_macro(
        app=app,
        workbook=workbook,
        addin_name=addin_name,
        macro_name=AFTER_REFRESH_MACRO,
    )


# =============================================================================
# Файловые операции CREATE / RECREATE
# =============================================================================

def unique_archive_path(archive_dir: Path, old_workbook: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = archive_dir / f"{old_workbook.stem}_backup_{timestamp}{old_workbook.suffix}"
    counter = 1
    while candidate.exists():
        candidate = archive_dir / (
            f"{old_workbook.stem}_backup_{timestamp}_{counter}{old_workbook.suffix}"
        )
        counter += 1
    return candidate


def temporary_workbook_path(folder: Path, final_name: str) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    return folder / f"{Path(final_name).stem}__new__{timestamp}.xlsx"


def build_new_workbook(
    *,
    app: xw.App,
    addin_name: str,
    template_path: Path,
    temp_path: Path,
    refresh_project: str,
    restore_from_db: bool,
) -> None:
    workbook: xw.Book | None = None
    shutil.copy2(template_path, temp_path)

    try:
        disable_excel_prompts(app)
        workbook = app.books.open(
            str(temp_path),
            update_links=True,
            ignore_read_only_recommended=True,
        )
        disable_excel_prompts(app)
        unprotect_all_sheets(workbook)

        progress(
            f"Записываю номер проекта в {PROJECT_CELL_SHEET}!{PROJECT_CELL_ADDRESS}: {refresh_project}"
        )
        set_project_cell(workbook, refresh_project)
        workbook.save()

        if restore_from_db:
            run_load_all_studies_for_book(
                workbook_path=temp_path,
                refresh_project=refresh_project,
            )
            load_forms_from_tables(
                app=app,
                workbook=workbook,
                addin_name=addin_name,
            )

        run_refresh_tasks_for_book(
            workbook_path=temp_path,
            refresh_project=refresh_project,
        )
        run_after_refresh_macro(
            app=app,
            workbook=workbook,
            addin_name=addin_name,
        )

        # Hash/status считаются только VBA и только по финальному состоянию формы.
        set_all_study_baselines(
            app=app,
            workbook=workbook,
            addin_name=addin_name,
        )

        protect_all_sheets(workbook)
        progress("Сохраняю новую форму")
        disable_excel_prompts(app)
        workbook.save()

    finally:
        if workbook is not None:
            try:
                disable_excel_prompts(app)
                workbook.close()
            except Exception:
                pass


def update_existing_workbook(
    *,
    app: xw.App,
    addin_name: str,
    workbook_path: Path,
    refresh_project: str,
) -> None:
    workbook: xw.Book | None = None
    try:
        disable_excel_prompts(app)
        workbook = app.books.open(
            str(workbook_path),
            update_links=True,
            ignore_read_only_recommended=True,
        )
        disable_excel_prompts(app)
        unprotect_all_sheets(workbook)

        run_refresh_tasks_for_book(
            workbook_path=workbook_path,
            refresh_project=refresh_project,
        )
        run_after_refresh_macro(
            app=app,
            workbook=workbook,
            addin_name=addin_name,
        )

        protect_all_sheets(workbook)
        progress("Сохраняю существующую форму")
        disable_excel_prompts(app)
        workbook.save()
    finally:
        if workbook is not None:
            try:
                disable_excel_prompts(app)
                workbook.close()
            except Exception:
                pass


def validate_built_workbook(path: Path, release: ReleaseInfo) -> None:
    """Финальный контроль новой книги до подмены рабочей версии."""
    props = read_office_custom_properties(path)
    actual_version = props.get(FORM_VERSION_PROPERTY, "").strip()
    if actual_version != release.form_version:
        raise RuntimeError(
            f"Созданная книга имеет неверную PVTFormVersion: "
            f"{actual_version or '<нет>'}; ожидается {release.form_version}"
        )

    _states, blocking, safe = read_study_states(props)
    if not safe:
        raise RuntimeError(
            "После создания/загрузки VBA не сформировал безопасный baseline: "
            + ", ".join(blocking)
        )


def finalize_new_file(*, temp_path: Path, final_path: Path) -> None:
    if final_path.exists():
        raise FileExistsError(f"Финальный файл уже существует: {final_path}")
    temp_path.replace(final_path)


def finalize_recreated_file(
    *,
    temp_path: Path,
    old_workbook: Path,
    final_path: Path,
) -> Path:
    archive_dir = old_workbook.parent / ARCHIVE_DIR_NAME
    archive_dir.mkdir(parents=True, exist_ok=True)
    backup_path = unique_archive_path(archive_dir, old_workbook)

    progress(f"Переношу старую форму в архив: {backup_path.name}")
    old_workbook.replace(backup_path)

    try:
        if final_path.exists():
            raise FileExistsError(
                f"После архивации старого файла финальное имя всё ещё занято: {final_path}"
            )
        temp_path.replace(final_path)
    except Exception:
        try:
            if backup_path.exists() and not old_workbook.exists():
                backup_path.replace(old_workbook)
        except Exception:
            logger.exception("Не удалось откатить старую книгу из архива: %s", backup_path)
        raise

    return backup_path

# =============================================================================
# Классификация и исполнение update/recreate
# =============================================================================

def resolve_or_create_folder(
    *,
    refresh_project: str,
    task_info: TaskProjectInfo | None,
    allow_create: bool,
) -> Path:
    folders, _search_key, issue = find_project_folders(refresh_project, task_info)

    if issue:
        raise RuntimeError(issue)
    if len(folders) > 1:
        raise RuntimeError(
            "Найдено несколько папок проекта: " + ", ".join(str(path) for path in folders)
        )
    if folders:
        return folders[0]
    if not allow_create:
        raise FileNotFoundError(f"Папка проекта {refresh_project} не найдена")
    if task_info is None or not task_info.new_folder_name:
        raise RuntimeError(
            f"Нельзя создать папку {refresh_project}: в журнале нет шифра для первых 4 фрагментов"
        )

    folder = year_dir_for_project(refresh_project) / task_info.new_folder_name
    progress(f"Создаю папку проекта: {folder}")
    folder.mkdir(parents=True, exist_ok=False)
    return folder


def validate_existing_workbook_for_automation(
    *,
    inspection: ProjectInspection,
    mode: str,
) -> None:
    if not inspection.selected_workbook:
        return
    if inspection.excel_lock_present:
        raise RuntimeError(f"Книга, вероятно, открыта в Excel: {inspection.selected_workbook}")
    if inspection.version_status in {"MISSING_VERSION", "INVALID_VERSION", "INCOMPATIBLE_MAJOR"}:
        raise RuntimeError(
            f"Книга несовместима с текущим поколением формы: "
            f"{inspection.workbook_version or '<нет версии>'}"
        )
    if inspection.version_status == "AHEAD_OF_RELEASE":
        raise RuntimeError(
            "Версия книги выше версии текущего релиза; автоматическая обработка запрещена"
        )
    if mode == "recreate" and not inspection.safe_to_recreate:
        raise RuntimeError(
            "Пересоздание запрещено: есть несохранённые/неопределённые исследования: "
            + ", ".join(inspection.blocking_studies)
        )


def create_or_update_one_project(
    *,
    app: xw.App,
    addin_name: str,
    mode: str,
    refresh_project: str,
    task_info: TaskProjectInfo | None,
    release: ReleaseInfo,
    projects_with_db_results: set[str],
) -> ProjectRunResult:
    inspection = inspect_project(refresh_project, task_info, release)

    try:
        validate_existing_workbook_for_automation(inspection=inspection, mode=mode)

        if inspection.recommended_action in {
            "BLOCKED_FOLDER_MISMATCH",
            "BLOCKED_MULTIPLE_FOLDERS",
            "BLOCKED_MULTIPLE_WORKBOOKS",
            "BLOCKED_NO_FOLDER_INFO",
            "BLOCKED_METADATA_ERROR",
            "BLOCKED_WORKBOOK_OPEN",
            "BLOCKED_VERSION_AHEAD",
            "MIGRATION_REQUIRED",
        }:
            raise RuntimeError(inspection.issue or inspection.recommended_action)

        folder = resolve_or_create_folder(
            refresh_project=refresh_project,
            task_info=task_info,
            allow_create=True,
        )

        search_key = workbook_search_key(folder, task_info)
        workbooks = find_project_workbooks(folder, search_key)
        if len(workbooks) > 1:
            raise RuntimeError(
                "Найдено несколько рабочих книг проекта: "
                + ", ".join(path.name for path in workbooks)
            )
        existing_workbook = workbooks[0] if workbooks else None

        if existing_workbook is not None:
            # Повторная проверка непосредственно перед изменением.
            inspection = inspect_project(refresh_project, task_info, release)
            validate_existing_workbook_for_automation(inspection=inspection, mode=mode)

        has_db_results = refresh_project in projects_with_db_results

        if mode == "update" and existing_workbook is not None:
            progress(f"UPDATE: {existing_workbook.name}")
            update_existing_workbook(
                app=app,
                addin_name=addin_name,
                workbook_path=existing_workbook,
                refresh_project=refresh_project,
            )
            return ProjectRunResult(
                project=refresh_project,
                ok=True,
                action="UPDATE",
                message="Существующая форма обновлена",
                folder=str(folder),
                workbook=str(existing_workbook),
            )

        # Если книги нет — и update, и recreate выполняют CREATE.
        # При наличии OP/OPOH в БД данные восстанавливаются.
        if existing_workbook is None:
            final_path = folder / f"{folder.name}_v{release.form_version}.xlsx"
            temp_path = temporary_workbook_path(folder, final_path.name)
            progress(f"CREATE: {refresh_project}; restore_from_db={has_db_results}")

            try:
                build_new_workbook(
                    app=app,
                    addin_name=addin_name,
                    template_path=Path(release.template_path),
                    temp_path=temp_path,
                    refresh_project=refresh_project,
                    restore_from_db=has_db_results,
                )
                validate_built_workbook(temp_path, release)
                finalize_new_file(temp_path=temp_path, final_path=final_path)
            except Exception:
                if temp_path.exists():
                    try:
                        temp_path.unlink()
                    except Exception:
                        logger.exception("Не удалось удалить временный файл %s", temp_path)
                raise

            return ProjectRunResult(
                project=refresh_project,
                ok=True,
                action="CREATE",
                message=(
                    "Новая форма создана и восстановлена из БД"
                    if has_db_results else "Новая форма создана"
                ),
                folder=str(folder),
                workbook=str(final_path),
            )

        if mode != "recreate":
            raise RuntimeError(f"Неожиданная ветка режима: {mode}")

        if any(state == "synced" for state in inspection.study_states.values()) and not has_db_results:
            raise RuntimeError(
                "В metadata есть synced-исследования, но в БД не найден OP/OPOH проекта"
            )

        final_path = folder / f"{folder.name}_v{release.form_version}.xlsx"
        temp_path = temporary_workbook_path(folder, final_path.name)
        progress(f"RECREATE: {existing_workbook.name} -> {final_path.name}")

        try:
            build_new_workbook(
                app=app,
                addin_name=addin_name,
                template_path=Path(release.template_path),
                temp_path=temp_path,
                refresh_project=refresh_project,
                restore_from_db=has_db_results,
            )
            validate_built_workbook(temp_path, release)
            backup_path = finalize_recreated_file(
                temp_path=temp_path,
                old_workbook=existing_workbook,
                final_path=final_path,
            )
        except Exception:
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    logger.exception("Не удалось удалить временный файл %s", temp_path)
            raise

        return ProjectRunResult(
            project=refresh_project,
            ok=True,
            action="RECREATE",
            message="Форма пересоздана; старая версия перенесена в архив",
            folder=str(folder),
            workbook=str(final_path),
            backup=str(backup_path),
        )

    except Exception as exc:
        logger.exception("Ошибка обработки проекта %s", refresh_project)
        return ProjectRunResult(
            project=refresh_project,
            ok=False,
            action="BLOCKED" if isinstance(exc, RuntimeError) else "ERROR",
            message=str(exc),
            folder=inspection.selected_folder or None,
            workbook=inspection.selected_workbook or None,
            error_type=type(exc).__name__,
        )


def projects_for_execution(
    *,
    mode: str,
    projects_text: str,
    task_infos: dict[str, TaskProjectInfo],
) -> list[str]:
    requested = parse_projects_input(projects_text)
    if requested:
        return requested
    if mode == "update":
        return latest_task_projects(task_infos, RECENT_PROJECT_LIMIT)
    if mode == "recreate":
        raise ValueError("Для режима recreate проекты должны быть указаны явно")
    raise ValueError(f"Неожиданный режим: {mode}")


def run_mutating_mode(
    *,
    release: ReleaseInfo,
    mode: str,
    projects_text: str,
) -> RunReport:
    if mode not in {"update", "recreate"}:
        raise ValueError(mode)

    task_rows = read_external_task_rows(TASKS_WORKBOOK, TASKS_PARENT_SHEET, TASKS_PARENT_TABLE)
    task_infos = build_task_project_infos(task_rows)
    projects = projects_for_execution(mode=mode, projects_text=projects_text, task_infos=task_infos)
    progress(f"{mode.upper()}: проектов к обработке: {len(projects)}")

    # БД не используется для поиска папок/книг; только для решения, нужно ли
    # восстанавливать результаты в создаваемую/пересоздаваемую форму.
    projects_with_db_results = get_projects_with_db_results()

    app = xw.App(visible=True, add_book=False)
    results: list[ProjectRunResult] = []

    try:
        # Для инициализации xlam события должны быть включены.
        app.display_alerts = False
        app.api.DisplayAlerts = False
        app.api.AskToUpdateLinks = False
        app.api.AlertBeforeOverwriting = False
        app.api.EnableEvents = True
        app.screen_updating = False

        progress("Подключаю установленную Excel-надстройку")
        addin = load_installed_addin(app, Path(release.addin_path), release.addin_version)
        addin_name = str(addin.Name)

        # Проектные книги открываются уже без событий и диалогов.
        disable_excel_prompts(app)

        for index, project in enumerate(projects, start=1):
            progress("-" * 80)
            progress(f"{mode.upper()}: {index}/{len(projects)} {project}")

            result = create_or_update_one_project(
                app=app,
                addin_name=addin_name,
                mode=mode,
                refresh_project=project,
                task_info=task_infos.get(project),
                release=release,
                projects_with_db_results=projects_with_db_results,
            )
            results.append(result)
            progress(
                f"{project}: {result.action} "
                f"{'OK' if result.ok else ''} — {result.message}"
            )
    finally:
        try:
            app.quit()
        except Exception:
            pass

    succeeded = sum(1 for item in results if item.ok)
    failed = len(results) - succeeded
    return RunReport(
        ok=failed == 0,
        mode=mode,
        message=(
            f"{mode.upper()} завершён. Успешно: {succeeded}, "
            f"ошибок/блокировок: {failed}"
        ),
        form_version=release.form_version,
        addin_version=release.addin_version,
        processed=len(results),
        succeeded=succeeded,
        failed=failed,
        projects=results,
    )


# =============================================================================
# UI
# =============================================================================

def ask_ui_options(*, default_year: int) -> UiOptions:
    try:
        import tkinter as tk
        from tkinter import messagebox

        result = UiOptions(mode="scan", projects_text="", year=default_year, cancelled=True)
        root = tk.Tk()
        root.title("Формы проектов ГТИ")
        root.resizable(False, False)

        frame = tk.Frame(root, padx=14, pady=14)
        frame.pack(fill="both", expand=True)

        tk.Label(
            frame,
            text=(
                "Проекты: 26-F001 или 26-F001...26-F015\n"
                "Можно перечислять через пробел, запятую или ;"
            ),
            justify="left",
            anchor="w",
        ).pack(fill="x")

        projects_var = tk.StringVar(value="")
        entry = tk.Entry(frame, textvariable=projects_var, width=54)
        entry.pack(fill="x", pady=(8, 12))

        mode_frame = tk.LabelFrame(frame, text="Режим", padx=10, pady=8)
        mode_frame.pack(fill="x", pady=(0, 12))
        mode_var = tk.StringVar(value="scan")

        tk.Radiobutton(
            mode_frame,
            text="Сканирование — ничего не изменять, создать Excel-отчёт",
            variable=mode_var,
            value="scan",
        ).pack(anchor="w")
        tk.Radiobutton(
            mode_frame,
            text=(
                f"Создание / обновление — пустой список = последние "
                f"{RECENT_PROJECT_LIMIT} проектов"
            ),
            variable=mode_var,
            value="update",
        ).pack(anchor="w")
        tk.Radiobutton(
            mode_frame,
            text="Пересоздание — проекты нужно указать явно",
            variable=mode_var,
            value="recreate",
        ).pack(anchor="w")

        year_frame = tk.Frame(frame)
        year_frame.pack(fill="x", pady=(0, 12))
        tk.Label(year_frame, text="Год для scan без списка проектов:").pack(side="left")
        year_var = tk.StringVar(value=str(default_year))
        tk.Entry(year_frame, textvariable=year_var, width=8).pack(side="left", padx=(8, 0))

        tk.Label(
            frame,
            justify="left",
            anchor="w",
            text=(
                "SCAN без списка: весь выбранный год, Excel не запускается.\n"
                f"UPDATE без списка: последние {RECENT_PROJECT_LIMIT} проектов из журнала.\n"
                "RECREATE: только явно указанные проекты; dirty/unsaved/error/unknown "
                "блокируют пересоздание."
            ),
        ).pack(fill="x", pady=(0, 12))

        buttons = tk.Frame(frame)
        buttons.pack(fill="x")

        def submit() -> None:
            mode = mode_var.get().strip().lower()
            projects_text = projects_var.get().strip()
            try:
                year = int(year_var.get().strip())
            except ValueError:
                messagebox.showerror("Ошибка", "Год должен быть целым числом")
                return
            if mode == "recreate" and not projects_text:
                messagebox.showerror("Ошибка", "Для пересоздания укажите проекты явно")
                return
            result.mode = mode
            result.projects_text = projects_text
            result.year = year
            result.cancelled = False
            root.destroy()

        def cancel() -> None:
            result.cancelled = True
            root.destroy()

        tk.Button(buttons, text="Запустить", width=14, command=submit).pack(side="left")
        tk.Button(buttons, text="Отмена", width=14, command=cancel).pack(side="left", padx=(8, 0))
        root.bind("<Return>", lambda _event: submit())
        root.bind("<Escape>", lambda _event: cancel())
        root.protocol("WM_DELETE_WINDOW", cancel)

        root.update_idletasks()
        width, height = root.winfo_reqwidth(), root.winfo_reqheight()
        x = max((root.winfo_screenwidth() - width) // 2, 0)
        y = max((root.winfo_screenheight() - height) // 2, 0)
        root.geometry(f"{width}x{height}+{x}+{y}")
        entry.focus_set()
        root.mainloop()
        return result

    except Exception:
        print("Режимы: scan / update / recreate")
        mode = input("Режим: ").strip().lower()
        if mode not in MODES:
            return UiOptions(mode="", projects_text="", year=default_year, cancelled=True)
        projects_text = input("Проекты (можно оставить пустым, кроме recreate): ").strip()
        if mode == "recreate" and not projects_text:
            return UiOptions(mode="", projects_text="", year=default_year, cancelled=True)
        year = default_year
        if mode == "scan" and not projects_text:
            raw_year = input(f"Год [{default_year}]: ").strip()
            if raw_year:
                year = int(raw_year)
        return UiOptions(mode=mode, projects_text=projects_text, year=year, cancelled=False)


# =============================================================================
# CLI / main
# =============================================================================

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Сканирование, создание/обновление и безопасное пересоздание форм проектов ГТИ"
        )
    )
    parser.add_argument("--mode", choices=sorted(MODES), help="scan, update или recreate")
    parser.add_argument(
        "--projects",
        default="",
        help="Проект/список/диапазон, например 26-F001 или 26-F001...26-F015",
    )
    parser.add_argument(
        "--year",
        type=int,
        default=datetime.now().year,
        help="Год для scan без --projects. По умолчанию текущий год.",
    )
    parser.add_argument("--report", default="", help="Путь к xlsx-отчёту режима scan")
    return parser


def report_to_json(report: RunReport) -> str:
    return json.dumps(asdict(report), ensure_ascii=False, indent=2)


def execute(*, mode: str, projects_text: str, year: int, report_path: str) -> RunReport:
    if mode not in MODES:
        raise ValueError(f"Неизвестный режим: {mode}")

    # Общий preflight выполняется до любых изменений.
    release = preflight_release()

    if not BASE_PROJECTS_DIR.exists():
        raise FileNotFoundError(f"Корневая папка проектов не найдена: {BASE_PROJECTS_DIR}")

    if mode == "scan":
        return run_scan(
            release=release,
            projects_text=projects_text,
            year=year,
            report_path=Path(report_path) if report_path else None,
        )

    return run_mutating_mode(
        release=release,
        mode=mode,
        projects_text=projects_text,
    )


def main(argv: list[str] | None = None) -> int:
    args_list = list(sys.argv[1:] if argv is None else argv)

    # Без параметров — UI.
    if not args_list:
        try:
            preflight_release()
        except Exception as exc:
            logger.exception("Ошибка preflight")
            print(json.dumps({"ok": False, "message": f"{type(exc).__name__}: {exc}"}, ensure_ascii=False, indent=2))
            return 1

        options = ask_ui_options(default_year=datetime.now().year)
        if options.cancelled:
            progress("Операция отменена пользователем")
            return 0

        mode = options.mode
        projects_text = options.projects_text
        year = options.year
        report_path = ""

    else:
        parser = build_parser()
        args = parser.parse_args(args_list)
        if not args.mode:
            parser.error("--mode обязателен при запуске с параметрами")

        mode = args.mode
        projects_text = args.projects
        year = args.year
        report_path = args.report

        if mode == "recreate" and not projects_text.strip():
            parser.error("для режима recreate нужно явно указать --projects")

    try:
        report = execute(
            mode=mode,
            projects_text=projects_text,
            year=year,
            report_path=report_path,
        )
    except Exception as exc:
        logger.exception("Критическая ошибка")
        report = RunReport(
            ok=False,
            mode=mode,
            message=f"{type(exc).__name__}: {exc}. Подробности в логе: {LOG_PATH}",
            form_version="",
            addin_version="",
            processed=0,
            succeeded=0,
            failed=1,
        )

    print(report_to_json(report))
    return 0 if report.ok else 1


if __name__ == "__main__":
    sys.exit(main())
